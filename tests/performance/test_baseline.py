import os
import json
import threading
import time
from datetime import datetime, timezone
from pathlib import Path

import pytest
import psutil
import httpx
from openpyxl import Workbook

from excel_auditor.engine import compare_workbook
from excel_auditor.models import AuditReport, RuleSet
from excel_auditor.reporting import write_differences_jsonl, write_json_report
from excel_auditor.models import StandardSourceConfig
from excel_auditor.rendering import ExcelRenderer
from excel_auditor.service import AuditService
from excel_auditor.pandera_adapter import StandardDataValidator
from excel_auditor.snapshots import SpilledRecords, create_snapshot
from excel_auditor.standard_files import load_standard_file
from excel_auditor.standard_sources import ConnectionRegistry, ManagedHttpSource
from excel_auditor.workbook import inspect_workbook
from excel_auditor.workbook import SheetSnapshot, WorkbookSnapshot
from excel_auditor.product_workflow import (
    CatalogFieldDefinition,
    CatalogFieldSource,
    CategoryDefinition,
    InMemoryCatalogAdapter,
    normalize_product_workbook,
)


def _python_cpu_calibration_seconds() -> float:
    """Measure the current runner's single-process Python throughput."""
    value = 0
    started = time.process_time()
    for index in range(10_000_000):
        value = (value * 1_664_525 + index + 1_013_904_223) & 0xFFFFFFFF
    elapsed = time.process_time() - started
    assert value == 4_094_895_424 and elapsed > 0
    return elapsed


@pytest.mark.performance
def test_product_normalization_baseline():
    if os.environ.get("RUN_PRODUCT_PERFORMANCE") != "1":
        pytest.skip("set RUN_PRODUCT_PERFORMANCE=1 to execute product normalization baseline")
    record_count = int(os.environ.get("PRODUCT_PERF_ROWS", "10000"))
    attribute_count = int(os.environ.get("PRODUCT_PERF_ATTRIBUTES", "20"))
    assert record_count >= 1 and attribute_count >= 1
    platform_fields = [
        CatalogFieldDefinition(
            field_id=f"attribute_{index}",
            title=f"平台属性{index}",
            source=CatalogFieldSource.PLATFORM_ATTRIBUTE,
            category_id="phone",
            attribute_id=f"attribute_{index}",
            display_order=index,
        )
        for index in range(attribute_count)
    ]
    rules = RuleSet.model_validate({
        "schema_id": "product-performance",
        "schema_version": "1.0.0",
        "name": "Product performance",
        "workbook": {
            "max_rows_per_sheet": max(100_000, record_count),
            "max_in_memory_cells": int(os.environ.get("PRODUCT_PERF_MAX_IN_MEMORY_CELLS", "500000")),
        },
        "sheets": [{
            "id": "products",
            "name": "Products",
            "primary_key": ["product_id"],
            "columns": [
                {"name": "product_id", "title": "商品ID", "required": True},
                {"name": "platform_category_id", "title": "平台类目ID"},
                {"name": "merchant_category", "title": "商家类目", "required": True},
            ],
        }],
        "product_workflow": {
            "sheet_id": "products",
            "catalog_connection_id": "performance",
            "category": {
                "attributes": {"path_template": "/catalog/{category_id}/attributes"},
                "specifications": {"path_template": "/catalog/{category_id}/specifications"},
            },
        },
    })
    headers = ["商品ID", "平台类目ID", "商家类目", *[field.title for field in platform_fields], "商家备注"]
    rows = [(1, headers)] + [
        (
            index + 2,
            [f"P{index:06d}", "phone", "手机", *[f"V{index}-{field_index}" for field_index in range(attribute_count)], f"备注{index}"],
        )
        for index in range(record_count)
    ]
    snapshot = WorkbookSnapshot(
        Path("product-performance.xlsx"),
        "0" * 64,
        {"Products": SheetSnapshot("Products", record_count + 1, len(headers), rows)},
    )
    catalog = InMemoryCatalogAdapter(
        [CategoryDefinition(category_id="phone", name="手机")],
        {"phone": platform_fields},
    )
    process = psutil.Process()
    baseline_rss = process.memory_info().rss
    peak_rss = [baseline_rss]
    stopped = threading.Event()

    def sample_memory():
        while not stopped.wait(0.02):
            peak_rss[0] = max(peak_rss[0], process.memory_info().rss)

    sampler = threading.Thread(target=sample_memory, daemon=True)
    sampler.start()
    cpu_calibration_seconds = _python_cpu_calibration_seconds()
    started = time.perf_counter()
    cpu_started = time.process_time()
    result = normalize_product_workbook(snapshot, rules, catalog)
    serialized_row_count = len(result.model_dump(mode="json")["category_sheets"][0]["rows"])
    elapsed = time.perf_counter() - started
    cpu_seconds = time.process_time() - cpu_started
    stopped.set()
    sampler.join()
    rss_delta_mib = (peak_rss[0] - baseline_rss) / 1024 / 1024
    metrics = {
        "benchmark_version": 2,
        "rows": record_count,
        "platform_attributes": attribute_count,
        "cpu_calibration_seconds": round(cpu_calibration_seconds, 6),
        "cpu_seconds": round(cpu_seconds, 3),
        "normalized_cpu_units": round(cpu_seconds / cpu_calibration_seconds, 3),
        "elapsed_seconds": round(elapsed, 3),
        "peak_rss_delta_mib": round(rss_delta_mib, 2),
        "issues": len(result.issues),
        "reviews": len(result.review_items),
    }
    print(metrics)
    if output := os.environ.get("PRODUCT_PERF_RESULT_PATH"):
        Path(output).write_text(json.dumps(metrics, indent=2), encoding="utf-8")
    assert len(result.category_sheets) == 1
    assert len(result.category_sheets[0].rows) == record_count
    assert serialized_row_count == record_count
    assert result.category_sheets[0].rows[-1][f"attribute_{attribute_count - 1}"]
    assert not result.requires_manual_review and not result.issues
    assert elapsed <= float(os.environ.get("PRODUCT_PERF_MAX_SECONDS", "60"))
    assert rss_delta_mib <= float(os.environ.get("PRODUCT_PERF_MAX_RSS_MIB", "512"))
    result.close()


@pytest.mark.performance
def test_configured_workbook_baseline(tmp_path):
    if os.environ.get("RUN_PERFORMANCE") != "1":
        pytest.skip("set RUN_PERFORMANCE=1 to execute performance baseline")
    rows = int(os.environ.get("PERF_ROWS", "10000"))
    fields = int(os.environ.get("PERF_COLUMNS", "20"))
    density = float(os.environ.get("PERF_DENSITY", "1"))
    difference_rate = float(os.environ.get("PERF_DIFFERENCE_RATE", "0"))
    sheet_count = int(os.environ.get("PERF_SHEETS", "1"))
    assert 0 <= density <= 1 and 0 <= difference_rate <= 1
    assert 1 <= sheet_count <= 50 and rows >= sheet_count
    columns = [{"name": "id", "title": "ID", "required": True}] + [{"name": f"f{i}", "title": f"F{i}"} for i in range(1, fields)]
    sheets = [
        {
            "id": "data" if sheet_count == 1 else f"data_{sheet_index + 1}",
            "name": "Data" if sheet_count == 1 else f"Data{sheet_index + 1}",
            "primary_key": ["id"],
            "columns": columns,
        }
        for sheet_index in range(sheet_count)
    ]
    rules = RuleSet.model_validate({"schema_id": "performance", "schema_version": "1.0.0", "name": "Performance", "sheets": sheets})
    path = tmp_path / "baseline.xlsx"
    book = Workbook(write_only=True)
    standard: dict[str, list[dict]] = {sheet_rule["id"]: [] for sheet_rule in sheets}
    populated_fields = int((fields - 1) * density)
    expected_differences = 0
    global_index = 0
    for sheet_index, sheet_rule in enumerate(sheets):
        worksheet = book.create_sheet(sheet_rule["name"])
        worksheet.append([column["title"] for column in columns])
        sheet_rows = rows // sheet_count + (1 if sheet_index < rows % sheet_count else 0)
        for _local_index in range(sheet_rows):
            record = {"id": f"R{global_index:06d}", **{f"f{i}": f"V{global_index}-{i}" for i in range(1, populated_fields + 1)}}
            standard[sheet_rule["id"]].append(record)
            excel_record = dict(record)
            if populated_fields and global_index < int(rows * difference_rate):
                excel_record["f1"] = f"DIFF-{global_index}"
                expected_differences += 1
            worksheet.append([excel_record.get(column["name"]) for column in columns])
            global_index += 1
    book.save(path)
    cpu_calibration_seconds = _python_cpu_calibration_seconds()
    process = psutil.Process()
    baseline_rss = process.memory_info().rss
    peak_rss = [baseline_rss]
    stopped = threading.Event()
    def sample_memory():
        while not stopped.wait(0.02):
            peak_rss[0] = max(peak_rss[0], process.memory_info().rss)
    sampler = threading.Thread(target=sample_memory, daemon=True)
    sampler.start()
    started = time.perf_counter()
    cpu_started = time.process_time()
    snapshot = inspect_workbook(path, rules)
    inspected = time.perf_counter()
    result = compare_workbook(snapshot, standard, rules)
    report_path = tmp_path / "report.json"
    write_json_report(AuditReport(job_id="job_performance", created_at=datetime.now(timezone.utc), schema_id=rules.schema_id, schema_version=rules.schema_version, schema_sha256=rules.content_sha256, input_sha256=snapshot.sha256, standard_snapshot_id="std_performance", standard_sha256="0" * 64, header_mappings=result.mappings, differences=result.differences, summary=result.summary), report_path)
    elapsed = time.perf_counter() - started
    cpu_seconds = time.process_time() - cpu_started
    stopped.set(); sampler.join()
    metrics = {"benchmark_version": 5, "rows": rows, "columns": fields, "sheets": sheet_count, "density": density, "difference_rate": difference_rate, "large_mode": snapshot.large_mode, "join_backends": sorted(set(result.join_backends or [])), "inspect_seconds": round(inspected - started, 3), "compare_and_report_seconds": round(elapsed - (inspected - started), 3), "cpu_calibration_seconds": round(cpu_calibration_seconds, 6), "cpu_seconds": round(cpu_seconds, 3), "normalized_cpu_units": round(cpu_seconds / cpu_calibration_seconds, 3), "elapsed_seconds": round(elapsed, 3), "peak_rss_delta_mib": round((peak_rss[0] - baseline_rss) / 1024 / 1024, 2), "report_bytes": report_path.stat().st_size}
    print(metrics)
    if output := os.environ.get("PERF_RESULT_PATH"):
        Path(output).write_text(json.dumps(metrics, indent=2), encoding="utf-8")
    assert result.summary.matched_records == rows
    assert result.summary.mismatched_cells == expected_differences
    assert elapsed <= float(os.environ.get("PERF_MAX_SECONDS", "900"))
    assert metrics["peak_rss_delta_mib"] <= float(os.environ.get("PERF_MAX_RSS_MIB", "6144"))


@pytest.mark.performance
def test_paginated_standard_source_baseline(tmp_path):
    if os.environ.get("PERF_RUN_HTTP") != "1":
        pytest.skip("set PERF_RUN_HTTP=1 to execute paginated HTTP baseline")
    record_count = int(os.environ.get("PERF_HTTP_RECORDS", "50000"))
    page_size = int(os.environ.get("PERF_HTTP_PAGE_SIZE", "500"))
    registry_path = tmp_path / "connections.json"
    registry_path.write_text(json.dumps({"connections": [{
        "id": "performance-http",
        "base_url": "https://standards.example.test/",
        "allowed_paths": ["/records"],
        "max_records": record_count,
        "max_response_bytes": 64 * 1024 * 1024,
    }]}), encoding="utf-8")

    def handler(request: httpx.Request) -> httpx.Response:
        page = int(request.url.params["page"])
        size = int(request.url.params["size"])
        start = (page - 1) * size
        stop = min(record_count, start + size)
        return httpx.Response(200, json={"data": [{"id": f"R{index:06d}"} for index in range(start, stop)], "total": record_count})

    spill_after_records = int(os.environ.get("PERF_HTTP_SPILL_AFTER_RECORDS", "10000"))
    source = ManagedHttpSource(ConnectionRegistry(registry_path), transport=httpx.MockTransport(handler), resolver=lambda _host: ["93.184.216.34"], spill_after_records=spill_after_records)
    config = StandardSourceConfig.model_validate({
        "type": "managed_http",
        "connection_id": "performance-http",
        "path": "/records",
        "data_json_path": "$.data",
        "pagination": {"size": page_size, "total_json_path": "$.total", "max_pages": (record_count // page_size) + 2},
    })
    started = time.perf_counter()
    cpu_started = time.process_time()
    records, metadata = source.fetch_with_metadata(config)
    try:
        elapsed = time.perf_counter() - started
        metrics = {"benchmark_version": 3, "records": len(records), "page_size": page_size, "pages": len(metadata["pages"]), "cpu_seconds": round(time.process_time() - cpu_started, 3), "elapsed_seconds": round(elapsed, 3), "response_bytes": metadata["response_bytes"], "record_storage": metadata["record_storage"]}
        print(metrics)
        if output := os.environ.get("PERF_HTTP_RESULT_PATH"):
            Path(output).write_text(json.dumps(metrics, indent=2), encoding="utf-8")
        assert len(records) == record_count
        assert len(metadata["pages"]) == (record_count + page_size - 1) // page_size
        assert isinstance(records, SpilledRecords) and metadata["record_storage"] == "disk_spill"
        assert elapsed <= float(os.environ.get("PERF_HTTP_MAX_SECONDS", "60"))
    finally:
        records.close()


@pytest.mark.performance
def test_maximum_standard_comparison_baseline(tmp_path):
    if os.environ.get("PERF_RUN_LARGE_STANDARD") != "1":
        pytest.skip("set PERF_RUN_LARGE_STANDARD=1 to execute the maximum standard comparison")
    record_count = int(os.environ.get("PERF_LARGE_STANDARD_RECORDS", "500000"))
    excel_rows = int(os.environ.get("PERF_LARGE_EXCEL_ROWS", "99999"))
    spill_after = int(os.environ.get("PERF_LARGE_DIFFERENCE_SPILL", "10000"))
    assert 1 <= excel_rows <= 99_999 and excel_rows <= record_count
    rules = RuleSet.model_validate({
        "schema_id": "maximum-standard",
        "schema_version": "1.0.0",
        "name": "Maximum standard comparison",
        "workbook": {"max_standard_records": record_count},
        "sheets": [{
            "id": "data",
            "name": "Data",
            "primary_key": ["id"],
            "columns": [{"name": "id", "title": "ID", "required": True}],
        }],
    })
    workbook_path = tmp_path / "maximum-standard.xlsx"
    book = Workbook(write_only=True)
    sheet = book.create_sheet("Data")
    sheet.append(["ID"])
    for index in range(excel_rows):
        sheet.append([f"R{index:06d}"])
    book.save(workbook_path)
    standard = SpilledRecords()
    for index in range(record_count):
        standard.append({"id": f"R{index:06d}"})

    cpu_calibration_seconds = _python_cpu_calibration_seconds()
    process = psutil.Process()
    baseline_rss = process.memory_info().rss
    peak_rss = [baseline_rss]
    stopped = threading.Event()

    def sample_memory():
        while not stopped.wait(0.02):
            peak_rss[0] = max(peak_rss[0], process.memory_info().rss)

    sampler = threading.Thread(target=sample_memory, daemon=True)
    sampler.start()
    started = time.perf_counter()
    cpu_started = process.cpu_times()
    snapshot = inspect_workbook(workbook_path, rules)
    inspected = time.perf_counter()
    result = None
    try:
        result = compare_workbook(
            snapshot,
            {"data": standard},
            rules,
            job_id="job_maximum_standard",
            difference_spill_threshold=spill_after,
        )
        compared = time.perf_counter()
        report = AuditReport(
            job_id="job_maximum_standard",
            created_at=datetime.now(timezone.utc),
            schema_id=rules.schema_id,
            schema_version=rules.schema_version,
            schema_sha256=rules.content_sha256,
            input_sha256=snapshot.sha256,
            standard_snapshot_id="std_maximum_standard",
            standard_sha256="0" * 64,
            header_mappings=result.mappings,
            differences=result.differences,
            summary=result.summary,
        )
        report_path = tmp_path / "maximum-standard-report.json"
        jsonl_path = tmp_path / "maximum-standard-differences.jsonl"
        write_json_report(report, report_path)
        write_differences_jsonl(report.differences, jsonl_path)
        elapsed = time.perf_counter() - started
        cpu_finished = process.cpu_times()
        cpu_seconds = (cpu_finished.user + cpu_finished.system) - (cpu_started.user + cpu_started.system)
        metrics = {
            "benchmark_version": 4,
            "standard_records": record_count,
            "excel_rows": excel_rows,
            "matched_records": result.summary.matched_records,
            "missing_records": result.summary.missing_records,
            "differences": result.summary.differences,
            "join_backends": sorted(set(result.join_backends or [])),
            "storage_backends": sorted(set(result.storage_backends or [])),
            "report_only": result.report_only,
            "inspect_seconds": round(inspected - started, 3),
            "compare_seconds": round(compared - inspected, 3),
            "report_seconds": round(elapsed - (compared - started), 3),
            "render_seconds": 0,
            "cpu_calibration_seconds": round(cpu_calibration_seconds, 6),
            "cpu_seconds": round(cpu_seconds, 3),
            "normalized_cpu_units": round(cpu_seconds / cpu_calibration_seconds, 3),
            "elapsed_seconds": round(elapsed, 3),
            "peak_rss_delta_mib": round((peak_rss[0] - baseline_rss) / 1024 / 1024, 2),
            "report_bytes": report_path.stat().st_size,
            "jsonl_bytes": jsonl_path.stat().st_size,
        }
        print(metrics)
        if output := os.environ.get("PERF_LARGE_STANDARD_RESULT_PATH"):
            Path(output).write_text(json.dumps(metrics, indent=2), encoding="utf-8")
        assert result.summary.matched_records == excel_rows
        assert result.summary.missing_records == record_count - excel_rows
        assert result.summary.differences == record_count - excel_rows
        expected_join = "polars_partitioned" if record_count + excel_rows >= int(os.environ.get("EXCEL_AUDITOR_POLARS_JOIN_THRESHOLD", "50000")) else "python_in_memory"
        assert result.join_backends == [expected_join]
        assert result.report_only and result.storage_backends == ["disk_differences", "disk_standard_records"]
        with jsonl_path.open("r", encoding="utf-8") as handle:
            assert sum(1 for _line in handle) == result.summary.differences
        assert elapsed <= float(os.environ.get("PERF_LARGE_STANDARD_MAX_SECONDS", "900"))
        assert metrics["peak_rss_delta_mib"] <= float(os.environ.get("PERF_LARGE_STANDARD_MAX_RSS_MIB", "2048"))
    finally:
        stopped.set()
        sampler.join()
        if result is not None:
            result.close()
        snapshot.close()
        standard.close()


@pytest.mark.performance
def test_maximum_managed_service_pipeline_baseline(tmp_path):
    if os.environ.get("PERF_RUN_LARGE_SERVICE") != "1":
        pytest.skip("set PERF_RUN_LARGE_SERVICE=1 to execute the maximum managed service pipeline")
    record_count = int(os.environ.get("PERF_SERVICE_STANDARD_RECORDS", "500000"))
    excel_rows = int(os.environ.get("PERF_SERVICE_EXCEL_ROWS", "99999"))
    page_size = int(os.environ.get("PERF_SERVICE_PAGE_SIZE", "5000"))
    assert 1 <= excel_rows <= 99_999 and excel_rows <= record_count
    registry_path = tmp_path / "service-connections.json"
    registry_path.write_text(json.dumps({"connections": [{
        "id": "maximum-service",
        "base_url": "https://standards.example.test/",
        "allowed_paths": ["/records"],
        "max_records": record_count,
        "max_response_bytes": 64 * 1024 * 1024,
    }]}), encoding="utf-8")

    def handler(request: httpx.Request) -> httpx.Response:
        page = int(request.url.params["page"])
        size = int(request.url.params["size"])
        start = (page - 1) * size
        stop = min(record_count, start + size)
        return httpx.Response(200, json={"data": [{"id": f"R{index:06d}"} for index in range(start, stop)], "total": record_count})

    managed_source = ManagedHttpSource(
        ConnectionRegistry(registry_path),
        transport=httpx.MockTransport(handler),
        resolver=lambda _host: ["93.184.216.34"],
        spill_after_records=10_000,
    )
    rules = RuleSet.model_validate({
        "schema_id": "maximum-managed-service",
        "schema_version": "1.0.0",
        "name": "Maximum managed service",
        "workbook": {"max_standard_records": record_count},
        "standard_source": {
            "type": "managed_http",
            "connection_id": "maximum-service",
            "path": "/records",
            "data_json_path": "$.data",
            "pagination": {"size": page_size, "total_json_path": "$.total", "max_pages": (record_count // page_size) + 2},
        },
        "sheets": [{
            "id": "data",
            "name": "Data",
            "primary_key": ["id"],
            "columns": [{"name": "id", "title": "ID", "required": True}],
        }],
    })
    workbook_path = tmp_path / "maximum-managed-service.xlsx"
    book = Workbook(write_only=True)
    sheet = book.create_sheet("Data")
    sheet.append(["ID"])
    for index in range(excel_rows):
        sheet.append([f"R{index:06d}"])
    book.save(workbook_path)

    class MustNotRender(ExcelRenderer):
        def render(self, source, destination, workbook, rules, comparison, report_payload):
            raise AssertionError("maximum managed pipeline must switch to report-only mode")

    cpu_calibration_seconds = _python_cpu_calibration_seconds()
    process = psutil.Process()
    baseline_rss = process.memory_info().rss
    peak_rss = [baseline_rss]
    stopped = threading.Event()

    def sample_memory():
        while not stopped.wait(0.02):
            peak_rss[0] = max(peak_rss[0], process.memory_info().rss)

    sampler = threading.Thread(target=sample_memory, daemon=True)
    sampler.start()
    started = time.perf_counter()
    cpu_started = time.process_time()
    service = AuditService(tmp_path / "service-runtime", renderer=MustNotRender(), managed_http=managed_source)
    job_id = service.create_job()
    try:
        service.run(job_id, workbook_path, None, rules)
        elapsed = time.perf_counter() - started
        status = service.status(job_id)
        assert status["status"] == "completed", status
        report_path = service.artifact(job_id, "json")
        jsonl_path = service.artifact(job_id, "differences_jsonl")
        html_path = service.artifact(job_id, "html")
        with jsonl_path.open("r", encoding="utf-8") as handle:
            jsonl_lines = sum(1 for _line in handle)
        cpu_seconds = time.process_time() - cpu_started
        metrics = {
            "benchmark_version": 4,
            "standard_records": record_count,
            "excel_rows": excel_rows,
            "page_size": page_size,
            "pages": (record_count + page_size - 1) // page_size,
            "matched_records": status["summary"]["matched_records"],
            "missing_records": status["summary"]["missing_records"],
            "differences": status["summary"]["differences"],
            "mode": status.get("mode"),
            "cpu_calibration_seconds": round(cpu_calibration_seconds, 6),
            "cpu_seconds": round(cpu_seconds, 3),
            "normalized_cpu_units": round(cpu_seconds / cpu_calibration_seconds, 3),
            "elapsed_seconds": round(elapsed, 3),
            "peak_rss_delta_mib": round((peak_rss[0] - baseline_rss) / 1024 / 1024, 2),
            "report_bytes": report_path.stat().st_size,
            "jsonl_bytes": jsonl_path.stat().st_size,
            "html_bytes": html_path.stat().st_size,
        }
        print(metrics)
        if output := os.environ.get("PERF_SERVICE_RESULT_PATH"):
            Path(output).write_text(json.dumps(metrics, indent=2), encoding="utf-8")
        assert status["mode"] == "report_only"
        assert status["summary"]["matched_records"] == excel_rows
        assert status["summary"]["missing_records"] == record_count - excel_rows
        assert status["summary"]["differences"] == record_count - excel_rows
        assert "comparison_storage:disk_standard_records" in status["warnings"]
        expected_difference_storage = (
            "disk_differences"
            if record_count - excel_rows > int(os.environ.get("EXCEL_AUDITOR_DIFFERENCE_SPILL_THRESHOLD", "50000"))
            else "memory_differences"
        )
        assert f"comparison_storage:{expected_difference_storage}" in status["warnings"]
        assert jsonl_lines == record_count - excel_rows
        assert elapsed <= float(os.environ.get("PERF_SERVICE_MAX_SECONDS", "900"))
        assert metrics["peak_rss_delta_mib"] <= float(os.environ.get("PERF_SERVICE_MAX_RSS_MIB", "2048"))
    finally:
        stopped.set()
        sampler.join()


@pytest.mark.performance
def test_maximum_uploaded_json_pipeline_baseline(tmp_path):
    if os.environ.get("PERF_RUN_LARGE_UPLOAD") != "1":
        pytest.skip("set PERF_RUN_LARGE_UPLOAD=1 to execute the maximum uploaded JSON pipeline")
    record_count = int(os.environ.get("PERF_UPLOAD_STANDARD_RECORDS", "500000"))
    spill_after = int(os.environ.get("PERF_UPLOAD_SPILL_AFTER_RECORDS", "100000"))
    rules = RuleSet.model_validate({
        "schema_id": "maximum-uploaded-standard",
        "schema_version": "1.0.0",
        "name": "Maximum uploaded standard",
        "workbook": {"max_standard_records": record_count},
        "sheets": [{
            "id": "data",
            "name": "Data",
            "primary_key": ["id"],
            "columns": [{"name": "id", "title": "ID", "required": True}],
        }],
    })
    upload_path = tmp_path / "maximum-upload.json"
    with upload_path.open("w", encoding="utf-8", newline="") as handle:
        handle.write('{"data":[')
        for index in range(record_count):
            if index:
                handle.write(",")
            handle.write(f'{{"id":"R{index:06d}"}}')
        handle.write("]}")

    cpu_calibration_seconds = _python_cpu_calibration_seconds()
    process = psutil.Process()
    baseline_rss = process.memory_info().rss
    peak_rss = [baseline_rss]
    stopped = threading.Event()

    def sample_memory():
        while not stopped.wait(0.02):
            peak_rss[0] = max(peak_rss[0], process.memory_info().rss)

    sampler = threading.Thread(target=sample_memory, daemon=True)
    sampler.start()
    started = time.perf_counter()
    cpu_started = time.process_time()
    standard = None
    try:
        standard = load_standard_file(upload_path, rules, spill_after_records=spill_after)
        loaded = time.perf_counter()
        StandardDataValidator().validate(standard, rules)
        validated = time.perf_counter()
        snapshot = create_snapshot(standard, tmp_path / "snapshot", {"upload_format": "json"})
        elapsed = time.perf_counter() - started
        cpu_seconds = time.process_time() - cpu_started
        rows = standard["data"]
        metrics = {
            "benchmark_version": 4,
            "source_format": "json_upload",
            "standard_records": record_count,
            "record_storage": "disk_spill" if isinstance(rows, SpilledRecords) else "memory",
            "load_seconds": round(loaded - started, 3),
            "validate_seconds": round(validated - loaded, 3),
            "snapshot_seconds": round(elapsed - (validated - started), 3),
            "cpu_calibration_seconds": round(cpu_calibration_seconds, 6),
            "cpu_seconds": round(cpu_seconds, 3),
            "normalized_cpu_units": round(cpu_seconds / cpu_calibration_seconds, 3),
            "elapsed_seconds": round(elapsed, 3),
            "peak_rss_delta_mib": round((peak_rss[0] - baseline_rss) / 1024 / 1024, 2),
            "upload_bytes": upload_path.stat().st_size,
            "snapshot_bytes": snapshot.path.stat().st_size,
            "snapshot_records": snapshot.record_count,
            "snapshot_sha256": snapshot.sha256,
        }
        print(metrics)
        if output := os.environ.get("PERF_UPLOAD_RESULT_PATH"):
            Path(output).write_text(json.dumps(metrics, indent=2), encoding="utf-8")
        assert isinstance(rows, SpilledRecords) and len(rows) == record_count
        assert rows[0]["id"] == "R000000" and rows[-1]["id"] == f"R{record_count - 1:06d}"
        assert snapshot.record_count == record_count and len(snapshot.sha256) == 64
        assert elapsed <= float(os.environ.get("PERF_UPLOAD_MAX_SECONDS", "300"))
        assert metrics["peak_rss_delta_mib"] <= float(os.environ.get("PERF_UPLOAD_MAX_RSS_MIB", "1024"))
    finally:
        stopped.set()
        sampler.join()
        if standard is not None:
            for rows in standard.values():
                close = getattr(rows, "close", None)
                if close is not None:
                    close()
