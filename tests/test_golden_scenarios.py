import json
from collections import Counter
from pathlib import Path

import pytest
from excel_auditor.engine import compare_workbook
from excel_auditor.models import RuleSet
from excel_auditor.workbook import inspect_workbook
from excel_auditor.service import AuditService
from openpyxl import load_workbook


GOLDEN = json.loads(Path("tests/golden_files/core_scenarios.json").read_text(encoding="utf-8"))
SEMANTICS = json.loads(Path("tests/golden_files/core_expected_semantics.json").read_text(encoding="utf-8"))


@pytest.mark.parametrize("scenario", GOLDEN["scenarios"], ids=lambda item: item["name"])
def test_golden_difference_counts(tmp_path, scenario):
    rules = RuleSet.model_validate(GOLDEN["rule"])
    path = Path("tests/golden_files/workbooks") / f"{scenario['name']}.xlsx"
    import hashlib
    assert hashlib.sha256(path.read_bytes()).hexdigest() == scenario["input_sha256"]
    result = compare_workbook(inspect_workbook(path, rules), {"staff": scenario["standard"]}, rules)
    actual = Counter(item.type.value for item in result.differences)
    # Golden expectations are exhaustive: any unexpected difference is a
    # regression, even when every previously expected count still matches.
    assert dict(sorted(actual.items())) == dict(sorted(scenario["expected"].items()))
    projection = [
        {
            key: item.model_dump(mode="json")[key]
            for key in ("type", "cell", "excel_row", "canonical_field", "business_key", "excel_raw_value", "standard_raw_value", "render_action", "repair_status")
        }
        for item in result.differences
    ]
    assert projection == SEMANTICS[scenario["name"]]["differences"]


@pytest.mark.parametrize("scenario", GOLDEN["scenarios"], ids=lambda item: f"render-{item['name']}")
def test_golden_rendered_structure_and_fills(tmp_path, scenario):
    rules = RuleSet.model_validate(GOLDEN["rule"])
    source = Path("tests/golden_files/workbooks") / f"{scenario['name']}.xlsx"
    standard = tmp_path / "standard.json"
    standard.write_text(json.dumps({"staff": scenario["standard"]}, ensure_ascii=False), encoding="utf-8")
    service = AuditService(tmp_path / "runtime")
    job_id = service.create_job()
    service.run(job_id, source, standard, rules)
    assert service.status(job_id)["status"] == "completed"
    rendered = load_workbook(service.artifact(job_id, "excel"), data_only=False)
    sheet = rendered[rules.sheets[0].name]
    expected = SEMANTICS[scenario["name"]]
    assert [sheet.cell(rules.sheets[0].header.row, index).value for index in range(1, len(expected["rendered_headers"]) + 1)] == expected["rendered_headers"]
    for coordinate, color in expected["fills"].items():
        assert sheet[coordinate].fill.fgColor.rgb.endswith(color)
    for difference in expected["differences"]:
        if difference["type"] != "MISSING_HEADER" or difference["repair_status"] != "planned":
            continue
        header_cell = next(
            cell
            for cell in sheet[rules.sheets[0].header.row]
            if cell.value == difference["standard_raw_value"]
        )
        assert header_cell.comment is not None
        assert "原值：（缺失）" in header_cell.comment.text
        assert f"标准值：{difference['standard_raw_value']}" in header_cell.comment.text
        assert f"{difference['canonical_field']}.missing_column" in header_cell.comment.text
    assert "核验报告" in rendered.sheetnames
    rendered.close()
