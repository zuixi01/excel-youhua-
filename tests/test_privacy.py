import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from excel_auditor.engine import compare_workbook
from excel_auditor.models import Difference, DifferenceType, RuleSet
from excel_auditor.persistence import AuditEventRow, DatabaseRepository
from excel_auditor.rendering import ExcelRenderer, _repair_provenance_comment
from excel_auditor.service import AuditService
from excel_auditor.workbook import inspect_workbook


def test_sensitive_values_are_masked_in_differences(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "pii", "schema_version": "1.0.0", "name": "PII",
        "sheets": [{
            "id": "people", "name": "人员", "primary_key": ["id"],
            "columns": [
                {"name": "id", "title": "身份证号", "required": True, "type": "id_code", "sensitive": True},
                {"name": "phone", "title": "手机号", "type": "phone", "sensitive": True},
            ],
        }],
    })
    path = tmp_path / "pii.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "人员"
    sheet.append(["身份证号", "手机号"])
    sheet.append(["123456789012345678", "13800138000"])
    book.save(path)
    result = compare_workbook(inspect_workbook(path, rules), {"people": [{"id": "123456789012345678", "phone": "13900139000"}]}, rules)
    mismatch = next(item for item in result.differences if item.canonical_field == "phone")
    assert mismatch.excel_raw_value == "13***00"
    assert mismatch.standard_raw_value == "13***00"
    assert mismatch.business_key == {"id": "12***78"}


def test_repair_provenance_comment_preserves_required_fields_within_excel_limit():
    difference = Difference(
        difference_id="diff_comment_limit",
        type=DifferenceType.VALUE_MISMATCH,
        sheet_id="data",
        sheet_name="Data",
        excel_raw_value="😀" * 10_000,
        standard_raw_value="🧪" * 10_000,
        message="mismatch",
    )
    comment = _repair_provenance_comment("自动修复", "value.overwrite_mismatch", difference)

    assert len(comment.encode("utf-16-le")) // 2 <= 32767
    assert "原值：" in comment and "标准值：" in comment
    assert "规则：value.overwrite_mismatch" in comment
    assert comment.count("…") == 2


def test_invalid_sensitive_primary_key_is_masked():
    rules = RuleSet.model_validate({
        "schema_id": "sensitive-key", "schema_version": "1.0.0", "name": "Sensitive key",
        "sheets": [{"id": "data", "name": "Data", "primary_key": ["id"], "columns": [
            {"name": "id", "title": "ID", "required": True, "type": "integer", "sensitive": True},
        ]}],
    })
    from excel_auditor.workbook import SheetSnapshot, WorkbookSnapshot

    workbook = WorkbookSnapshot(Path("sensitive-key.xlsx"), "0" * 64, {
        "Data": SheetSnapshot("Data", 2, 1, [(1, ["ID"]), (2, ["secret-primary-key"])])
    })
    result = compare_workbook(workbook, {"data": []}, rules)
    difference = next(item for item in result.differences if item.type.value == "EMPTY_PRIMARY_KEY")
    assert difference.excel_raw_value == "se***ey"


def test_sensitive_values_do_not_leak_to_secondary_artifacts_or_database_audit(tmp_path):
    excel_secret = "13800138000"
    standard_secret = "13900139000"
    rules = RuleSet.model_validate({
        "schema_id": "pii-artifacts", "schema_version": "1.0.0", "name": "PII artifacts",
        "sheets": [{
            "id": "people", "name": "People", "primary_key": ["id"],
            "actions": {"overwrite_mismatch": True},
            "columns": [
                {"name": "id", "title": "ID", "required": True},
                {"name": "phone", "title": "Phone", "type": "phone", "sensitive": True},
            ],
        }],
    })
    workbook_path, standard_path = tmp_path / "pii-artifacts.xlsx", tmp_path / "pii-artifacts.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "People"
    sheet.append(["ID", "Phone"])
    sheet.append(["E1", excel_secret])
    book.save(workbook_path)
    standard_path.write_text(json.dumps({"people": [{"id": "E1", "phone": standard_secret}]}), encoding="utf-8")
    database = DatabaseRepository(f"sqlite:///{(tmp_path / 'audit.db').as_posix()}")
    service = AuditService(tmp_path / "runtime", database=database)
    job_id = service.create_job()
    service.run(job_id, workbook_path, standard_path, rules)

    status = service.status(job_id)
    assert status["status"] == "completed", status
    for artifact in ("json", "differences_jsonl", "html", "manifest"):
        payload = service.artifact(job_id, artifact).read_text(encoding="utf-8")
        assert excel_secret not in payload and standard_secret not in payload
    assert excel_secret not in json.dumps(status) and standard_secret not in json.dumps(status)
    assert not (service.job_directory(job_id) / "render-manifest.private.json").exists()
    assert not (service.job_directory(job_id) / "report-render.json").exists()

    rendered = load_workbook(service.artifact(job_id, "excel"), data_only=False)
    assert rendered["People"]["B2"].value == standard_secret  # authorized business output
    repair_comment = rendered["People"]["B2"].comment.text or ""
    assert excel_secret not in repair_comment
    assert standard_secret not in repair_comment
    assert "原值：13***00" in repair_comment and "标准值：13***00" in repair_comment
    assert "规则：phone.overwrite_mismatch" in repair_comment
    embedded_report = "\n".join(
        "|".join("" if value is None else str(value) for value in row)
        for row in rendered["核验报告"].iter_rows(values_only=True)
    )
    assert excel_secret not in embedded_report and standard_secret not in embedded_report
    rendered.close()
    with Session(database.engine) as session:
        audit_events = session.scalars(select(AuditEventRow)).all()
        audit_payload = json.dumps([row.metadata_json for row in audit_events])
    assert excel_secret not in audit_payload and standard_secret not in audit_payload
    repair_event = next(row for row in audit_events if row.action == "comparison.auto_repair_operation")
    assert repair_event.metadata_json["rule_id"] == "phone.overwrite_mismatch"
    assert repair_event.metadata_json["difference_rule_id"] == "phone.exact"
    assert repair_event.metadata_json["excel_raw_value"] == "13***00"
    assert repair_event.metadata_json["standard_raw_value"] == "13***00"
    assert repair_event.metadata_json["status"] == "applied"


def test_failure_status_and_diagnostic_never_store_exception_message(tmp_path):
    secret = "sensitive-renderer-detail-must-not-leak"

    class FailingRenderer(ExcelRenderer):
        def render(self, source, destination, workbook, rules, comparison, report_payload):
            raise RuntimeError(f"RENDER_FAILED: {secret}")

    rules = RuleSet.model_validate({
        "schema_id": "safe-failure", "schema_version": "1.0.0", "name": "Safe failure",
        "sheets": [{"id": "data", "name": "Data", "primary_key": ["id"], "columns": [{"name": "id", "title": "ID", "required": True}]}],
    })
    workbook_path = tmp_path / "input.xlsx"
    book = Workbook()
    book.active.title = "Data"
    book.active.append(["ID"])
    book.active.append(["E1"])
    book.save(workbook_path)
    standard_path = tmp_path / "standard.json"
    standard_path.write_text(json.dumps({"data": [{"id": "E1"}]}), encoding="utf-8")
    service = AuditService(tmp_path / "runtime", renderer=FailingRenderer())
    job_id = service.create_job()
    service.run(job_id, workbook_path, standard_path, rules)
    status = service.status(job_id)
    diagnostic = (service.job_directory(job_id) / "diagnostic.log").read_text(encoding="utf-8")
    assert status["status"] == "failed" and status["error_code"] == "RENDER_FAILED"
    assert status["error_message_safe"] == "Workbook rendering failed."
    assert secret not in json.dumps(status) and secret not in diagnostic
