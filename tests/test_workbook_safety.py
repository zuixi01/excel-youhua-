import json
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from excel_auditor.models import RuleSet
from excel_auditor.rules import load_rules
from excel_auditor.service import AuditService
from excel_auditor.workbook import SpilledRows, WorkbookSafetyError, inspect_workbook


def test_package_feature_inventory_detects_drawings(tmp_path):
    path = tmp_path / "drawing.xlsx"
    book = Workbook()
    book.save(path)
    with zipfile.ZipFile(path, "a") as archive:
        archive.writestr("xl/drawings/drawing1.xml", "<drawing/>")
    snapshot = inspect_workbook(path, load_rules(Path("configs/examples/employee-roster.yaml")))
    assert "workbook: drawings" in snapshot.warnings
    assert "workbook: drawings" in snapshot.manual_review_reasons


@pytest.mark.parametrize(
    ("object_type", "requires_review"),
    [("Note", False), ("Radio", True), ("Scroll", True)],
)
def test_vml_inventory_distinguishes_comments_from_legacy_controls(tmp_path, object_type, requires_review):
    path = tmp_path / f"vml-{object_type}.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID"])
    book.save(path)
    vml = (
        '<xml xmlns:v="urn:schemas-microsoft-com:vml" '
        'xmlns:x="urn:schemas-microsoft-com:office:excel">'
        f'<v:shape id="shape1"><x:ClientData ObjectType="{object_type}"/></v:shape>'
        '</xml>'
    )
    with zipfile.ZipFile(path, "a") as archive:
        archive.writestr("xl/drawings/vmlDrawing1.vml", vml)
    rules = RuleSet.model_validate({
        "schema_id": "vml", "schema_version": "1.0.0", "name": "VML",
        "sheets": [{"id": "data", "name": "Data", "primary_key": ["id"], "columns": [
            {"name": "id", "title": "ID", "required": True},
        ]}],
    })

    snapshot = inspect_workbook(path, rules)
    assert ("workbook: legacy_controls" in snapshot.manual_review_reasons) is requires_review


def test_unsafe_package_feature_cannot_be_bypassed_by_report_action(tmp_path):
    path = tmp_path / "drawing-report.xlsx"
    book = Workbook()
    book.active.title = "Data"
    book.active.append(["ID"])
    book.save(path)
    with zipfile.ZipFile(path, "a") as archive:
        archive.writestr("xl/drawings/drawing1.xml", "<drawing/>")
    rules = RuleSet.model_validate({
        "schema_id": "drawing-report",
        "schema_version": "1.0.0",
        "name": "Drawing report",
        "workbook": {"unsupported_feature_action": "report"},
        "sheets": [{
            "id": "data",
            "name": "Data",
            "primary_key": ["id"],
            "columns": [{"name": "id", "title": "ID", "required": True}],
        }],
    })
    snapshot = inspect_workbook(path, rules)
    assert snapshot.manual_review_reasons == ["workbook: drawings"]
    standard = tmp_path / "standard.json"
    standard.write_text(json.dumps({"data": []}), encoding="utf-8")
    service = AuditService(tmp_path / "runtime")
    job_id = service.create_job()
    service.run(job_id, path, standard, rules)
    status = service.status(job_id)
    assert status["status"] == "manual_review"
    assert "excel" not in status["artifacts"]


def test_dtd_is_rejected_before_workbook_parsing(tmp_path):
    path = tmp_path / "unsafe.xlsx"
    book = Workbook()
    book.save(path)
    with zipfile.ZipFile(path, "a") as archive:
        archive.writestr("xl/unsafe.xml", '<!DOCTYPE x [<!ENTITY leak SYSTEM "file:///etc/passwd">]><x>&leak;</x>')
    with pytest.raises(WorkbookSafetyError, match="DTD"):
        inspect_workbook(path, load_rules(Path("configs/examples/employee-roster.yaml")))


def test_excessive_xml_depth_is_rejected(tmp_path):
    path = tmp_path / "deep.xlsx"
    book = Workbook()
    book.save(path)
    with zipfile.ZipFile(path, "a") as archive:
        archive.writestr("xl/deep.xml", "<x>" * 101 + "</x>" * 101)
    with pytest.raises(WorkbookSafetyError, match="nesting depth"):
        inspect_workbook(path, load_rules(Path("configs/examples/employee-roster.yaml")))


def test_streaming_mode_keeps_rows_and_detects_xml_level_features(tmp_path):
    path = tmp_path / "streaming.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID", "Value"])
    sheet.append(["E1", "A"])
    sheet.auto_filter.ref = "A1:B2"
    validation = DataValidation(type="list", formula1='"A,B"')
    validation.add("B2")
    sheet.add_data_validation(validation)
    book.save(path)
    rules = RuleSet.model_validate({"schema_id": "stream", "schema_version": "1.0.0", "name": "Stream", "sheets": [{"id": "data", "name": "Data", "primary_key": ["id"], "columns": [{"name": "id", "title": "ID", "required": True}, {"name": "value", "title": "Value"}]}]})
    snapshot = inspect_workbook(path, rules, max_in_memory_cells=1)
    assert isinstance(snapshot.sheets["Data"].rows, SpilledRows)
    assert snapshot.sheets["Data"].rows[1][1] == ["E1", "A"]
    assert {"auto_filter", "data_validations"} <= set(snapshot.sheets["Data"].risky_features)
    assert not snapshot.manual_review_reasons
    assert not snapshot.sheets["Data"].rows._file.closed
    snapshot.close()
    assert snapshot.sheets["Data"].rows._file.closed


def test_workbook_sheet_limit_and_structure_inventory(tmp_path):
    path = tmp_path / "too-many-sheets.xlsx"
    book = Workbook()
    for index in range(50):
        book.create_sheet(f"S{index}")
    book.save(path)
    with pytest.raises(WorkbookSafetyError, match="more than 50"):
        inspect_workbook(path, load_rules(Path("configs/examples/employee-roster.yaml")))

    structure = tmp_path / "structure.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID"])
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].hidden = True
    book.save(structure)
    rules = RuleSet.model_validate({"schema_id": "structure", "schema_version": "1.0.0", "name": "Structure", "sheets": [{"id": "data", "name": "Data", "primary_key": ["id"], "columns": [{"name": "id", "title": "ID", "required": True}]}]})
    snapshot = inspect_workbook(structure, rules)
    assert {"freeze_panes", "hidden_columns"} <= set(snapshot.sheets["Data"].risky_features)
    assert not snapshot.manual_review_reasons


def test_row_limit_is_counted_from_the_configured_data_start(tmp_path):
    path = tmp_path / "data-start-limit.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet["A1"] = "ID"
    sheet["A5"] = "E1"
    sheet["A6"] = "E2"
    sheet["A7"] = "E3"
    book.save(path)
    rules = RuleSet.model_validate({
        "schema_id": "data-start-limit", "schema_version": "1.0.0", "name": "Data start limit",
        "workbook": {"max_rows_per_sheet": 3},
        "sheets": [{
            "id": "data", "name": "Data", "data_region": {"start_row": 5},
            "primary_key": ["id"], "columns": [{"name": "id", "title": "ID", "required": True}],
        }],
    })

    snapshot = inspect_workbook(path, rules)
    assert snapshot.sheets["Data"].max_row == 7
    snapshot.close()

    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet["A1"] = "ID"
    sheet["A8"] = "E4"
    book.save(path)
    with pytest.raises(WorkbookSafetyError, match="FILE_LIMIT_EXCEEDED"):
        inspect_workbook(path, rules)


@pytest.mark.parametrize("max_in_memory_cells", [10_000, 1])
def test_row_limit_is_counted_from_the_auto_detected_header(tmp_path, max_in_memory_cells):
    path = tmp_path / f"auto-header-limit-{max_in_memory_cells}.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet["A1"] = "Report title"
    sheet["A3"] = "ID"
    sheet["A4"] = "E1"
    sheet["A5"] = "E2"
    book.save(path)
    rules = RuleSet.model_validate({
        "schema_id": "auto-header-limit", "schema_version": "1.0.0", "name": "Auto header limit",
        "workbook": {"max_rows_per_sheet": 2},
        "sheets": [{
            "id": "data", "name": "Data", "header": {"auto_detect": True},
            "primary_key": ["id"], "columns": [{"name": "id", "title": "ID", "required": True}],
        }],
    })

    snapshot = inspect_workbook(path, rules, max_in_memory_cells=max_in_memory_cells)
    assert snapshot.sheets["Data"].max_row == 5
    snapshot.close()

    sheet["A6"] = "E3"
    book.save(path)
    with pytest.raises(WorkbookSafetyError, match="FILE_LIMIT_EXCEEDED"):
        inspect_workbook(path, rules, max_in_memory_cells=max_in_memory_cells)


def test_only_merged_header_requires_review(tmp_path):
    path = tmp_path / "merged.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID", "Value"])
    sheet.append(["E1", "A"])
    sheet.merge_cells("A2:B2")
    book.save(path)
    rules = RuleSet.model_validate({
        "schema_id": "merged",
        "schema_version": "1.0.0",
        "name": "Merged",
        "sheets": [{
            "id": "data",
            "name": "Data",
            "primary_key": ["id"],
            "columns": [
                {"name": "id", "title": "ID", "required": True},
                {"name": "value", "title": "Value"},
            ],
        }],
    })
    body_merge = inspect_workbook(path, rules)
    assert "Data: merged_cells" in body_merge.warnings
    assert not body_merge.manual_review_reasons

    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID", None])
    sheet.merge_cells("A1:B1")
    book.save(path)
    header_merge = inspect_workbook(path, rules)
    assert "Data: merged_header" in header_merge.manual_review_reasons


def test_auto_detected_header_drives_formula_and_merge_safety_checks(tmp_path):
    formula_path = tmp_path / "auto-header-formula.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["Report title"])
    sheet.append(["ID", "Value"])
    sheet.append(["E1", "=1+1"])
    book.save(formula_path)
    rules = RuleSet.model_validate({
        "schema_id": "auto-header-safety", "schema_version": "1.0.0", "name": "Auto header safety",
        "sheets": [{
            "id": "data", "name": "Data", "header": {"auto_detect": True}, "primary_key": ["id"],
            "columns": [
                {"name": "id", "title": "ID", "required": True},
                {"name": "value", "title": "Value", "compare": {"formula_mode": "formula"}},
            ],
        }],
    })
    formula_snapshot = inspect_workbook(formula_path, rules)
    assert "Data: formulas" not in formula_snapshot.manual_review_reasons
    formula_snapshot.close()

    merged_path = tmp_path / "auto-header-merged.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["Report title"])
    sheet.append(["ID", "Value"])
    sheet.merge_cells("A2:B2")
    book.save(merged_path)
    merged_snapshot = inspect_workbook(merged_path, rules)
    assert "Data: merged_header" in merged_snapshot.manual_review_reasons
    merged_snapshot.close()
