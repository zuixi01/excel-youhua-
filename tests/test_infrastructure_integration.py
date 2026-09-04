import json
import os
import importlib
from pathlib import Path

import boto3
import pytest
from alembic import command
from alembic.config import Config
from openpyxl import Workbook
from rq import SimpleWorker
from sqlalchemy import select
from sqlalchemy.orm import Session
from fastapi.testclient import TestClient

from excel_auditor.models import RuleSet
from excel_auditor.persistence import AuditEventRow, ComparisonJobRow, DatabaseRepository
from excel_auditor.queueing import RedisJobQueue
from excel_auditor.service import AuditService
from excel_auditor.storage import S3ArtifactStore
from excel_auditor.product_workflow import (
    CatalogFieldDefinition,
    CatalogFieldSource,
    CategoryDefinition,
    InMemoryCatalogAdapter,
    ProductReviewDecision,
)
from excel_auditor.product_workflow.service import ProductWorkflowService


@pytest.mark.integration
def test_postgres_redis_rq_minio_renderer_end_to_end(tmp_path, monkeypatch):
    if os.environ.get("RUN_INFRA_INTEGRATION") != "1":
        pytest.skip("infrastructure integration is opt-in")
    database_url = os.environ["INTEGRATION_DATABASE_URL"]
    redis_url = os.environ["INTEGRATION_REDIS_URL"]
    endpoint = os.environ["INTEGRATION_S3_ENDPOINT"]
    bucket = "excel-auditor-integration"
    monkeypatch.setenv("DATABASE_URL", database_url)
    monkeypatch.setenv("REDIS_URL", redis_url)
    monkeypatch.setenv("S3_BUCKET", bucket)
    monkeypatch.setenv("S3_ENDPOINT_URL", endpoint)
    monkeypatch.setenv("S3_SERVER_SIDE_ENCRYPTION", "")
    migration = Config("alembic.ini")
    migration.set_main_option("sqlalchemy.url", database_url)
    command.upgrade(migration, "head")
    s3 = boto3.client("s3", endpoint_url=endpoint, region_name="us-east-1")
    s3.create_bucket(Bucket=bucket)
    database = DatabaseRepository(database_url, create_schema=False)
    store = S3ArtifactStore(bucket, endpoint, "us-east-1", None)
    root = tmp_path / "runtime"
    service = AuditService(root, database=database, artifact_store=store)
    queue = RedisJobQueue(redis_url)
    queue.connection.flushdb()
    rules = RuleSet.model_validate({
        "schema_id": "infra", "schema_version": "1.0.0", "name": "Infrastructure",
        "sheets": [{"id": "data", "name": "Data", "primary_key": ["id"], "columns": [
            {"name": "id", "title": "ID", "required": True}, {"name": "value", "title": "Value"},
        ]}],
    })
    job_id = service.create_job(tenant_id="integration", user_id="ci")
    directory = service.job_directory(job_id)
    excel, standard = directory / "upload.xlsx", directory / "standard.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID", "Value"])
    sheet.append(["E1", "old"])
    book.save(excel)
    standard.write_text(json.dumps({"data": [{"id": "E1", "value": "new"}]}), encoding="utf-8")
    queue.enqueue(root, job_id, excel, standard, rules, {})
    assert SimpleWorker([queue.queue], connection=queue.connection).work(burst=True, logging_level="WARNING")
    status = service.status(job_id)
    assert status["status"] == "completed", status
    assert status["object_keys"]
    for key in status["object_keys"].values():
        assert s3.head_object(Bucket=bucket, Key=key)["ContentLength"] > 0
    with Session(database.engine) as session:
        row = session.scalar(select(ComparisonJobRow).where(ComparisonJobRow.id == job_id))
        assert row is not None and row.status == "completed" and row.tenant_id == "integration"

    product_rules = RuleSet.model_validate({
        "schema_id": "infra-product", "schema_version": "1.0.0", "name": "Infrastructure product",
        "sheets": [{"id": "products", "name": "Products", "primary_key": ["product_id"], "columns": [
            {"name": "product_id", "title": "Product ID", "required": True},
            {"name": "category_id", "title": "Category ID"},
            {"name": "category_name", "title": "Category"},
        ]}],
        "product_workflow": {
            "sheet_id": "products", "catalog_connection_id": "integration-catalog", "mapping_fuzzy_threshold": 50,
            "category": {
                "source_field": "category_name", "id_field": "category_id",
                "attributes": {"path_template": "/catalog/{category_id}/attributes"},
                "specifications": {"path_template": "/catalog/{category_id}/specifications"},
            },
        },
    })
    product_job = service.create_job(tenant_id="integration", user_id="ci")
    product_input = service.job_directory(product_job) / "product-input.xlsx"
    book = Workbook(); sheet = book.active; sheet.title = "Products"
    sheet.append(["Product ID", "Category ID", "Category", "Brnad"])
    sheet.append(["P-1", "phone", "Phone", "Example"])
    book.save(product_input)
    product_catalog = InMemoryCatalogAdapter(
        [CategoryDefinition(category_id="phone", name="Phone")],
        {"phone": [CatalogFieldDefinition(
            field_id="brand", title="Brand", source=CatalogFieldSource.PLATFORM_ATTRIBUTE,
            category_id="phone", attribute_id="brand", required=True,
        )]},
        connection_id="integration-catalog",
    )
    product_workflow = ProductWorkflowService(service, database)
    product_workflow.run(
        product_job, product_input, product_rules, product_catalog,
        tenant_id="integration", actor_id="ci",
    )
    review = next(item for item in product_workflow.list_reviews(product_job, tenant_id="integration") if item["review_type"] == "field_mapping")
    product_workflow.decide_review(
        product_job,
        review["review_id"],
        ProductReviewDecision(action="confirm_mapping", field_id="brand", raw_header="Brnad"),
        tenant_id="integration",
        actor_id="ci",
    )
    queue.enqueue_product_revision(
        root, product_job, product_rules, "integration", "ci", 1,
    )
    assert SimpleWorker([queue.queue], connection=queue.connection).work(burst=True, logging_level="WARNING")
    product_status = service.status(product_job)
    assert product_status["status"] == "completed" and product_status["revision_number"] == 2
    assert s3.head_object(Bucket=bucket, Key=product_status["object_keys"]["product-result.xlsx"])["ContentLength"] > 0


@pytest.mark.integration
def test_real_http_api_queue_download_cancel_failure_and_idempotence(tmp_path, monkeypatch):
    if os.environ.get("RUN_INFRA_INTEGRATION") != "1":
        pytest.skip("infrastructure integration is opt-in")
    database_url = os.environ["INTEGRATION_DATABASE_URL"]
    redis_url = os.environ["INTEGRATION_REDIS_URL"]
    endpoint = os.environ["INTEGRATION_S3_ENDPOINT"]
    bucket = "excel-auditor-api-integration"
    monkeypatch.setenv("DATABASE_URL", database_url)
    monkeypatch.setenv("REDIS_URL", redis_url)
    monkeypatch.setenv("S3_BUCKET", bucket)
    monkeypatch.setenv("S3_ENDPOINT_URL", endpoint)
    monkeypatch.setenv("S3_SERVER_SIDE_ENCRYPTION", "")
    monkeypatch.setenv("EXCEL_AUDITOR_DATA", str(tmp_path / "api-runtime"))
    monkeypatch.setenv("EXCEL_AUDITOR_REQUIRE_AUTH", "1")
    monkeypatch.setenv("EXCEL_AUDITOR_API_TOKENS_JSON", json.dumps({"integration-token": {"tenant_id": "integration-api", "user_id": "ci"}}))
    migration = Config("alembic.ini")
    migration.set_main_option("sqlalchemy.url", database_url)
    command.upgrade(migration, "head")
    s3 = boto3.client("s3", endpoint_url=endpoint, region_name="us-east-1")
    try:
        s3.create_bucket(Bucket=bucket)
    except s3.exceptions.BucketAlreadyOwnedByYou:
        pass
    import excel_auditor.api as api_module
    api_module = importlib.reload(api_module)
    assert api_module.task_queue is not None
    api_module.task_queue.connection.flushdb()
    client = TestClient(api_module.app)
    headers = {"Authorization": "Bearer integration-token", "X-Request-ID": "trace-integration"}
    rules = {
        "schema_id": "infra-api", "schema_version": "1.0.0", "name": "Infrastructure API",
        "sheets": [{"id": "data", "name": "Data", "primary_key": ["id"], "columns": [
            {"name": "id", "title": "ID", "required": True}, {"name": "value", "title": "Value"},
        ]}],
    }
    assert client.post("/api/v1/schemas/publish", headers=headers, json=rules).status_code == 201
    workbook = tmp_path / "api.xlsx"
    book = Workbook(); sheet = book.active; sheet.title = "Data"; sheet.append(["ID", "Value"]); sheet.append(["E1", "old"]); book.save(workbook)
    files = {
        "excel_file": ("api.xlsx", workbook.read_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        "standard_data": ("standard.json", json.dumps({"data": [{"id": "E1", "value": "new"}]}).encode(), "application/json"),
    }
    created = client.post("/api/v1/comparisons", headers={**headers, "Idempotency-Key": "infra-api-success"}, data={"schema_id": "infra-api", "schema_version": "1.0.0"}, files=files)
    assert created.status_code == 202 and created.json()["created_at"]
    job_id = created.json()["job_id"]
    replay = client.post("/api/v1/comparisons", headers={**headers, "Idempotency-Key": "infra-api-success"}, data={"schema_id": "infra-api", "schema_version": "1.0.0"}, files=files)
    assert replay.status_code == 200 and replay.json()["job_id"] == job_id and replay.json()["idempotent_replay"]
    assert SimpleWorker([api_module.task_queue.queue], connection=api_module.task_queue.connection).work(burst=True, logging_level="WARNING")
    status = client.get(f"/api/v1/comparisons/{job_id}", headers=headers)
    assert status.status_code == 200 and status.json()["status"] == "completed" and status.json()["trace_id"] == "trace-integration"
    assert client.get(f"/api/v1/comparisons/{job_id}/differences", headers=headers).status_code == 200
    download = client.get(f"/api/v1/comparisons/{job_id}/artifacts/json", headers=headers, follow_redirects=False)
    assert download.status_code == 307 and download.headers["location"].startswith(endpoint)

    cancelled = client.post("/api/v1/comparisons", headers={**headers, "Idempotency-Key": "infra-api-cancel"}, data={"schema_id": "infra-api", "schema_version": "1.0.0"}, files=files)
    cancelled_id = cancelled.json()["job_id"]
    assert client.post(f"/api/v1/comparisons/{cancelled_id}/cancel", headers=headers).status_code == 202
    SimpleWorker([api_module.task_queue.queue], connection=api_module.task_queue.connection).work(burst=True, logging_level="WARNING")
    assert client.get(f"/api/v1/comparisons/{cancelled_id}", headers=headers).json()["status"] == "cancelled"

    failed_files = {"excel_file": files["excel_file"], "standard_data": ("standard.json", b'{"data":[{"value":"missing-id"}]}', "application/json")}
    failed = client.post("/api/v1/comparisons", headers={**headers, "Idempotency-Key": "infra-api-failure"}, data={"schema_id": "infra-api", "schema_version": "1.0.0"}, files=failed_files)
    failed_id = failed.json()["job_id"]
    SimpleWorker([api_module.task_queue.queue], connection=api_module.task_queue.connection).work(burst=True, logging_level="WARNING")
    failed_status = client.get(f"/api/v1/comparisons/{failed_id}", headers=headers).json()
    assert failed_status["status"] == "failed" and failed_status["error_code"] == "STANDARD_DATA_INVALID"
    with Session(api_module.database.engine) as session:
        downloads = session.scalars(select(AuditEventRow).where(AuditEventRow.action == "comparison.artifact_downloaded", AuditEventRow.tenant_id == "integration-api")).all()
        assert downloads
