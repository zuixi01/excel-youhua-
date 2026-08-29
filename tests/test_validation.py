from datetime import date, datetime

import pytest
from openpyxl import Workbook
from openpyxl.utils.datetime import MAC_EPOCH, WINDOWS_EPOCH, to_excel

from excel_auditor.engine import compare_workbook
from excel_auditor.models import RuleSet
from excel_auditor.pandera_adapter import StandardDataValidator
from excel_auditor.service import _field_statistics
from excel_auditor.snapshots import SpilledRecords
from excel_auditor.workbook import inspect_workbook


@pytest.mark.parametrize("epoch", [WINDOWS_EPOCH, MAC_EPOCH])
def test_unformatted_excel_date_serials_use_the_workbook_epoch(tmp_path, epoch):
    rules = RuleSet.model_validate({
        "schema_id": "date-epoch", "schema_version": "1.0.0", "name": "Date epoch",
        "sheets": [{
            "id": "data", "name": "Data", "primary_key": ["day"],
            "columns": [
                {"name": "id", "title": "ID"},
                {"name": "day", "title": "Day", "type": "date", "required": True},
                {"name": "moment", "title": "Moment", "type": "datetime", "compare": {"mode": "datetime", "timezone": "UTC"}},
            ],
        }],
    })
    path = tmp_path / f"date-epoch-{epoch.year}.xlsx"
    expected_day = date(2024, 1, 15)
    expected_moment = datetime(2024, 1, 15, 12, 30)
    book = Workbook()
    book.epoch = epoch
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID", "Day", "Moment"])
    sheet.append(["E1", to_excel(expected_day, epoch), to_excel(expected_moment, epoch)])
    sheet["B2"].number_format = "General"
    sheet["C2"].number_format = "General"
    book.save(path)

    snapshot = inspect_workbook(path, rules)
    assert snapshot.excel_epoch == epoch
    result = compare_workbook(snapshot, {"data": [{
        "id": "E1", "day": "2024-01-15", "moment": "2024-01-15T12:30:00Z",
    }]}, rules)
    assert result.summary.matched_records == 1
    assert result.summary.differences == 0
    snapshot.close()


def test_non_default_sheet_actions_are_honored(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "actions", "schema_version": "1.0.0", "name": "Actions",
        "sheets": [{
            "id": "data", "name": "Data", "primary_key": ["id"],
            "actions": {"extra_header": "report_only", "extra_record": "report_only", "mismatched_value": "report_only", "invalid_value": "report_only", "duplicate_key": "report_only"},
            "columns": [
                {"name": "id", "title": "ID", "required": True},
                {"name": "amount", "title": "Amount", "type": "integer"},
            ],
        }],
    })
    path = tmp_path / "actions.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID", "Amount", "Extra"])
    sheet.append(["E1", "bad", "x"])
    sheet.append(["E2", "2", "y"])
    book.save(path)
    result = compare_workbook(inspect_workbook(path, rules), {"data": [{"id": "E1", "amount": 1}]}, rules)
    configurable = {
        "EXTRA_HEADER", "EXTRA_RECORD", "VALUE_MISMATCH", "INVALID_VALUE", "VALIDATION_ERROR", "DUPLICATE_PRIMARY_KEY"
    }
    assert all(item.render_action == "report_only" for item in result.differences if item.type.value in configurable)


def test_header_auto_detection_and_order_mismatch(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "header-detect", "schema_version": "1.0.0", "name": "Header detect",
        "sheets": [{
            "id": "data", "name": "Data", "header": {"row": 1, "auto_detect": True}, "primary_key": ["id"],
            "columns": [
                {"name": "id", "title": "ID", "required": True},
                {"name": "name", "title": "Name"},
            ],
        }],
    })
    path = tmp_path / "header-detect.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["Report title"])
    sheet.append(["Name", "ID"])
    sheet.append(["Alice", "E1"])
    book.save(path)
    result = compare_workbook(inspect_workbook(path, rules), {"data": [{"id": "E1", "name": "Alice"}]}, rules)
    assert not result.manual_review_reasons
    assert any(item.type.value == "HEADER_ORDER_MISMATCH" for item in result.differences)
    assert result.summary.matched_records == 1


def test_ambiguous_auto_detected_header_requires_review(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "header-ambiguous", "schema_version": "1.0.0", "name": "Header ambiguous",
        "sheets": [{"id": "data", "name": "Data", "header": {"auto_detect": True}, "primary_key": ["id"], "columns": [{"name": "id", "title": "ID", "required": True}]}],
    })
    path = tmp_path / "header-ambiguous.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID"])
    sheet.append(["ID"])
    book.save(path)
    result = compare_workbook(inspect_workbook(path, rules), {"data": []}, rules)
    assert result.manual_review_reasons
    assert [item.type.value for item in result.differences] == ["HEADER_NOT_FOUND"]


def test_auto_detected_header_must_precede_explicit_data_region(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "header-data-overlap", "schema_version": "1.0.0", "name": "Header overlap",
        "sheets": [{
            "id": "data", "name": "Data", "header": {"auto_detect": True},
            "data_region": {"start_row": 3}, "primary_key": ["id"],
            "columns": [{"name": "id", "title": "ID", "required": True}],
        }],
    })
    path = tmp_path / "header-data-overlap.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["Report title"])
    sheet.append(["More preamble"])
    sheet.append(["ID"])
    sheet.append(["E1"])
    book.save(path)

    result = compare_workbook(inspect_workbook(path, rules), {"data": [{"id": "E1"}]}, rules)
    assert result.manual_review_reasons == ["Data: data_region_overlaps_header"]
    assert [item.type.value for item in result.differences] == ["HEADER_NOT_FOUND"]
    assert result.summary.matched_records == 0


def test_unmatched_records_still_receive_field_and_cross_field_validation(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "quality", "schema_version": "1.0.0", "name": "Quality",
        "sheets": [{
            "id": "staff", "name": "员工", "primary_key": ["id"],
            "columns": [
                {"name": "id", "title": "编号", "required": True},
                {"name": "email", "title": "邮箱", "validation": {"regex": "^[^@]+@[^@]+$", "unique": True}},
                {"name": "status", "title": "状态"},
                {"name": "end_date", "title": "离职日期", "type": "date"},
            ],
            "cross_field_rules": [{"rule_id": "end-required", "validator": "conditional_required", "params": {"when_field": "status", "equals": "离职", "required_field": "end_date"}}],
        }],
    })
    path = tmp_path / "quality.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "员工"
    sheet.append(["编号", "邮箱", "状态", "离职日期"])
    sheet.append(["E001", "bad", "离职", None])
    sheet.append(["E002", "bad", "在职", None])
    book.save(path)
    result = compare_workbook(inspect_workbook(path, rules), {"staff": []}, rules)
    rule_ids = [item.rule_id for item in result.differences]
    assert "email.validation" in rule_ids
    assert rule_ids.count("email.unique") == 2
    assert "end-required" in rule_ids


def test_fuzzy_string_is_suggestion_only_even_when_overwrite_is_enabled(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "fuzzy", "schema_version": "1.0.0", "name": "Fuzzy",
        "sheets": [{"id": "data", "name": "数据", "primary_key": ["id"], "actions": {"overwrite_mismatch": True}, "columns": [
            {"name": "id", "title": "编号", "required": True},
            {"name": "address", "title": "地址", "type": "fuzzy_string"},
        ]}],
    })
    path = tmp_path / "fuzzy.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "数据"
    sheet.append(["编号", "地址"])
    sheet.append(["E1", "上海市浦东新区"])
    book.save(path)
    result = compare_workbook(inspect_workbook(path, rules), {"data": [{"id": "E1", "address": "上海浦东新区"}]}, rules)
    difference = next(item for item in result.differences if item.canonical_field == "address")
    assert difference.severity == "warning" and difference.repair_status == "not_requested"
    assert not result.repairs


def test_omitted_optional_standard_field_never_clears_excel_but_explicit_null_can(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "omitted-standard", "schema_version": "1.0.0", "name": "Omitted standard",
        "sheets": [{"id": "data", "name": "Data", "primary_key": ["id"],
            "actions": {"overwrite_mismatch": True},
            "columns": [
                {"name": "id", "title": "ID", "required": True},
                {"name": "note", "title": "Note"},
            ],
        }],
    })
    path = tmp_path / "omitted-standard.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID", "Note"])
    sheet.append(["E1", "keep me"])
    book.save(path)

    omitted = compare_workbook(inspect_workbook(path, rules), {"data": [{"id": "E1"}]}, rules)
    assert omitted.summary.matched_records == 1
    assert not [item for item in omitted.differences if item.canonical_field == "note"]
    assert not omitted.repairs

    explicit_null = compare_workbook(inspect_workbook(path, rules), {"data": [{"id": "E1", "note": None}]}, rules)
    note_difference = next(item for item in explicit_null.differences if item.canonical_field == "note")
    assert note_difference.type.value == "VALUE_MISMATCH"
    assert note_difference.repair_status == "planned"
    assert [(repair.type, repair.canonical_field, repair.value) for repair in explicit_null.repairs] == [
        ("set_cell", "note", None)
    ]


def test_empty_primary_key_policy_can_match_by_row_number_without_silent_default(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "empty-key", "schema_version": "1.0.0", "name": "Empty key",
        "sheets": [{"id": "data", "name": "Data", "primary_key": ["id"], "empty_primary_key_action": "use_row_number", "columns": [
            {"name": "id", "title": "ID", "required": True}, {"name": "value", "title": "Value"},
        ]}],
    })
    path = tmp_path / "empty-key.xlsx"
    book = Workbook(); sheet = book.active; sheet.title = "Data"; sheet.append(["ID", "Value"]); sheet.append([None, "A"]); sheet.append([None, "B"]); book.save(path)
    standard = {"data": [{"id": None, "value": "A"}, {"id": None, "value": "B"}]}
    StandardDataValidator().validate(standard, rules)
    result = compare_workbook(inspect_workbook(path, rules), standard, rules)
    assert result.summary.matched_records == 2
    assert all(item.type.value != "EMPTY_PRIMARY_KEY" for item in result.differences)


def test_empty_primary_key_policy_can_explicitly_skip_rows(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "skip-key", "schema_version": "1.0.0", "name": "Skip key",
        "sheets": [{"id": "data", "name": "Data", "primary_key": ["id"], "empty_primary_key_action": "skip_row", "columns": [
            {"name": "id", "title": "ID", "required": True},
            {"name": "required_value", "title": "Required value", "required": True},
        ]}],
    })
    path = tmp_path / "skip-key.xlsx"
    book = Workbook(); sheet = book.active; sheet.title = "Data"; sheet.append(["ID", "Required value"]); sheet.append([None, "ignored"]); book.save(path)
    standard = {"data": [{"id": None}]}
    StandardDataValidator().validate(standard, rules)
    result = compare_workbook(inspect_workbook(path, rules), standard, rules)
    assert result.summary.differences == 0 and result.summary.matched_records == 0


def test_row_number_fallback_only_relaxes_the_empty_business_key():
    rules = RuleSet.model_validate({
        "schema_id": "row-fallback-validation", "schema_version": "1.0.0", "name": "Fallback validation",
        "sheets": [{"id": "data", "name": "Data", "primary_key": ["id"], "empty_primary_key_action": "use_row_number", "columns": [
            {"name": "id", "title": "ID", "required": True},
            {"name": "amount", "title": "Amount", "type": "integer", "required": True},
        ]}],
    })
    validator = StandardDataValidator()
    validator.validate({"data": [{"id": None, "amount": "1"}]}, rules)
    with pytest.raises(ValueError, match="typed failures.*amount"):
        validator.validate({"data": [{"id": None, "amount": "not-an-integer"}]}, rules)


def test_row_number_primary_key_never_appends_to_a_different_physical_row(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "row-number-append", "schema_version": "1.0.0", "name": "Row append",
        "sheets": [{
            "id": "data", "name": "Data", "primary_key_mode": "row_number",
            "actions": {"missing_record": "append_and_mark_green"},
            "columns": [{"name": "value", "title": "Value"}],
        }],
    })
    path = tmp_path / "row-number-append.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["Value"])
    sheet.append(["A"])
    book.save(path)
    standard = {"data": [
        {"__row_number__": 2, "value": "A"},
        {"__row_number__": 5, "value": "must-not-be-written-to-row-3"},
    ]}

    result = compare_workbook(inspect_workbook(path, rules), standard, rules)
    missing = next(item for item in result.differences if item.type.value == "MISSING_RECORD")
    assert missing.business_key == {"__row_number__": 5}
    assert missing.render_action == "report_only"
    assert missing.repair_status == "not_requested"
    assert "物理行不一致" in missing.message
    assert not result.repairs


@pytest.mark.parametrize("invalid_row", [True, False, 2.5, "2.0", " 2 ", 0, -1, 1_048_577])
def test_row_number_primary_key_rejects_lossy_or_out_of_range_values(invalid_row):
    rules = RuleSet.model_validate({
        "schema_id": "strict-row-number", "schema_version": "1.0.0", "name": "Strict row number",
        "sheets": [{
            "id": "data", "name": "Data", "primary_key_mode": "row_number",
            "columns": [{"name": "value", "title": "Value"}],
        }],
    })
    validator = StandardDataValidator()
    with pytest.raises(ValueError, match="invalid __row_number__"):
        validator.validate({"data": [{"__row_number__": invalid_row, "value": "A"}]}, rules)


@pytest.mark.parametrize("valid_row", [2, "2", "0002"])
def test_row_number_primary_key_accepts_exact_integer_csv_representations(valid_row):
    rules = RuleSet.model_validate({
        "schema_id": "valid-row-number", "schema_version": "1.0.0", "name": "Valid row number",
        "sheets": [{
            "id": "data", "name": "Data", "primary_key_mode": "row_number",
            "columns": [{"name": "value", "title": "Value"}],
        }],
    })
    StandardDataValidator().validate({"data": [{"__row_number__": valid_row, "value": "A"}]}, rules)


def test_standard_validation_is_chunked_and_checks_uniqueness_across_chunks():
    rules = RuleSet.model_validate({
        "schema_id": "chunked-standard",
        "schema_version": "1.0.0",
        "name": "Chunked standard",
        "sheets": [{
            "id": "data",
            "name": "Data",
            "primary_key": ["id"],
            "columns": [
                {"name": "id", "title": "ID", "required": True},
                {"name": "code", "title": "Code", "validation": {"unique": True}},
            ],
        }],
    })
    rows = SpilledRecords()
    for record in ({"id": "E1", "code": "A"}, {"id": "E2", "code": "B"}, {"id": "E3", "code": "A"}):
        rows.append(record)
    try:
        with pytest.raises(ValueError, match="is not unique at record 3"):
            StandardDataValidator(chunk_size=2).validate({"data": rows}, rules)
    finally:
        rows.close()


@pytest.mark.parametrize("equivalent_values", [("1.0", "1.00"), ("0", "-0")])
def test_standard_decimal_uniqueness_uses_normalized_numeric_equality(equivalent_values):
    rules = RuleSet.model_validate({
        "schema_id": "decimal-unique", "schema_version": "1.0.0", "name": "Decimal unique",
        "sheets": [{
            "id": "data", "name": "Data", "primary_key": ["id"],
            "columns": [
                {"name": "id", "title": "ID", "required": True},
                {"name": "amount", "title": "Amount", "type": "decimal", "validation": {"unique": True}},
            ],
        }],
    })
    rows = [
        {"id": "E1", "amount": equivalent_values[0]},
        {"id": "E2", "amount": equivalent_values[1]},
    ]

    with pytest.raises(ValueError, match=r"data\.amount is not unique at record 2"):
        StandardDataValidator(chunk_size=1).validate({"data": rows}, rules)


def test_datetime_uniqueness_distinguishes_dst_folds_but_matches_equal_instants():
    rules = RuleSet.model_validate({
        "schema_id": "datetime-unique", "schema_version": "1.0.0", "name": "Datetime unique",
        "sheets": [{
            "id": "data", "name": "Data", "primary_key": ["id"],
            "columns": [
                {"name": "id", "title": "ID", "required": True},
                {
                    "name": "event_at", "title": "Event", "type": "datetime",
                    "compare": {"mode": "datetime", "timezone": "America/New_York"},
                    "validation": {"unique": True},
                },
            ],
        }],
    })
    distinct_folds = [
        {"id": "E1", "event_at": "2024-11-03T01:30:00-04:00"},
        {"id": "E2", "event_at": "2024-11-03T01:30:00-05:00"},
    ]
    StandardDataValidator(chunk_size=1).validate({"data": distinct_folds}, rules)

    equal_instants = [
        {"id": "E1", "event_at": "2024-11-03T01:30:00-04:00"},
        {"id": "E2", "event_at": "2024-11-03T05:30:00Z"},
    ]
    with pytest.raises(ValueError, match=r"data\.event_at is not unique at record 2"):
        StandardDataValidator(chunk_size=1).validate({"data": equal_instants}, rules)


def test_excel_datetime_uniqueness_distinguishes_dst_folds(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "excel-datetime-unique", "schema_version": "1.0.0", "name": "Excel datetime unique",
        "sheets": [{
            "id": "data", "name": "Data", "primary_key": ["id"],
            "columns": [
                {"name": "id", "title": "ID", "required": True},
                {
                    "name": "event_at", "title": "Event", "type": "datetime",
                    "compare": {"mode": "datetime", "timezone": "America/New_York"},
                    "validation": {"unique": True},
                },
            ],
        }],
    })
    path = tmp_path / "dst-fold-unique.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID", "Event"])
    sheet.append(["E1", "2024-11-03T01:30:00-04:00"])
    sheet.append(["E2", "2024-11-03T01:30:00-05:00"])
    book.save(path)

    snapshot = inspect_workbook(path, rules)
    result = compare_workbook(snapshot, {"data": [
        {"id": "E1", "event_at": "2024-11-03T01:30:00-04:00"},
        {"id": "E2", "event_at": "2024-11-03T01:30:00-05:00"},
    ]}, rules)
    try:
        assert not any(item.rule_id == "event_at.unique" for item in result.differences)
    finally:
        result.close()
        snapshot.close()


def test_standard_json_field_rejects_duplicate_object_keys():
    rules = RuleSet.model_validate({
        "schema_id": "strict-json-field", "schema_version": "1.0.0", "name": "Strict JSON field",
        "sheets": [{
            "id": "data", "name": "Data", "primary_key": ["id"],
            "columns": [
                {"name": "id", "title": "ID", "required": True},
                {"name": "payload", "title": "Payload", "type": "json"},
            ],
        }],
    })

    with pytest.raises(ValueError, match=r"typed failures.*payload"):
        StandardDataValidator().validate({"data": [{"id": "E1", "payload": '{"a":1,"a":2}'}]}, rules)


def test_excel_json_field_reports_duplicate_object_keys_as_invalid(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "excel-strict-json", "schema_version": "1.0.0", "name": "Excel strict JSON",
        "sheets": [{
            "id": "data", "name": "Data", "primary_key": ["id"],
            "columns": [
                {"name": "id", "title": "ID", "required": True},
                {"name": "payload", "title": "Payload", "type": "json"},
            ],
        }],
    })
    path = tmp_path / "duplicate-json-key.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID", "Payload"])
    sheet.append(["E1", '{"a":1,"a":2}'])
    book.save(path)

    snapshot = inspect_workbook(path, rules)
    result = compare_workbook(snapshot, {"data": [{"id": "E1", "payload": {"a": 2}}]}, rules)
    try:
        invalid = [item for item in result.differences if item.rule_id == "payload.parse"]
        assert len(invalid) == 1
        assert invalid[0].type.value == "INVALID_VALUE"
        assert "duplicate object key" in invalid[0].message
    finally:
        result.close()
        snapshot.close()


def test_field_statistics_use_per_sheet_and_per_issue_denominators(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "field-statistics", "schema_version": "1.0.0", "name": "Field statistics",
        "sheets": [
            {"id": "a", "name": "A", "primary_key": ["id"], "columns": [
                {"name": "id", "title": "ID", "required": True},
                {"name": "value", "title": "Value", "type": "integer"},
            ]},
            {"id": "b", "name": "B", "primary_key": ["id"], "columns": [
                {"name": "id", "title": "ID", "required": True},
                {"name": "value", "title": "Value", "type": "integer"},
            ]},
        ],
    })
    path = tmp_path / "field-statistics.xlsx"
    book = Workbook()
    first = book.active
    first.title = "A"
    first.append(["ID", "Value"])
    first.append(["A1", 1])
    first.append(["A2", "bad"])
    second = book.create_sheet("B")
    second.append(["ID", "Value"])
    for index in range(1, 4):
        second.append([f"B{index}", index])
    book.save(path)
    standard = {
        "a": [{"id": "A1", "value": 2}],
        "b": [{"id": f"B{index}", "value": index} for index in range(1, 4)],
    }

    result = compare_workbook(inspect_workbook(path, rules), standard, rules)
    statistics = _field_statistics(result)

    assert result.summary.matched_records == 4
    assert statistics["a.value"] == {
        "sheet_id": "a", "canonical_field": "value",
        "compared_records": 1, "difference_count": 1, "difference_rate": 1.0,
        "validated_records": 2, "validation_error_count": 1, "validation_error_rate": 0.5,
    }
    assert statistics["b.value"]["compared_records"] == 3
    assert statistics["b.value"]["difference_count"] == 0
    assert statistics["b.value"]["difference_rate"] == 0.0
    assert statistics["b.value"]["validation_error_rate"] == 0.0
    result.report_only = True
    assert _field_statistics(result) == statistics
