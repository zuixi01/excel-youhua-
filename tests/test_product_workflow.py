from __future__ import annotations

from copy import deepcopy
import json
from datetime import datetime
from decimal import Decimal
from pathlib import Path

import pytest
from pydantic import ValidationError
from openpyxl import Workbook, load_workbook

from excel_auditor.models import ColumnRule, FieldType, ProductWorkflowConfig, RuleSet, ValidationConfig
from excel_auditor.product_workflow import (
    CatalogFieldDefinition,
    CatalogFieldSource,
    CatalogSchemaSnapshot,
    CategoryCatalogSnapshot,
    CategoryDefinition,
    InMemoryCatalogAdapter,
    ManagedHttpCatalogAdapter,
    ProductReviewDecision,
    build_dynamic_schema,
    map_product_headers,
    resolve_categories,
    normalize_product_workbook,
)
from excel_auditor.rules import load_rules
from excel_auditor.persistence import DatabaseRepository
from excel_auditor.service import AuditService
from excel_auditor.product_workflow.service import ProductWorkflowService
from excel_auditor.spill import SpillableSequence
from excel_auditor.workbook import SheetSnapshot, WorkbookSnapshot


def _platform_field(
    field_id: str,
    title: str,
    *,
    source: CatalogFieldSource = CatalogFieldSource.PLATFORM_ATTRIBUTE,
    order: int = 0,
    required: bool = False,
    aliases: list[str] | None = None,
) -> CatalogFieldDefinition:
    return CatalogFieldDefinition(
        field_id=field_id,
        title=title,
        source=source,
        display_order=order,
        required=required,
        category_id="cat-phone",
        attribute_id=field_id,
        aliases=aliases or [],
    )


def test_product_workflow_rule_configuration_is_strict_and_safe():
    payload = load_rules(Path("configs/examples/employee-roster.yaml")).model_dump(mode="json")
    payload["product_workflow"] = {
        "sheet_id": "employees",
        "catalog_connection_id": "platform-main",
        "category": {
            "source_field": "employee_name",
            "id_field": "employee_id",
            "category_list_path": "/api/categories",
            "attributes": {"path_template": "/api/categories/{category_id}/attributes"},
            "specifications": {"path_template": "/api/categories/{category_id}/specifications"},
        },
    }
    rules = RuleSet.model_validate(payload)
    assert rules.product_workflow is not None
    assert rules.product_workflow.output.merchant_extra_mode == "append_right"

    unsafe = deepcopy(payload)
    unsafe["product_workflow"]["category"]["attributes"]["path_template"] = "https://evil.invalid/{category_id}"
    with pytest.raises(ValidationError, match="normalized absolute path"):
        RuleSet.model_validate(unsafe)

    typo = deepcopy(payload)
    typo["product_workflow"]["output"] = {"merchant_extra_mod": "append_right"}
    with pytest.raises(ValidationError, match="Extra inputs are not permitted"):
        RuleSet.model_validate(typo)


def test_product_normalization_example_is_directly_loadable():
    rules = load_rules(Path("configs/examples/product-normalization.yaml"))
    assert rules.product_workflow is not None
    assert rules.product_workflow.sheet_id == "products"
    assert rules.product_workflow.category.category_list_pagination is not None


def test_header_mapping_accepts_only_exact_or_confirmed_aliases():
    fields = [
        _platform_field("brand", "品牌", aliases=["商品品牌"]),
        _platform_field("model", "型号"),
    ]
    mappings = map_product_headers(
        ["商品品牌", "型號", "内部备注"],
        fields,
        confirmed_aliases={"平台型号": "model"},
        fuzzy_threshold=50,
    )

    assert mappings[0].status == "accepted"
    assert mappings[0].field_id == "brand"
    assert mappings[1].match_type == "fuzzy_suggestion"
    assert mappings[1].status == "manual_review"
    assert mappings[1].field_id is None
    assert mappings[2].status in {"manual_review", "unmapped"}


def test_duplicate_header_never_silently_overwrites_first_mapping():
    fields = [_platform_field("brand", "品牌")]
    mappings = map_product_headers(["品牌", "品牌"], fields)
    assert mappings[0].status == "accepted"
    assert mappings[1].match_type == "duplicate"
    assert mappings[1].status == "manual_review"


def test_dynamic_schema_order_is_fixed_then_attributes_specs_and_merchant_extras():
    fixed = [
        ColumnRule(name="product_name", title="商品名称", required=True),
        ColumnRule(name="merchant_category", title="商家类目", required=True),
    ]
    platform = [
        _platform_field(
            "color",
            "颜色",
            source=CatalogFieldSource.PLATFORM_SPECIFICATION,
            order=2,
        ),
        _platform_field("brand", "品牌", order=20, required=True),
        _platform_field("material", "材质", order=10),
        _platform_field(
            "size",
            "尺码",
            source=CatalogFieldSource.PLATFORM_SPECIFICATION,
            order=1,
        ),
    ]
    headers = ["内部货号", "商品名称", "品牌", "商家类目", "采购备注"]
    targets = [
        CatalogFieldDefinition(
            field_id=column.name,
            title=column.title,
            aliases=column.aliases,
            source=CatalogFieldSource.FIXED,
            field_type=column.type,
            required=column.required,
        )
        for column in fixed
    ] + platform
    mappings = map_product_headers(headers, targets)
    plan = build_dynamic_schema(
        category_id="cat-phone",
        category_name="手机通讯",
        fixed_columns=fixed,
        platform_fields=platform,
        headers=headers,
        mappings=mappings,
    )

    fields = plan.fields
    assert [item.field.field_id for item in fields[:6]] == [
        "product_name",
        "merchant_category",
        "material",
        "brand",
        "size",
        "color",
    ]
    assert [item.field.title for item in fields[6:]] == ["内部货号", "采购备注"]
    assert all(item.field.source == CatalogFieldSource.MERCHANT_EXTRA for item in fields[6:])
    assert fields[0].present is True
    assert next(item for item in fields if item.field.field_id == "material").present is False


def test_category_resolution_is_deterministic_and_fuzzy_is_review_only():
    categories = [
        CategoryDefinition(category_id="100", name="手机", aliases=["智能手机"]),
        CategoryDefinition(category_id="200", name="手机配件"),
    ]
    rows = [
        {"platform_category_id": "100", "merchant_category": "任意文本"},
        {"merchant_category": "智能手机"},
        {"merchant_category": "手机配牛"},
        {},
    ]
    resolved = resolve_categories(rows, categories, fuzzy_threshold=60)
    assert resolved[0].status == "resolved" and resolved[0].match_type == "id"
    assert resolved[1].category_id == "100" and resolved[1].status == "resolved"
    assert resolved[2].status == "manual_review"
    assert resolved[2].category_id is None
    assert resolved[3].status == "unresolved"


def test_category_id_name_conflict_and_invalid_id_are_never_auto_resolved():
    categories = [
        CategoryDefinition(category_id="cat-a", name="Category A"),
        CategoryDefinition(category_id="cat-b", name="Category B"),
    ]

    resolved = resolve_categories([
        {"platform_category_id": "cat-a", "merchant_category": "Category B"},
        {"platform_category_id": "deleted-id", "merchant_category": "Category A"},
    ], categories)

    assert resolved[0].status == "manual_review"
    assert resolved[0].match_type == "id_name_conflict"
    assert {candidate.field_id for candidate in resolved[0].candidates} == {"cat-a", "cat-b"}
    assert resolved[1].status == "manual_review"
    assert resolved[1].match_type == "invalid_id"
    assert [candidate.field_id for candidate in resolved[1].candidates] == ["cat-a"]


def test_category_catalog_snapshot_is_order_independent_and_tamper_evident():
    categories = [
        CategoryDefinition(category_id="cat-b", name="Category B"),
        CategoryDefinition(category_id="cat-a", name="Category A"),
    ]
    snapshot = CategoryCatalogSnapshot.create(
        snapshot_id="pcat_snapshot",
        connection_id="platform-main",
        categories=categories,
    )
    reordered = CategoryCatalogSnapshot.create(
        snapshot_id="pcat_reordered",
        connection_id="platform-main",
        categories=list(reversed(categories)),
    )

    assert [category.category_id for category in snapshot.categories] == ["cat-a", "cat-b"]
    assert snapshot.content_sha256 == reordered.content_sha256
    tampered = snapshot.model_dump(mode="json")
    tampered["categories"][0]["name"] = "Changed"
    with pytest.raises(ValidationError, match="content hash"):
        CategoryCatalogSnapshot.model_validate(tampered)


def test_catalog_snapshot_hash_detects_field_changes():
    fields = [_platform_field("brand", "品牌")]
    snapshot = CatalogSchemaSnapshot.create(
        snapshot_id="01HPRODUCTSNAPSHOT",
        connection_id="platform-main",
        category_id="cat-phone",
        fields=fields,
    )
    assert len(snapshot.content_sha256) == 64
    tampered = snapshot.model_dump(mode="json")
    tampered["fields"][0]["title"] = "被篡改"
    with pytest.raises(ValidationError, match="content hash"):
        CatalogSchemaSnapshot.model_validate(tampered)


def test_catalog_field_contract_rejects_incomplete_platform_identity():
    with pytest.raises(ValidationError, match="require category_id and attribute_id"):
        CatalogFieldDefinition(
            field_id="brand",
            title="品牌",
            source=CatalogFieldSource.PLATFORM_ATTRIBUTE,
            field_type=FieldType.STRING,
        )


class _FakeManagedSource:
    def __init__(self) -> None:
        self.requests = []

    def fetch_with_metadata(self, config, _parameters=None):
        self.requests.append(config)
        if config.path == "/catalog/categories":
            return [{"id": "cat-phone", "name": "手机", "aliases": ["智能手机"]}], {"kind": "categories"}
        if config.path.endswith("/attributes"):
            return [{
                "id": "brand",
                "title": "品牌",
                "required": True,
                "type": "string",
                "display_order": 1,
            }], {"kind": "attributes"}
        if config.path.endswith("/specifications"):
            return [{
                "id": "color",
                "title": "颜色",
                "multiple": True,
                "type": "enum",
                "enum_values": ["黑色", "白色"],
                "display_order": 1,
            }], {"kind": "specifications"}
        raise AssertionError(config.path)


def test_managed_catalog_adapter_uses_safe_paths_and_snapshots_platform_contract():
    config = ProductWorkflowConfig.model_validate({
        "sheet_id": "products",
        "catalog_connection_id": "platform-main",
        "category": {
            "category_list_path": "/catalog/categories",
            "attributes": {"path_template": "/catalog/categories/{category_id}/attributes"},
            "specifications": {"path_template": "/catalog/categories/{category_id}/specifications"},
        },
    })
    source = _FakeManagedSource()
    adapter = ManagedHttpCatalogAdapter(source, config)

    categories = adapter.list_categories()
    snapshot = adapter.fetch_schema("cat-phone")

    assert categories[0].category_id == "cat-phone"
    assert [request.path for request in source.requests] == [
        "/catalog/categories",
        "/catalog/categories/cat-phone/attributes",
        "/catalog/categories/cat-phone/specifications",
    ]
    assert [field.source for field in snapshot.fields] == [
        CatalogFieldSource.PLATFORM_ATTRIBUTE,
        CatalogFieldSource.PLATFORM_SPECIFICATION,
    ]
    assert snapshot.fields[1].enum_values == ["黑色", "白色"]
    assert snapshot.source_metadata["attributes"]["kind"] == "attributes"


def test_managed_catalog_adapter_rejects_loose_platform_types():
    config = ProductWorkflowConfig.model_validate({
        "sheet_id": "products",
        "catalog_connection_id": "platform-main",
        "category": {
            "attributes": {"path_template": "/catalog/{category_id}/attributes"},
            "specifications": {"path_template": "/catalog/{category_id}/specifications"},
        },
    })

    class InvalidSource(_FakeManagedSource):
        def fetch_with_metadata(self, config, _parameters=None):
            if config.path.endswith("/attributes"):
                return [{"id": "brand", "title": "品牌", "required": "yes"}], {}
            return [], {}

    with pytest.raises(ValueError, match="field.required must be a boolean"):
        ManagedHttpCatalogAdapter(InvalidSource(), config).fetch_schema("cat-phone")


def test_managed_catalog_adapter_supports_explicit_platform_payload_mappings_and_pagination():
    config = ProductWorkflowConfig.model_validate({
        "sheet_id": "products",
        "catalog_connection_id": "platform-main",
        "category": {
            "category_list_path": "/platform/categories",
            "category_list_pagination": {"size": 100, "total_json_path": "$.total"},
            "record_mapping": {"id_key": "categoryCode", "name_key": "categoryLabel"},
            "attributes": {
                "path_template": "/platform/{category_id}/attributes",
                "record_mapping": {
                    "id_key": "attributeCode",
                    "title_key": "attributeLabel",
                    "required_key": "isRequired",
                    "display_order_key": "sortIndex",
                },
            },
            "specifications": {
                "path_template": "/platform/{category_id}/specifications",
                "record_mapping": {"id_key": "specCode", "title_key": "specLabel"},
            },
        },
    })

    class MappedSource:
        def fetch_with_metadata(self, source, _parameters=None):
            if source.path == "/platform/categories":
                assert source.pagination is not None and source.pagination.size == 100
                return [{"categoryCode": "phone", "categoryLabel": "手机"}], {}
            if source.path.endswith("/attributes"):
                return [{
                    "attributeCode": "brand",
                    "attributeLabel": "品牌",
                    "isRequired": True,
                    "sortIndex": 3,
                }], {}
            return [{"specCode": "color", "specLabel": "颜色"}], {}

    adapter = ManagedHttpCatalogAdapter(MappedSource(), config)
    assert adapter.list_categories()[0].category_id == "phone"
    schema = adapter.fetch_schema("phone")
    assert [(field.attribute_id, field.title, field.required, field.display_order) for field in schema.fields] == [
        ("brand", "品牌", True, 3),
        ("color", "颜色", False, 0),
    ]


def test_managed_catalog_adapter_maps_type_aliases_and_validation_contracts():
    config = ProductWorkflowConfig.model_validate({
        "sheet_id": "products",
        "catalog_connection_id": "platform-main",
        "category": {
            "attributes": {
                "path_template": "/catalog/{category_id}/attributes",
                "record_mapping": {"type_value_aliases": {"money": "decimal"}},
            },
            "specifications": {"path_template": "/catalog/{category_id}/specifications"},
        },
    })

    class ValidatedSource(_FakeManagedSource):
        def fetch_with_metadata(self, source, _parameters=None):
            if source.path.endswith("/attributes"):
                return [{
                    "id": "price",
                    "title": "Price",
                    "type": " MONEY ",
                    "required": True,
                    "nullable": False,
                    "unique": True,
                    "min": "0.01",
                    "max": 999.99,
                }], {}
            return [], {}

    field = ManagedHttpCatalogAdapter(ValidatedSource(), config).fetch_schema("cat-phone").fields[0]
    assert field.field_type == FieldType.DECIMAL
    assert field.validation == ValidationConfig(
        nullable=False,
        unique=True,
        min=Decimal("0.01"),
        max=Decimal("999.99"),
    )

    class InvalidConstraintSource(ValidatedSource):
        def fetch_with_metadata(self, source, _parameters=None):
            records, metadata = super().fetch_with_metadata(source, _parameters)
            if records:
                records[0]["unique"] = "yes"
            return records, metadata

    with pytest.raises(ValueError, match="field.unique must be a boolean"):
        ManagedHttpCatalogAdapter(InvalidConstraintSource(), config).fetch_schema("cat-phone")


def test_product_normalizer_splits_categories_preserves_extras_and_validates_values(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "products",
        "schema_version": "1.0.0",
        "name": "商品整理",
        "sheets": [{
            "id": "products",
            "name": "商品信息",
            "primary_key": ["product_id"],
            "columns": [
                {"name": "product_id", "title": "商品ID", "required": True},
                {"name": "platform_category_id", "title": "平台类目ID"},
                {"name": "merchant_category", "title": "商家类目", "required": True},
            ],
        }],
        "product_workflow": {
            "sheet_id": "products",
            "catalog_connection_id": "platform-main",
            "category": {
                "attributes": {"path_template": "/catalog/{category_id}/attributes"},
                "specifications": {"path_template": "/catalog/{category_id}/specifications"},
            },
            "minimum_category_confidence": 70,
        },
    })
    brand = _platform_field("brand", "品牌", order=1, required=True)
    price = CatalogFieldDefinition(
        field_id="price",
        title="价格",
        source=CatalogFieldSource.PLATFORM_ATTRIBUTE,
        field_type=FieldType.DECIMAL,
        category_id="cat-phone",
        attribute_id="price",
        display_order=2,
        validation=ValidationConfig(min=Decimal("0")),
    )
    color = CatalogFieldDefinition(
        field_id="color",
        title="颜色",
        source=CatalogFieldSource.PLATFORM_SPECIFICATION,
        field_type=FieldType.ENUM,
        category_id="cat-phone",
        attribute_id="color",
        multiple=True,
        enum_values=["黑色", "白色"],
    )
    size = CatalogFieldDefinition(
        field_id="size",
        title="尺寸",
        source=CatalogFieldSource.PLATFORM_SPECIFICATION,
        field_type=FieldType.ENUM,
        category_id="cat-phone",
        attribute_id="size",
        multiple=True,
        enum_values=["小", "大"],
    )
    catalog = InMemoryCatalogAdapter(
        [
            CategoryDefinition(category_id="cat-phone", name="手机", aliases=["智能手机"]),
            CategoryDefinition(category_id="cat-accessory", name="手机配件"),
        ],
        {"cat-phone": [brand, price, color, size], "cat-accessory": []},
    )
    headers = ["内部备注", "商品ID", "平台类目ID", "商家类目", "品牌", "价格", "颜色", "尺寸"]
    snapshot = WorkbookSnapshot(
        path=tmp_path / "products.xlsx",
        sha256="0" * 64,
        sheets={"商品信息": SheetSnapshot(
            name="商品信息",
            max_row=6,
            max_column=len(headers),
            rows=[
                (1, headers),
                (3, ["保留我", "P-1", "cat-phone", "智能手机", "", "12.50", "黑色、白色", "小、大"]),
                (5, ["也保留", "P-2", "", "未知手机类目", "某品牌", "-1", "黑色", "小"]),
                (6, ["高精度", "P-3", "cat-phone", "手机", "某品牌", "1234567890123456", "黑色", "小"]),
            ],
        )},
        excel_epoch=datetime(1899, 12, 30),
    )

    result = normalize_product_workbook(snapshot, rules, catalog)

    assert len(result.category_sheets) == 1
    product_sheet = result.category_sheets[0]
    assert product_sheet.source_excel_rows == [3, 6]
    assert [item.field.source for item in product_sheet.plan.fields][-1] == CatalogFieldSource.MERCHANT_EXTRA
    assert product_sheet.rows[0][product_sheet.plan.fields[-1].field.field_id] == "保留我"
    assert product_sheet.rows[0]["price"] == Decimal("12.50")
    assert product_sheet.rows[0]["color"] == "黑色、白色"
    assert [(row["color"], row["size"]) for row in product_sheet.sku_rows[:4]] == [
        ("黑色", "小"), ("黑色", "大"), ("白色", "小"), ("白色", "大"),
    ]
    assert product_sheet.sku_source_excel_rows[:4] == [3, 3, 3, 3]
    assert any(issue.field_id == "brand" and issue.issue_type == "required_missing" for issue in result.issues)
    assert any(
        issue.excel_row == 6 and issue.field_id == "price" and issue.issue_type == "excel_write_unsafe"
        for issue in result.issues
    )
    assert product_sheet.rows[1]["price"] == "1234567890123456"
    assert result.unresolved_rows[0].excel_row == 5
    assert result.requires_manual_review is True

    database = DatabaseRepository(f"sqlite:///{(tmp_path / 'workflow.db').as_posix()}")
    database.create_job("job_01PRODUCTWORKFLOW000000000", user_id="operator")
    database.save_product_category_snapshot(result.category_catalog_snapshot)
    for catalog_snapshot in result.catalog_snapshots:
        database.save_product_catalog_snapshot(catalog_snapshot)
    revision = database.create_product_revision(
        "job_01PRODUCTWORKFLOW000000000",
        result,
        actor_id="operator",
    )
    reviews = database.list_product_reviews("job_01PRODUCTWORKFLOW000000000")
    assert revision["revision_number"] == 1
    assert reviews and reviews[0]["status"] == "pending"
    decided = database.decide_product_review(
        "job_01PRODUCTWORKFLOW000000000",
        reviews[0]["review_id"],
        {"action": "confirm_category", "category_id": "cat-phone"},
        actor_id="operator",
    )
    assert decided["status"] == "resolved"
    with pytest.raises(ValueError, match="already been decided"):
        database.decide_product_review(
            "job_01PRODUCTWORKFLOW000000000",
            reviews[0]["review_id"],
            {"action": "confirm_category", "category_id": "cat-phone"},
            actor_id="operator",
        )


def test_product_normalizer_checks_keys_unique_fields_and_cross_field_rules_globally(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "product-quality",
        "schema_version": "1.0.0",
        "name": "Product quality",
        "sheets": [{
            "id": "products",
            "name": "Products",
            "primary_key": ["product_id"],
            "columns": [
                {"name": "product_id", "title": "Product ID", "required": True},
                {"name": "category_id", "title": "Category ID"},
                {"name": "category_name", "title": "Category"},
                {"name": "serial", "title": "Serial", "validation": {"unique": True}},
                {"name": "status", "title": "Status"},
                {"name": "end_date", "title": "End date"},
            ],
            "cross_field_rules": [{
                "rule_id": "end-required",
                "validator": "conditional_required",
                "params": {"when_field": "status", "equals": "inactive", "required_field": "end_date"},
            }],
        }],
        "product_workflow": {
            "sheet_id": "products",
            "catalog_connection_id": "platform-main",
            "category": {
                "source_field": "category_name",
                "id_field": "category_id",
                "attributes": {"path_template": "/catalog/{category_id}/attributes"},
                "specifications": {"path_template": "/catalog/{category_id}/specifications"},
            },
        },
    })
    platform_code_a = CatalogFieldDefinition(
        field_id="platform.platform_attribute.code",
        attribute_id="code",
        category_id="cat-a",
        title="Platform code",
        source=CatalogFieldSource.PLATFORM_ATTRIBUTE,
        validation=ValidationConfig(unique=True),
    )
    platform_code_b = platform_code_a.model_copy(update={"category_id": "cat-b"})
    catalog = InMemoryCatalogAdapter(
        [
            CategoryDefinition(category_id="cat-a", name="Category A"),
            CategoryDefinition(category_id="cat-b", name="Category B"),
        ],
        {"cat-a": [platform_code_a], "cat-b": [platform_code_b]},
    )
    headers = ["Product ID", "Category ID", "Category", "Serial", "Status", "End date", "Platform code"]
    snapshot = WorkbookSnapshot(
        path=tmp_path / "quality.xlsx",
        sha256="1" * 64,
        sheets={"Products": SheetSnapshot(
            name="Products",
            max_row=4,
            max_column=len(headers),
            rows=[
                (1, headers),
                (2, ["P-1", "cat-a", "Category A", "S-1", "inactive", None, "X-1"]),
                (3, ["P-1", "cat-b", "Category B", "S-1", "active", None, "X-1"]),
                (4, ["P-2", "cat-a", "Category A", "S-2", "active", None, "X-1"]),
            ],
        )},
        excel_epoch=datetime(1899, 12, 30),
    )

    result = normalize_product_workbook(snapshot, rules, catalog)
    by_type = {}
    for issue in result.issues:
        by_type.setdefault(issue.issue_type, []).append(issue)

    assert {issue.excel_row for issue in by_type["duplicate_primary_key"]} == {2, 3}
    fixed_unique = [issue for issue in by_type["unique_value"] if issue.field_id == "serial"]
    platform_unique = [
        issue
        for issue in by_type["unique_value"]
        if issue.field_id == "platform.platform_attribute.code"
    ]
    assert {issue.excel_row for issue in fixed_unique} == {2, 3}
    assert {issue.excel_row for issue in platform_unique} == {2, 4}
    assert [(issue.excel_row, issue.field_id) for issue in by_type["cross_field_error"]] == [(2, "end_date")]


def test_product_workflow_service_renders_category_and_sku_workbooks(tmp_path):
    golden = json.loads(Path("tests/golden_files/product_workflow_expected.json").read_text(encoding="utf-8"))
    rules = RuleSet.model_validate({
        "schema_id": "product-output",
        "schema_version": "1.0.0",
        "name": "商品输出",
        "sheets": [{
            "id": "products",
            "name": "商品信息",
            "primary_key": ["product_id"],
            "columns": [
                {"name": "product_id", "title": "商品ID", "required": True},
                {"name": "platform_category_id", "title": "平台类目ID"},
                {"name": "merchant_category", "title": "商家类目", "required": True},
            ],
        }],
        "product_workflow": {
            "sheet_id": "products",
            "catalog_connection_id": "platform-main",
            "category": {
                "attributes": {"path_template": "/catalog/{category_id}/attributes"},
                "specifications": {"path_template": "/catalog/{category_id}/specifications"},
            },
        },
    })
    brand = _platform_field("brand", "品牌", required=True)
    color = CatalogFieldDefinition(
        field_id="color",
        title="颜色",
        source=CatalogFieldSource.PLATFORM_SPECIFICATION,
        field_type=FieldType.ENUM,
        category_id="cat-phone",
        attribute_id="color",
        multiple=True,
        enum_values=["黑色", "白色"],
    )
    catalog = InMemoryCatalogAdapter(
        [CategoryDefinition(category_id="cat-phone", name="手机")],
        {"cat-phone": [brand, color]},
    )
    source = tmp_path / "source.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "商品信息"
    sheet.append(["内部备注", "商品ID", "平台类目ID", "商家类目", "品牌", "颜色"])
    sheet.append(["保留", "P-1", "cat-phone", "手机", "某品牌", "黑色、白色"])
    book.save(source)
    audit = AuditService(tmp_path / "runtime")
    workflow = ProductWorkflowService(audit)
    job_id = audit.create_job()
    job_source = audit.job_directory(job_id) / "product-input.xlsx"
    job_source.write_bytes(source.read_bytes())

    workflow.run(job_id, job_source, rules, catalog)

    status = audit.status(job_id)
    assert status["status"] == "completed", status
    output = load_workbook(audit.artifact(job_id, "product_excel"), data_only=False)
    actual = {}
    for key in ("category_sheet", "sku_sheet"):
        expected = golden[key]
        rendered_sheet = output[expected["name"]]
        actual[key] = {
            "name": rendered_sheet.title,
            "headers": [cell.value for cell in rendered_sheet[1]],
            "rows": [list(row) for row in rendered_sheet.iter_rows(min_row=2, values_only=True)],
            "header_colors": [cell.fill.fgColor.rgb[-6:] for cell in rendered_sheet[1]],
            "freeze_panes": str(rendered_sheet.freeze_panes),
            "auto_filter": rendered_sheet.auto_filter.ref,
        }
    metadata = output[golden["metadata_sheet"]["name"]]
    actual["metadata_sheet"] = {"name": metadata.title, "state": metadata.sheet_state}
    assert actual == golden
    assert any(item.type == "list" for item in output[golden["category_sheet"]["name"]].data_validations.dataValidation)
    validation_sheets = [sheet for sheet in output.worksheets if sheet.title.startswith("__ExcelAuditorLists")]
    assert len(validation_sheets) == 1 and validation_sheets[0].sheet_state == "veryHidden"
    output.close()


def test_review_decision_creates_a_new_revision_without_auto_accepting_fuzzy_mapping(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "product-review",
        "schema_version": "1.0.0",
        "name": "商品审核",
        "sheets": [{
            "id": "products",
            "name": "商品信息",
            "primary_key": ["product_id"],
            "columns": [
                {"name": "product_id", "title": "商品ID", "required": True},
                {"name": "platform_category_id", "title": "平台类目ID"},
                {"name": "merchant_category", "title": "商家类目", "required": True},
            ],
        }],
        "product_workflow": {
            "sheet_id": "products",
            "catalog_connection_id": "platform-main",
            "mapping_fuzzy_threshold": 50,
            "category": {
                "attributes": {"path_template": "/catalog/{category_id}/attributes"},
                "specifications": {"path_template": "/catalog/{category_id}/specifications"},
            },
        },
    })
    catalog = InMemoryCatalogAdapter(
        [CategoryDefinition(category_id="cat-phone", name="手机")],
        {"cat-phone": [_platform_field("brand", "品牌", required=True)]},
    )
    source = tmp_path / "review.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "商品信息"
    sheet.append(["商品ID", "平台类目ID", "商家类目", "品脾"])
    sheet.append(["P-1", "cat-phone", "手机", "某品牌"])
    book.save(source)
    audit = AuditService(tmp_path / "review-runtime")
    workflow = ProductWorkflowService(audit)
    job_id = audit.create_job()
    job_source = audit.job_directory(job_id) / "product-input.xlsx"
    job_source.write_bytes(source.read_bytes())

    workflow.run(job_id, job_source, rules, catalog)

    assert audit.status(job_id)["status"] == "manual_review"
    issue_lines = audit.artifact(job_id, "product_issues").read_text(encoding="utf-8").splitlines()
    assert issue_lines and json.loads(issue_lines[0])["issue_type"] == "required_missing"
    review_output = load_workbook(audit.artifact(job_id, "product_excel"), data_only=False)
    assert "问题清单" in review_output.sheetnames
    review_output.close()
    reviews = workflow.list_reviews(job_id)
    mapping_review = next(item for item in reviews if item["review_type"] == "field_mapping")
    assert mapping_review["status"] == "pending"
    workflow.decide_review(
        job_id,
        mapping_review["review_id"],
        ProductReviewDecision(
            action="confirm_mapping",
            field_id="brand",
            raw_header="品脾",
        ),
        actor_id="operator",
    )
    frozen_category_hash = json.loads(
        (audit.job_directory(job_id) / "product-result-r1.json").read_text(encoding="utf-8")
    )["category_catalog_snapshot"]["content_sha256"]
    catalog._categories = [CategoryDefinition(category_id="cat-replacement", name="Replacement")]
    catalog._fields = {"cat-replacement": []}
    revision_lock = audit.job_directory(job_id) / ".revision.lock"
    revision_lock.mkdir()
    with pytest.raises(ValueError, match="already running"):
        workflow.rerun_after_reviews(job_id, rules, actor_id="operator")
    revision_lock.rmdir()
    workflow.rerun_after_reviews(job_id, rules, actor_id="operator")

    status = audit.status(job_id)
    assert status["status"] == "completed", status
    state = json.loads((audit.job_directory(job_id) / "product-workflow.json").read_text(encoding="utf-8"))
    assert [item["revision_number"] for item in state["revision_history"]] == [1, 2]
    rerun_result = json.loads(
        (audit.job_directory(job_id) / "product-result-r2.json").read_text(encoding="utf-8")
    )
    assert rerun_result["category_catalog_snapshot"]["content_sha256"] == frozen_category_hash
    assert rerun_result["category_sheets"][0]["category_id"] == "cat-phone"
    output = load_workbook(audit.artifact(job_id, "product_excel"), data_only=False)
    assert output["手机"]["D2"].value == "某品牌"
    output.close()


def test_product_workflow_honors_cancel_and_processing_timeout_before_work_starts(tmp_path, monkeypatch):
    rules = load_rules(Path("configs/examples/product-normalization.yaml"))
    catalog = InMemoryCatalogAdapter([], {})

    cancelled_audit = AuditService(tmp_path / "cancel-runtime")
    cancelled_workflow = ProductWorkflowService(cancelled_audit)
    cancelled_job = cancelled_audit.create_job()
    cancelled_audit.request_cancel(cancelled_job)
    cancelled_workflow.run(cancelled_job, tmp_path / "missing.xlsx", rules, catalog)
    assert cancelled_audit.status(cancelled_job)["status"] == "cancelled"

    timed_audit = AuditService(tmp_path / "timeout-runtime")
    timed_workflow = ProductWorkflowService(timed_audit)
    timed_job = timed_audit.create_job()
    clock = iter([0.0, float(rules.workbook.processing_timeout_seconds) + 1.0])
    monkeypatch.setattr("excel_auditor.product_workflow.service.time.monotonic", lambda: next(clock))
    timed_workflow.run(timed_job, tmp_path / "missing.xlsx", rules, catalog)
    status = timed_audit.status(timed_job)
    assert status["status"] == "failed"
    assert status["error_code"] == "PRODUCT_PROCESSING_TIMEOUT"


def test_product_normalizer_spills_large_output_rows_and_serializes_losslessly(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "product-spill",
        "schema_version": "1.0.0",
        "name": "Product spill",
        "workbook": {"max_in_memory_cells": 10000},
        "sheets": [{
            "id": "products",
            "name": "Products",
            "primary_key": ["product_id"],
            "columns": [
                {"name": "product_id", "title": "Product ID", "required": True},
                {"name": "category_id", "title": "Category ID"},
                {"name": "category_name", "title": "Category"},
            ],
        }],
        "product_workflow": {
            "sheet_id": "products",
            "catalog_connection_id": "memory",
            "category": {
                "source_field": "category_name",
                "id_field": "category_id",
                "attributes": {"path_template": "/catalog/{category_id}/attributes"},
                "specifications": {"path_template": "/catalog/{category_id}/specifications"},
            },
        },
    })
    headers = ["Product ID", "Category ID", "Category", *[f"Extra {index}" for index in range(7)]]
    rows = [(1, headers)] + [
        (index + 2, [f"P-{index}", "cat-a", "Category A", *[f"V-{index}-{extra}" for extra in range(7)]])
        for index in range(1001)
    ]
    snapshot = WorkbookSnapshot(
        path=tmp_path / "spill.xlsx",
        sha256="2" * 64,
        sheets={"Products": SheetSnapshot("Products", len(rows), len(headers), rows)},
    )
    result = normalize_product_workbook(
        snapshot,
        rules,
        InMemoryCatalogAdapter([CategoryDefinition(category_id="cat-a", name="Category A")], {"cat-a": []}),
        forced_extra_columns=set(range(4, 11)),
    )
    try:
        assert isinstance(result.category_sheets[0].rows, SpillableSequence)
        assert result.category_sheets[0].rows.spilled is True
        assert len(result.model_dump(mode="json")["category_sheets"][0]["rows"]) == 1001
    finally:
        result.close()


def test_product_workflow_respects_inspector_report_only_mode(tmp_path, monkeypatch):
    rules = RuleSet.model_validate({
        "schema_id": "product-report-only",
        "schema_version": "1.0.0",
        "name": "Product report only",
        "sheets": [{
            "id": "products",
            "name": "Products",
            "primary_key": ["product_id"],
            "columns": [
                {"name": "product_id", "title": "Product ID", "required": True},
                {"name": "category_id", "title": "Category ID"},
                {"name": "category_name", "title": "Category"},
            ],
        }],
        "product_workflow": {
            "sheet_id": "products",
            "catalog_connection_id": "memory",
            "category": {
                "source_field": "category_name",
                "id_field": "category_id",
                "attributes": {"path_template": "/catalog/{category_id}/attributes"},
                "specifications": {"path_template": "/catalog/{category_id}/specifications"},
            },
        },
    })
    snapshot = WorkbookSnapshot(
        path=tmp_path / "report-only.xlsx",
        sha256="3" * 64,
        sheets={"Products": SheetSnapshot(
            "Products",
            2,
            3,
            [(1, ["Product ID", "Category ID", "Category"]), (2, ["P-1", "cat-a", "Category A"])],
        )},
        report_only=True,
    )
    monkeypatch.setattr("excel_auditor.product_workflow.service.inspect_workbook", lambda *_args, **_kwargs: snapshot)
    source = tmp_path / "input.xlsx"
    source.write_bytes(b"not-read-because-inspector-is-stubbed")
    audit = AuditService(tmp_path / "report-only-runtime")
    job_id = audit.create_job()
    ProductWorkflowService(audit).run(
        job_id,
        source,
        rules,
        InMemoryCatalogAdapter([CategoryDefinition(category_id="cat-a", name="Category A")], {"cat-a": []}),
    )
    status = audit.status(job_id)
    assert status["status"] == "completed"
    assert status["report_only"] is True
    assert "product_excel" not in status["artifacts"]
