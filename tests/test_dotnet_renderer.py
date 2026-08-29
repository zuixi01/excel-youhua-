import hashlib
import json
import os
import subprocess
import zipfile
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from xml.etree import ElementTree

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.table import Table
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils.datetime import MAC_EPOCH, WINDOWS_EPOCH

from excel_auditor.models import RuleSet
from excel_auditor.rendering import DotNetOpenXmlRenderer
from excel_auditor.service import AuditService, _safe_error_code


def test_dotnet_renderer_self_check_contract():
    command = os.environ.get("EXCEL_RENDERER_COMMAND")
    if not command:
        pytest.skip("set EXCEL_RENDERER_COMMAND to run the .NET renderer contract test")
    assert DotNetOpenXmlRenderer(Path(command)).self_check() == "0.1.0"


@pytest.mark.parametrize(
    "error_code",
    ["ARGUMENT_INVALID", "MANIFEST_OR_STRUCTURE_INVALID", "OUTPUT_VERIFICATION_FAILED", "RENDER_FAILED", "UNSUPPORTED_FEATURE"],
)
def test_dotnet_renderer_adapter_preserves_structured_error_codes(tmp_path, monkeypatch, error_code):
    command = tmp_path / "renderer.exe"
    command.write_bytes(b"placeholder")
    renderer = DotNetOpenXmlRenderer(command)
    rules = RuleSet.model_validate({
        "schema_id": "renderer-errors", "schema_version": "1.0.0", "name": "Renderer errors",
        "sheets": [{
            "id": "data", "name": "Data", "primary_key": ["id"],
            "columns": [{"name": "id", "title": "ID", "required": True}],
        }],
    })
    monkeypatch.setattr(subprocess, "run", lambda *_args, **_kwargs: subprocess.CompletedProcess(
        args=[],
        returncode=1,
        stdout="",
        stderr=json.dumps({"success": False, "error_code": error_code, "message": "private detail"}),
    ))

    with pytest.raises(RuntimeError, match=f"^{error_code}:") as raised:
        renderer.render(tmp_path / "input.xlsx", tmp_path / "output.xlsx", None, rules, None, {})
    assert "private detail" not in str(raised.value)
    assert _safe_error_code(raised.value) == error_code


def test_dotnet_renderer_cli_rejects_invalid_arguments_and_same_path_without_mutation(tmp_path):
    command = os.environ.get("EXCEL_RENDERER_COMMAND")
    if not command:
        pytest.skip("set EXCEL_RENDERER_COMMAND to run the .NET renderer contract test")

    unknown = subprocess.run([command, "--unknown", "value"], capture_output=True, text=True, check=False)
    assert unknown.returncode != 0
    assert json.loads(unknown.stderr)["error_code"] == "ARGUMENT_INVALID"

    duplicate = subprocess.run(
        [command, "--input", "one.xlsx", "--input", "two.xlsx", "--output", "out.xlsx", "--manifest", "manifest.json"],
        capture_output=True,
        text=True,
        check=False,
    )
    assert duplicate.returncode != 0
    assert json.loads(duplicate.stderr)["error_code"] == "ARGUMENT_INVALID"

    source, manifest_path = tmp_path / "input.xlsx", tmp_path / "manifest.json"
    book = Workbook()
    book.active.title = "Data"
    book.active.append(["ID"])
    book.save(source)
    source_hash = hashlib.sha256(source.read_bytes()).hexdigest()
    manifest_path.write_text(json.dumps({
        "manifest_version": "1.0",
        "job_id": "job_same_path",
        "input_sha256": source_hash,
        "operations": [],
    }), encoding="utf-8")
    same_path = subprocess.run(
        [command, "--input", str(source), "--output", str(source), "--manifest", str(manifest_path)],
        capture_output=True,
        text=True,
        check=False,
    )
    assert same_path.returncode != 0
    assert json.loads(same_path.stderr)["error_code"] == "ARGUMENT_INVALID"
    assert hashlib.sha256(source.read_bytes()).hexdigest() == source_hash


def test_dotnet_renderer_marks_sparse_row_across_used_columns(tmp_path):
    command = os.environ.get("EXCEL_RENDERER_COMMAND")
    if not command:
        pytest.skip("set EXCEL_RENDERER_COMMAND to run the .NET renderer contract test")
    source, output, manifest_path = tmp_path / "sparse.xlsx", tmp_path / "sparse-output.xlsx", tmp_path / "manifest.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID", "Name", "Amount"])
    sheet.append(["E1"])
    book.save(source)
    manifest_path.write_text(json.dumps({
        "manifest_version": "1.0",
        "job_id": "job_sparse_row",
        "input_sha256": hashlib.sha256(source.read_bytes()).hexdigest(),
        "operations": [{"type": "mark_row", "sheet": "Data", "row": 2, "fill_color": "F4CCCC", "comment": "extra"}],
    }), encoding="utf-8")
    completed = subprocess.run([command, "--input", str(source), "--output", str(output), "--manifest", str(manifest_path)], capture_output=True, text=True, check=False)
    assert completed.returncode == 0, completed.stderr
    rendered = load_workbook(output)
    assert all(rendered["Data"].cell(2, column).fill.fgColor.rgb.endswith("F4CCCC") for column in range(1, 4))
    assert rendered["Data"].calculate_dimension() == "A1:C2"
    rendered.close()


def test_dotnet_renderer_validates_hash_marks_and_inserts(tmp_path):
    command = os.environ.get("EXCEL_RENDERER_COMMAND")
    if not command:
        pytest.skip("set EXCEL_RENDERER_COMMAND to run the .NET renderer contract test")
    source = tmp_path / "input.xlsx"
    output = tmp_path / "output.xlsx"
    manifest_path = tmp_path / "manifest.json"
    report_path = tmp_path / "report.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "人员信息"
    sheet.append(["员工编号", "工资"])
    sheet.append(["E001", "99"])
    sheet["B2"].comment = Comment("用户原批注", "Alice")
    book.save(source)
    digest = hashlib.sha256(source.read_bytes()).hexdigest()
    report_path.write_text(json.dumps({"summary": {"matched_records": 1}, "differences": [{"type": "VALUE_MISMATCH", "sheet_name": "人员信息", "cell": "B2", "canonical_field": "salary", "message": "值不一致"}]}, ensure_ascii=False), encoding="utf-8")
    manifest_path.write_text(json.dumps({
        "manifest_version": "1.0",
        "job_id": "job_contract",
        "input_sha256": digest,
        "operations": [
            {"type": "mark_cell", "sheet": "人员信息", "cell": "B2", "fill_color": "FFE599", "comment": "值不一致"},
            {"type": "insert_column", "sheet": "人员信息", "before": "B", "canonical_field": "name", "header_row": 1, "header_value": "姓名", "fill_color": "D9EAD3"},
            {"type": "add_or_replace_report_sheet", "name": "核验报告", "source_json": "report.json"},
        ],
    }, ensure_ascii=False), encoding="utf-8")
    completed = subprocess.run([command, "--input", str(source), "--output", str(output), "--manifest", str(manifest_path)], capture_output=True, text=True, check=False)
    assert completed.returncode == 0, completed.stderr
    result = json.loads(completed.stdout)
    assert result["success"] is True
    assert {item["difference_id"] for item in result["operation_results"] if item["difference_id"]} == set()
    rendered = load_workbook(output)
    assert rendered["人员信息"]["B1"].value == "姓名"
    assert rendered["人员信息"]["C2"].value == "99"
    assert rendered["人员信息"]["C2"].fill.fgColor.rgb.endswith("FFE599")
    assert rendered["人员信息"]["C2"].comment is not None
    assert rendered["人员信息"]["C2"].comment.author == "Alice"
    assert rendered["人员信息"]["C2"].comment.text == "用户原批注\n\n[Excel Auditor]\n值不一致"
    assert rendered["核验报告"]["A2"].value == "matched_records"
    first_style_count = len(rendered._cell_styles)
    first_fill_count = len(rendered._fills)
    rendered.close()

    repeated = subprocess.run([command, "--input", str(source), "--output", str(output), "--manifest", str(manifest_path)], capture_output=True, text=True, check=False)
    assert repeated.returncode == 0, repeated.stderr
    assert json.loads(repeated.stdout)["success"] is True
    rerendered = load_workbook(output)
    assert [sheet_name for sheet_name in rerendered.sheetnames if sheet_name == "核验报告"] == ["核验报告"]
    assert [sheet_name for sheet_name in rerendered.sheetnames if sheet_name == "__ExcelAuditorMetadata"] == ["__ExcelAuditorMetadata"]
    assert rerendered["人员信息"]["B1"].value == "姓名"
    assert rerendered["人员信息"]["C2"].comment.text == "用户原批注\n\n[Excel Auditor]\n值不一致"
    assert len(rerendered._cell_styles) == first_style_count
    assert len(rerendered._fills) == first_fill_count
    rerendered.close()

    dry_output = tmp_path / "dry-run.xlsx"
    dry_output.write_bytes(b"pre-existing-output")
    dry = subprocess.run([command, "--input", str(source), "--output", str(dry_output), "--manifest", str(manifest_path), "--dry-run"], capture_output=True, text=True, check=False)
    assert dry.returncode == 0, dry.stderr
    dry_result = json.loads(dry.stdout)
    assert dry_result["success"] is True and dry_result["dry_run"] is True
    assert dry_output.read_bytes() == b"pre-existing-output"

    chained_output = tmp_path / "chained-output.xlsx"
    chained_manifest = tmp_path / "chained-manifest.json"
    chained_manifest.write_text(json.dumps({
        "manifest_version": "1.0",
        "job_id": "job_contract_chained",
        "input_sha256": hashlib.sha256(output.read_bytes()).hexdigest(),
        "operations": [{"type": "mark_cell", "sheet": "人员信息", "cell": "C2", "fill_color": "FFE599", "comment": "值不一致"}],
    }, ensure_ascii=False), encoding="utf-8")
    chained = subprocess.run([command, "--input", str(output), "--output", str(chained_output), "--manifest", str(chained_manifest)], capture_output=True, text=True, check=False)
    assert chained.returncode == 0, chained.stderr
    chained_book = load_workbook(chained_output)
    chained_comment = chained_book["人员信息"]["C2"].comment
    assert chained_comment.author == "Alice"
    assert chained_comment.text.count("用户原批注") == 1
    assert chained_comment.text.count("[Excel Auditor]") == 1
    assert chained_comment.text.count("值不一致") == 1
    chained_book.close()

    long_source, long_output, long_manifest = tmp_path / "long-comment.xlsx", tmp_path / "long-comment-output.xlsx", tmp_path / "long-comment-manifest.json"
    long_book = Workbook()
    long_book.active.title = "Data"
    long_book.active["A1"] = "value"
    long_book.active["A1"].comment = Comment("U" * 32760, "Alice")
    long_book.save(long_source)
    long_manifest.write_text(json.dumps({
        "manifest_version": "1.0",
        "job_id": "job_long_comment",
        "input_sha256": hashlib.sha256(long_source.read_bytes()).hexdigest(),
        "operations": [{"type": "mark_cell", "sheet": "Data", "cell": "A1", "fill_color": "FFE599", "comment": "audit"}],
    }), encoding="utf-8")
    long_result = subprocess.run([command, "--input", str(long_source), "--output", str(long_output), "--manifest", str(long_manifest)], capture_output=True, text=True, check=False)
    assert long_result.returncode != 0
    assert json.loads(long_result.stderr)["error_code"] == "UNSUPPORTED_FEATURE"
    assert not long_output.exists()


def test_dotnet_renderer_preserves_existing_note_vml_geometry_during_mark_and_insert(tmp_path):
    command = os.environ.get("EXCEL_RENDERER_COMMAND")
    if not command:
        pytest.skip("set EXCEL_RENDERER_COMMAND to run the .NET renderer contract test")
    source, output, manifest_path = tmp_path / "custom-note.xlsx", tmp_path / "custom-note-output.xlsx", tmp_path / "manifest.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID", "Amount"])
    sheet.append(["E1", 99])
    sheet["A2"].comment = Comment("second note", "Bob")
    sheet["B2"].comment = Comment("user note", "Alice")
    book.save(source)

    customized = tmp_path / "customized-note.xlsx"
    with zipfile.ZipFile(source) as archive:
        members = [(info, archive.read(info.filename)) for info in archive.infolist()]
    vml_name = next(info.filename for info, _data in members if info.filename.endswith(".vml"))
    vml_namespace = "{urn:schemas-microsoft-com:vml}"
    excel_namespace = "{urn:schemas-microsoft-com:office:excel}"
    rewritten_members = []
    for info, data in members:
        if info.filename == vml_name:
            root = ElementTree.fromstring(data)
            shapes = root.findall(f".//{vml_namespace}shape")
            shape = next(item for item in shapes if item.find(f".//{excel_namespace}Row").text == "1" and item.find(f".//{excel_namespace}Column").text == "0")
            shape.set("style", "position:absolute;margin-left:17pt;margin-top:19pt;width:222pt;height:111pt;z-index:7;visibility:visible")
            shape.set("fillcolor", "#abcdef")
            client_data = shape.find(f".//{excel_namespace}ClientData")
            anchor = shape.find(f".//{excel_namespace}Anchor")
            row = shape.find(f".//{excel_namespace}Row")
            column = shape.find(f".//{excel_namespace}Column")
            assert client_data is not None and row is not None and column is not None
            if anchor is None:
                anchor = ElementTree.Element(f"{excel_namespace}Anchor")
                client_data.insert(2, anchor)
            anchor.text, row.text, column.text = "0, 7, 1, 3, 4, 9, 6, 4", "1", "0"
            data = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
        rewritten_members.append((info, data))
    with zipfile.ZipFile(customized, "w") as archive:
        for info, data in rewritten_members:
            archive.writestr(info, data)
    customized.replace(source)

    manifest_path.write_text(json.dumps({
        "manifest_version": "1.0",
        "job_id": "job_custom_note",
        "input_sha256": hashlib.sha256(source.read_bytes()).hexdigest(),
        "operations": [
            {"type": "mark_cell", "sheet": "Data", "cell": "A2", "fill_color": "FFE599", "comment": "audit note"},
            {"type": "insert_column", "sheet": "Data", "before": "A", "canonical_field": "name", "header_row": 1, "header_value": "Name", "fill_color": "D9EAD3", "comment": "inserted header"},
        ],
    }), encoding="utf-8")
    completed = subprocess.run(
        [command, "--input", str(source), "--output", str(output), "--manifest", str(manifest_path)],
        capture_output=True,
        text=True,
        check=False,
    )
    assert completed.returncode == 0, completed.stderr
    rendered = load_workbook(output)
    preserved_comment = rendered["Data"]["B2"].comment
    assert preserved_comment is not None
    assert preserved_comment.author == "Bob"
    assert preserved_comment.text == "second note\n\n[Excel Auditor]\naudit note"
    assert rendered["Data"]["C2"].comment.author == "Alice"
    assert rendered["Data"]["C2"].comment.text == "user note"
    assert rendered["Data"]["A1"].comment.text == "inserted header"
    rendered.close()

    with zipfile.ZipFile(output) as archive:
        root = ElementTree.fromstring(archive.read(vml_name))
    shapes = root.findall(f".//{vml_namespace}shape")
    assert len(shapes) == 3
    preserved_shape = next(shape for shape in shapes if shape.get("fillcolor") == "#abcdef")
    assert preserved_shape.get("style") == "position:absolute;margin-left:17pt;margin-top:19pt;width:222pt;height:111pt;z-index:7;visibility:visible"
    assert preserved_shape.get("fillcolor") == "#abcdef"
    assert preserved_shape.find(f".//{excel_namespace}Anchor").text == "1, 7, 1, 3, 5, 9, 6, 4"
    assert preserved_shape.find(f".//{excel_namespace}Row").text == "1"
    assert preserved_shape.find(f".//{excel_namespace}Column").text == "1"
    assert sum("width:108pt;height:59pt" in (shape.get("style") or "") for shape in shapes) == 1

    malformed_source, malformed_output = tmp_path / "malformed-note.xlsx", tmp_path / "malformed-note-output.xlsx"
    with zipfile.ZipFile(source) as archive:
        members = [(info, archive.read(info.filename)) for info in archive.infolist()]
    with zipfile.ZipFile(malformed_source, "w") as archive:
        for info, data in members:
            if info.filename == vml_name:
                malformed_root = ElementTree.fromstring(data)
                malformed_root.find(f".//{excel_namespace}Anchor").text = "not-a-valid-anchor"
                data = ElementTree.tostring(malformed_root, encoding="utf-8", xml_declaration=True)
            archive.writestr(info, data)
    malformed_manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    malformed_manifest["job_id"] = "job_malformed_note"
    malformed_manifest["input_sha256"] = hashlib.sha256(malformed_source.read_bytes()).hexdigest()
    manifest_path.write_text(json.dumps(malformed_manifest), encoding="utf-8")
    rejected = subprocess.run(
        [command, "--input", str(malformed_source), "--output", str(malformed_output), "--manifest", str(manifest_path)],
        capture_output=True,
        text=True,
        check=False,
    )
    assert rejected.returncode != 0
    assert json.loads(rejected.stderr)["error_code"] == "UNSUPPORTED_FEATURE"
    assert not malformed_output.exists()


def test_dotnet_renderer_writes_typed_cells_formats_validation_metadata_and_rejects_unknown_ops(tmp_path):
    command = os.environ.get("EXCEL_RENDERER_COMMAND")
    if not command:
        pytest.skip("set EXCEL_RENDERER_COMMAND to run the .NET renderer contract test")
    source, output, manifest_path = tmp_path / "typed.xlsx", tmp_path / "typed-output.xlsx", tmp_path / "manifest.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID", "Amount"])
    sheet.append(["E1", 1])
    sheet.add_table(Table(displayName="TypedTable", ref="A1:B2"))
    existing_validation = DataValidation(type="whole", operator="greaterThan", formula1="0")
    existing_validation.add("B2")
    sheet.add_data_validation(existing_validation)
    book.save(source)
    manifest = {
        "manifest_version": "1.0",
        "job_id": "job_typed",
        "input_sha256": hashlib.sha256(source.read_bytes()).hexdigest(),
        "metadata": {
            "schema_id": "typed",
            "schema_version": "1.0.0",
            "schema_sha256": "a" * 64,
            "standard_snapshot_id": "std_example",
            "standard_sha256": "b" * 64,
            "result_sha256": "c" * 64,
        },
        "operations": [
            {"type": "set_cell", "sheet": "Data", "cell": "B2", "value": "12.50", "field_type": "decimal", "number_format": "0.00", "fill_color": "D9EAD3"},
            {"type": "insert_column", "sheet": "Data", "before": "B", "canonical_field": "score", "header_row": 1, "header_value": "Score", "fill_color": "D9EAD3", "field_type": "decimal", "number_format": "0.00", "formula_template": '=IF(ROW()>0,"https://example.test/a|b","embedded ""quote""")', "validation": {"type": "decimal", "min": "0", "max": "100", "allow_blank": False}},
            {"type": "append_row", "sheet": "Data", "row": 3, "values": [{"cell": "A3", "value": "E2", "field_type": "string"}, {"cell": "C3", "value": "9.75", "field_type": "decimal", "number_format": "0.00"}], "fill_color": "D9EAD3"},
        ],
    }
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    completed = subprocess.run([command, "--input", str(source), "--output", str(output), "--manifest", str(manifest_path)], capture_output=True, text=True, check=False)
    assert completed.returncode == 0, completed.stderr
    rendered = load_workbook(output, data_only=False)
    data = rendered["Data"]
    assert data["B2"].value == '=IF(ROW()>0,"https://example.test/a|b","embedded ""quote""")'
    assert data["B2"].number_format == "0.00"
    assert data["C2"].value == 12.5 and data["C2"].data_type == "n"
    assert data["C3"].value == 9.75 and data["C3"].data_type == "n"
    assert data["C3"].number_format == "0.00"
    assert data.tables["TypedTable"].ref == "A1:C3"
    assert sorted(str(item.sqref) for item in data.data_validations.dataValidation) == ["B2:B3", "C2:C3"]
    assert data.calculate_dimension() == "A1:C3"
    metadata = rendered["__ExcelAuditorMetadata"]
    assert metadata.sheet_state == "veryHidden"
    assert metadata["B2"].value == "job_typed"
    assert metadata["A9"].value == "result_content_sha256"
    assert metadata["B9"].value == json.loads(completed.stdout)["result_content_sha256"]
    assert len(metadata["B9"].value) == 64
    rendered.close()

    manifest["operations"] = [{"type": "unrecognized_operation", "sheet": "Data"}]
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    rejected = subprocess.run([command, "--input", str(source), "--output", str(output), "--manifest", str(manifest_path)], capture_output=True, text=True, check=False)
    assert rejected.returncode != 0
    assert json.loads(rejected.stderr)["error_code"] == "MANIFEST_OR_STRUCTURE_INVALID"


@pytest.mark.parametrize(
    ("epoch", "field_type", "written_value", "expected_serial", "expected_datetime"),
    [
        (WINDOWS_EPOCH, "date", "1900-01-01", 1.0, "1900-01-01T00:00:00"),
        (MAC_EPOCH, "date", "2026-08-30", 44802.0, "2026-08-30T00:00:00"),
        (MAC_EPOCH, "datetime", "2026-08-30T12:00:00+08:00", 44802.5, "2026-08-30T12:00:00"),
    ],
)
def test_dotnet_renderer_writes_dates_using_the_workbook_epoch(tmp_path, epoch, field_type, written_value, expected_serial, expected_datetime):
    command = os.environ.get("EXCEL_RENDERER_COMMAND")
    if not command:
        pytest.skip("set EXCEL_RENDERER_COMMAND to run the .NET renderer contract test")
    source, output, manifest_path = tmp_path / "dates.xlsx", tmp_path / "dates-output.xlsx", tmp_path / "manifest.json"
    book = Workbook()
    book.epoch = epoch
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["Date"])
    sheet.append([date(2020, 1, 1)])
    book.save(source)
    manifest_path.write_text(json.dumps({
        "manifest_version": "1.0",
        "job_id": "job_date_epoch",
        "input_sha256": hashlib.sha256(source.read_bytes()).hexdigest(),
        "operations": [{
            "type": "set_cell",
            "sheet": "Data",
            "cell": "A2",
            "value": written_value,
            "field_type": field_type,
            "number_format": "yyyy-mm-dd" if field_type == "date" else "yyyy-mm-dd hh:mm:ss",
            "fill_color": "D9EAD3",
        }],
    }), encoding="utf-8")

    completed = subprocess.run(
        [command, "--input", str(source), "--output", str(output), "--manifest", str(manifest_path)],
        capture_output=True,
        text=True,
        check=False,
    )

    assert completed.returncode == 0, completed.stderr
    with zipfile.ZipFile(output) as archive:
        worksheet = ElementTree.fromstring(archive.read("xl/worksheets/sheet1.xml"))
    namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    raw_value = worksheet.find(".//x:c[@r='A2']/x:v", namespace)
    assert raw_value is not None and float(raw_value.text) == pytest.approx(expected_serial)
    rendered = load_workbook(output, data_only=False)
    assert rendered.epoch == epoch
    observed = rendered["Data"]["A2"].value
    assert isinstance(observed, datetime) and observed.isoformat() == expected_datetime
    rendered.close()


@pytest.mark.parametrize(
    ("case", "operations", "metadata", "message"),
    [
        ("missing-operations", "__omit__", None, "operations are required"),
        ("null-operations", None, None, "operations are required"),
        ("null-operation", [None], None, "must not contain null"),
        ("invalid-cell", [{"type": "mark_cell", "sheet": "Data", "cell": "XFE1", "fill_color": "D9EAD3"}], None, "valid Excel cell"),
        ("invalid-row", [{"type": "mark_row", "sheet": "Data", "row": 0, "fill_color": "D9EAD3"}], None, "valid Excel row"),
        ("after-xfd", [{"type": "insert_column", "sheet": "Data", "after": "XFD", "canonical_field": "value", "header_row": 1, "header_value": "Value", "fill_color": "D9EAD3"}], None, "cannot insert after XFD"),
        ("formula-rows-without-template", [{"type": "insert_column", "sheet": "Data", "before": "B", "canonical_field": "value", "header_row": 1, "header_value": "Value", "fill_color": "D9EAD3", "formula_rows": [2]}], None, "require a formula_template"),
        ("unsafe-template", [{"type": "insert_column", "sheet": "Data", "before": "B", "canonical_field": "value", "header_row": 1, "header_value": "Value", "fill_color": "D9EAD3", "formula_template": "={other}"}], None, "unsafe or invalid"),
        ("unknown-validation", [{"type": "insert_column", "sheet": "Data", "before": "B", "canonical_field": "value", "header_row": 1, "header_value": "Value", "fill_color": "D9EAD3", "validation": {"type": "unknown"}}], None, "numeric validation requires"),
        ("empty-append", [{"type": "append_row", "sheet": "Data", "row": 2, "values": [], "fill_color": "D9EAD3"}], None, "non-empty values"),
        ("cross-row-append", [{"type": "append_row", "sheet": "Data", "row": 2, "values": [{"cell": "A3", "value": "E2", "field_type": "string"}], "fill_color": "D9EAD3"}], None, "declared row"),
        ("unknown-field-type", [{"type": "set_cell", "sheet": "Data", "cell": "A1", "value": "1", "field_type": "deciml", "fill_color": "D9EAD3"}], None, "unknown field_type"),
        ("missing-field-type", [{"type": "set_cell", "sheet": "Data", "cell": "A1", "value": "1", "fill_color": "D9EAD3"}], None, "requires field_type"),
        ("unsafe-integer", [{"type": "set_cell", "sheet": "Data", "cell": "A1", "value": "1234567890123456", "field_type": "integer", "fill_color": "D9EAD3"}], None, "exceeds Excel's safe numeric"),
        ("unsafe-decimal", [{"type": "set_cell", "sheet": "Data", "cell": "A1", "value": "0.1234567890123456", "field_type": "decimal", "fill_color": "D9EAD3"}], None, "exceeds Excel's safe numeric"),
        ("numeric-underflow", [{"type": "set_cell", "sheet": "Data", "cell": "A1", "value": "1e-309", "field_type": "decimal", "fill_color": "D9EAD3"}], None, "exceeds Excel's safe numeric"),
        ("numeric-overflow", [{"type": "set_cell", "sheet": "Data", "cell": "A1", "value": "1e308", "field_type": "decimal", "fill_color": "D9EAD3"}], None, "exceeds Excel's safe numeric"),
        ("numeric-boolean", [{"type": "set_cell", "sheet": "Data", "cell": "A1", "value": True, "field_type": "decimal", "fill_color": "D9EAD3"}], None, "must be JSON numbers or numeric strings"),
        ("ambiguous-datetime", [{"type": "set_cell", "sheet": "Data", "cell": "A1", "value": "2024-11-03T01:30:00-04:00", "field_type": "datetime", "timezone": "America/New_York", "fill_color": "D9EAD3"}], None, "ambiguous in the declared timezone"),
        ("nonexistent-datetime", [{"type": "set_cell", "sheet": "Data", "cell": "A1", "value": "2024-03-10T02:30:00-05:00", "field_type": "datetime", "timezone": "America/New_York", "fill_color": "D9EAD3"}], None, "nonexistent in the declared timezone"),
        ("datetime-offset-mismatch", [{"type": "set_cell", "sheet": "Data", "cell": "A1", "value": "2026-08-30T12:00:00Z", "field_type": "datetime", "timezone": "Asia/Shanghai", "fill_color": "D9EAD3"}], None, "offset does not match"),
        ("timezone-on-string", [{"type": "set_cell", "sheet": "Data", "cell": "A1", "value": "text", "field_type": "string", "timezone": "UTC", "fill_color": "D9EAD3"}], None, "only valid for datetime"),
        ("ambiguous-append-datetime", [{"type": "append_row", "sheet": "Data", "row": 2, "values": [{"cell": "A2", "value": "2024-11-03T01:30:00-05:00", "field_type": "datetime", "timezone": "America/New_York"}], "fill_color": "D9EAD3"}], None, "ambiguous in the declared timezone"),
        ("reserved-report-name", [{"type": "add_or_replace_report_sheet", "name": "__ExcelAuditorMetadata", "source_json": "report.json"}], None, "non-reserved name"),
        ("ignored-field", [{"type": "mark_cell", "sheet": "Data", "cell": "A1", "before": "B", "fill_color": "D9EAD3"}], None, "insert_column-only fields"),
        ("invalid-metadata-hash", [], {"schema_sha256": "not-a-hash"}, "must be SHA-256"),
    ],
)
def test_dotnet_renderer_rejects_invalid_or_ineffective_manifest_fields(tmp_path, case, operations, metadata, message):
    command = os.environ.get("EXCEL_RENDERER_COMMAND")
    if not command:
        pytest.skip("set EXCEL_RENDERER_COMMAND to run the .NET renderer contract test")
    source, output, manifest_path = tmp_path / "input.xlsx", tmp_path / "output.xlsx", tmp_path / "manifest.json"
    book = Workbook()
    book.active.title = "Data"
    book.active.append(["ID"])
    book.save(source)
    manifest = {
        "manifest_version": "1.0",
        "job_id": f"job_{case}",
        "input_sha256": hashlib.sha256(source.read_bytes()).hexdigest(),
    }
    if operations != "__omit__":
        manifest["operations"] = operations
    if metadata is not None:
        manifest["metadata"] = metadata
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    completed = subprocess.run(
        [command, "--input", str(source), "--output", str(output), "--manifest", str(manifest_path)],
        capture_output=True,
        text=True,
        check=False,
    )

    assert completed.returncode != 0
    failure = json.loads(completed.stderr)
    assert failure["error_code"] == "MANIFEST_OR_STRUCTURE_INVALID"
    assert message in failure["message"]
    assert not output.exists()


def test_dotnet_renderer_rejects_side_effect_formula_templates(tmp_path):
    command = os.environ.get("EXCEL_RENDERER_COMMAND")
    if not command:
        pytest.skip("set EXCEL_RENDERER_COMMAND to run the .NET renderer contract test")
    source, manifest_path = tmp_path / "input.xlsx", tmp_path / "manifest.json"
    book = Workbook()
    book.active.title = "Data"
    book.active.append(["ID"])
    book.save(source)
    digest = hashlib.sha256(source.read_bytes()).hexdigest()
    unsafe_formulas = [
        "=cmd|'/C calc'!A0",
        '=EXEC("calc")',
        '=REGISTER.ID("module","procedure","J")',
        "=_xlfn.IMAGE(A1)",
        "=STOCKHISTORY(A1)",
        '="unterminated',
    ]
    for index, formula in enumerate(unsafe_formulas):
        output = tmp_path / f"unsafe-formula-{index}.xlsx"
        manifest_path.write_text(json.dumps({
            "manifest_version": "1.0",
            "job_id": f"job_unsafe_formula_{index}",
            "input_sha256": digest,
            "operations": [{
                "type": "insert_column", "sheet": "Data", "before": "A", "canonical_field": "value",
                "header_row": 1, "header_value": "Value", "fill_color": "D9EAD3", "formula_template": formula,
            }],
        }), encoding="utf-8")
        completed = subprocess.run(
            [command, "--input", str(source), "--output", str(output), "--manifest", str(manifest_path)],
            capture_output=True,
            text=True,
            check=False,
        )
        assert completed.returncode != 0
        assert json.loads(completed.stderr)["error_code"] == "MANIFEST_OR_STRUCTURE_INVALID"
        assert not output.exists()


def test_service_repairs_use_normalized_typed_values_without_losing_raw_audit_values(tmp_path):
    command = os.environ.get("EXCEL_RENDERER_COMMAND")
    if not command:
        pytest.skip("set EXCEL_RENDERER_COMMAND to run the .NET renderer contract test")
    rules = RuleSet.model_validate({
        "schema_id": "typed-repairs", "schema_version": "1.0.0", "name": "Typed repairs",
        "sheets": [{
            "id": "data", "name": "Data", "primary_key": ["id"],
            "actions": {"overwrite_mismatch": True},
            "columns": [
                {"name": "id", "title": "ID", "required": True},
                {"name": "active", "title": "Active", "type": "boolean"},
                {"name": "status", "title": "Status", "type": "enum", "enum_values": ["active", "inactive"], "enum_aliases": {"A": "active"}},
                {"name": "event_at", "title": "Event", "type": "datetime", "compare": {"mode": "datetime", "timezone": "Asia/Shanghai"}},
                {"name": "tags", "title": "Tags", "type": "set"},
                {"name": "payload", "title": "Payload", "type": "json"},
                {"name": "ratio", "title": "Ratio", "type": "decimal", "normalize": ["percent_to_decimal"], "compare": {"mode": "numeric"}},
            ],
        }],
    })
    source, standard = tmp_path / "typed-repairs.xlsx", tmp_path / "typed-repairs.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID", "Active", "Status", "Event", "Tags", "Payload", "Ratio"])
    sheet.append(["E1", False, "inactive", "2026-01-01T00:00:00+08:00", "c", '{"z":0}', 0])
    book.save(source)
    standard.write_text(json.dumps({"data": [{
        "id": "E1", "active": "yes", "status": "A", "event_at": "2026-01-01T00:00:00Z",
        "tags": "b,a,a", "payload": {"b": 2, "a": 1}, "ratio": "10%",
    }]}), encoding="utf-8")
    service = AuditService(tmp_path / "typed-repair-runtime", renderer=DotNetOpenXmlRenderer(Path(command)))
    job_id = service.create_job()
    service.run(job_id, source, standard, rules)

    status = service.status(job_id)
    assert status["status"] == "completed", status
    assert status["summary"]["repairs_applied"] == 6
    rendered = load_workbook(service.artifact(job_id, "excel"), data_only=False)
    result = rendered["Data"]
    assert result["B2"].value is True and result["B2"].data_type == "b"
    assert "原值：false" in result["B2"].comment.text
    assert "标准值：yes" in result["B2"].comment.text
    assert "规则：active.overwrite_mismatch" in result["B2"].comment.text
    assert result["C2"].value == "active"
    assert result["D2"].value.isoformat() == "2026-01-01T08:00:00"
    assert result["E2"].value == "a,b"
    assert result["F2"].value == '{"a":1,"b":2}'
    assert result["G2"].value == pytest.approx(0.1) and result["G2"].data_type == "n"
    assert "原值：0" in result["G2"].comment.text
    assert "标准值：10%" in result["G2"].comment.text
    assert "规则：ratio.overwrite_mismatch" in result["G2"].comment.text
    rendered.close()
    report = json.loads(service.artifact(job_id, "json").read_text(encoding="utf-8"))
    raw_by_field = {item["canonical_field"]: item["standard_raw_value"] for item in report["differences"]}
    assert raw_by_field["active"] == "yes"
    assert raw_by_field["event_at"] == "2026-01-01T00:00:00Z"
    assert raw_by_field["ratio"] == "10%"


def test_service_blocks_lossy_numeric_repairs_but_writes_safe_values(tmp_path):
    command = os.environ.get("EXCEL_RENDERER_COMMAND")
    if not command:
        pytest.skip("set EXCEL_RENDERER_COMMAND to run the .NET renderer contract test")
    rules = RuleSet.model_validate({
        "schema_id": "numeric-write-precision", "schema_version": "1.0.0", "name": "Numeric write precision",
        "sheets": [{
            "id": "data", "name": "Data", "primary_key": ["id"],
            "actions": {"overwrite_mismatch": True, "missing_record": "append_and_mark_green"},
            "columns": [
                {"name": "id", "title": "ID", "required": True},
                {"name": "big", "title": "Big", "type": "integer"},
                {"name": "precise", "title": "Precise", "type": "decimal", "compare": {"mode": "numeric"}},
            ],
        }],
    })

    def run_case(name, rows, standard_rows):
        source, standard = tmp_path / f"{name}.xlsx", tmp_path / f"{name}.json"
        book = Workbook()
        sheet = book.active
        sheet.title = "Data"
        sheet.append(["ID", "Big", "Precise"])
        for row in rows:
            sheet.append(row)
        book.save(source)
        standard.write_text(json.dumps({"data": standard_rows}), encoding="utf-8")
        service = AuditService(tmp_path / f"{name}-runtime", renderer=DotNetOpenXmlRenderer(Path(command)))
        job_id = service.create_job()
        service.run(job_id, source, standard, rules)
        return service, job_id

    unsafe_service, unsafe_job = run_case(
        "unsafe-matched",
        [["E1", 1, 1]],
        [{"id": "E1", "big": "1234567890123456", "precise": "0.1234567890123456"}],
    )
    unsafe_status = unsafe_service.status(unsafe_job)
    assert unsafe_status["status"] == "manual_review", unsafe_status
    assert unsafe_status["summary"]["repairs_planned"] == 0
    assert "excel" not in unsafe_status["artifacts"]
    unsafe_report = json.loads(unsafe_service.artifact(unsafe_job, "json").read_text(encoding="utf-8"))
    unsafe_by_rule = {item["rule_id"]: item for item in unsafe_report["differences"] if item["rule_id"]}
    assert unsafe_by_rule["big.excel_write_precision"]["standard_raw_value"] == "1234567890123456"
    assert unsafe_by_rule["precise.excel_write_precision"]["standard_raw_value"] == "0.1234567890123456"
    assert all(
        item["repair_status"] == "not_requested"
        for item in unsafe_report["differences"]
        if item["type"] == "VALUE_MISMATCH"
    )

    append_service, append_job = run_case(
        "unsafe-append",
        [],
        [{"id": "E2", "big": "1234567890123456", "precise": "1"}],
    )
    append_status = append_service.status(append_job)
    assert append_status["status"] == "manual_review", append_status
    assert append_status["summary"]["repairs_planned"] == 0
    append_report = json.loads(append_service.artifact(append_job, "json").read_text(encoding="utf-8"))
    assert any(item["rule_id"] == "missing_record.numeric_write_blocked" for item in append_report["differences"])
    assert any(item["rule_id"] == "big.excel_write_precision" for item in append_report["differences"])

    default_rules = RuleSet.model_validate({
        "schema_id": "numeric-static-default", "schema_version": "1.0.0", "name": "Numeric static default",
        "sheets": [{
            "id": "data", "name": "Data", "primary_key": ["id"],
            "columns": [
                {"name": "id", "title": "ID", "required": True},
                {
                    "name": "big", "title": "Big", "type": "integer",
                    "fill_static_default": True, "static_default": "1234567890123456",
                },
            ],
        }],
    })
    default_source, default_standard = tmp_path / "unsafe-default.xlsx", tmp_path / "unsafe-default.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID", "Big"])
    sheet.append(["E1", None])
    book.save(default_source)
    default_standard.write_text(json.dumps({"data": [{"id": "E1"}]}), encoding="utf-8")
    default_service = AuditService(tmp_path / "unsafe-default-runtime", renderer=DotNetOpenXmlRenderer(Path(command)))
    default_job = default_service.create_job()
    default_service.run(default_job, default_source, default_standard, default_rules)
    default_status = default_service.status(default_job)
    assert default_status["status"] == "manual_review", default_status
    assert default_status["summary"]["repairs_planned"] == 0
    default_report = json.loads(default_service.artifact(default_job, "json").read_text(encoding="utf-8"))
    assert any(item["rule_id"] == "big.excel_write_precision" for item in default_report["differences"])

    range_service, range_job = run_case(
        "unsafe-range",
        [["E1", 1, 1]],
        [{"id": "E1", "big": "1", "precise": "1e-309"}],
    )
    range_status = range_service.status(range_job)
    assert range_status["status"] == "manual_review", range_status
    range_report = json.loads(range_service.artifact(range_job, "json").read_text(encoding="utf-8"))
    assert any(item["rule_id"] == "precise.excel_write_precision" for item in range_report["differences"])

    safe_service, safe_job = run_case(
        "safe-matched",
        [["E1", 1, 1]],
        [{"id": "E1", "big": "999999999999999", "precise": "0.123456789012345"}],
    )
    safe_status = safe_service.status(safe_job)
    assert safe_status["status"] == "completed", safe_status
    assert safe_status["summary"]["repairs_applied"] == 2
    rendered = load_workbook(safe_service.artifact(safe_job, "excel"), data_only=False)
    result = rendered["Data"]
    assert result["B2"].value == 999999999999999 and result["B2"].data_type == "n"
    assert Decimal(str(result["C2"].value)) == Decimal("0.123456789012345") and result["C2"].data_type == "n"
    rendered.close()


def test_service_blocks_lossy_dst_repairs_but_writes_unambiguous_datetimes(tmp_path):
    command = os.environ.get("EXCEL_RENDERER_COMMAND")
    if not command:
        pytest.skip("set EXCEL_RENDERER_COMMAND to run the .NET renderer contract test")
    rules = RuleSet.model_validate({
        "schema_id": "datetime-write-safety", "schema_version": "1.0.0", "name": "Datetime write safety",
        "sheets": [{
            "id": "data", "name": "Data", "primary_key": ["id"],
            "actions": {
                "overwrite_mismatch": True,
                "fill_empty_from_standard": True,
                "missing_record": "append_and_mark_green",
            },
            "columns": [
                {"name": "id", "title": "ID", "required": True},
                {
                    "name": "event_at", "title": "Event", "type": "datetime",
                    "compare": {"mode": "datetime", "timezone": "America/New_York"},
                },
            ],
        }],
    })

    def run_case(name, rows, standard_rows, case_rules=rules):
        source, standard = tmp_path / f"{name}.xlsx", tmp_path / f"{name}.json"
        book = Workbook()
        sheet = book.active
        sheet.title = "Data"
        sheet.append(["ID", "Event"])
        for row in rows:
            sheet.append(row)
        book.save(source)
        standard.write_text(json.dumps({"data": standard_rows}), encoding="utf-8")
        service = AuditService(tmp_path / f"{name}-runtime", renderer=DotNetOpenXmlRenderer(Path(command)))
        job_id = service.create_job()
        service.run(job_id, source, standard, case_rules)
        return service, job_id

    ambiguous = "2024-11-03T01:30:00-04:00"
    for name, rows in (
        ("dst-overwrite", [["E1", "2024-11-03T00:30:00-04:00"]]),
        ("dst-fill-empty", [["E1", None]]),
    ):
        blocked_service, blocked_job = run_case(name, rows, [{"id": "E1", "event_at": ambiguous}])
        blocked_status = blocked_service.status(blocked_job)
        assert blocked_status["status"] == "manual_review", blocked_status
        assert blocked_status["summary"]["repairs_planned"] == 0
        assert "excel" not in blocked_status["artifacts"]
        blocked_report = json.loads(blocked_service.artifact(blocked_job, "json").read_text(encoding="utf-8"))
        assert any(item["rule_id"] == "event_at.excel_write_timezone" for item in blocked_report["differences"])

    append_service, append_job = run_case(
        "dst-append",
        [],
        [{"id": "E2", "event_at": "2024-11-03T01:30:00-05:00"}],
    )
    append_status = append_service.status(append_job)
    assert append_status["status"] == "manual_review", append_status
    assert append_status["summary"]["repairs_planned"] == 0
    append_report = json.loads(append_service.artifact(append_job, "json").read_text(encoding="utf-8"))
    append_rule_ids = {item["rule_id"] for item in append_report["differences"]}
    assert "event_at.excel_write_timezone" in append_rule_ids
    assert "missing_record.datetime_write_blocked" in append_rule_ids

    default_rules = RuleSet.model_validate({
        "schema_id": "datetime-default-safety", "schema_version": "1.0.0", "name": "Datetime default safety",
        "sheets": [{
            "id": "data", "name": "Data", "primary_key": ["id"],
            "columns": [
                {"name": "id", "title": "ID", "required": True},
                {
                    "name": "event_at", "title": "Event", "type": "datetime",
                    "compare": {"mode": "datetime", "timezone": "America/New_York"},
                    "fill_static_default": True,
                    "static_default": ambiguous,
                },
            ],
        }],
    })
    default_service, default_job = run_case(
        "dst-default",
        [["E1", None]],
        [{"id": "E1"}],
        default_rules,
    )
    default_status = default_service.status(default_job)
    assert default_status["status"] == "manual_review", default_status
    default_report = json.loads(default_service.artifact(default_job, "json").read_text(encoding="utf-8"))
    assert any(item["rule_id"] == "event_at.excel_write_timezone" for item in default_report["differences"])

    safe_service, safe_job = run_case(
        "dst-safe",
        [["E1", "2024-11-03T00:30:00-04:00"]],
        [{"id": "E1", "event_at": "2024-11-03T03:30:00-05:00"}],
    )
    safe_status = safe_service.status(safe_job)
    assert safe_status["status"] == "completed", safe_status
    assert safe_status["summary"]["repairs_applied"] == 1
    rendered = load_workbook(safe_service.artifact(safe_job, "excel"), data_only=False)
    written = rendered["Data"]["B2"]
    assert written.value == datetime(2024, 11, 3, 3, 30)
    assert written.data_type == "d"
    rendered.close()
    public_manifest = json.loads(safe_service.artifact(safe_job, "manifest").read_text(encoding="utf-8"))
    datetime_operation = next(item for item in public_manifest["operations"] if item["type"] == "set_cell")
    assert datetime_operation["timezone"] == "America/New_York"

    safe_append_service, safe_append_job = run_case(
        "dst-safe-append",
        [],
        [{"id": "E3", "event_at": "2024-11-03T03:30:00-05:00"}],
    )
    safe_append_status = safe_append_service.status(safe_append_job)
    assert safe_append_status["status"] == "completed", safe_append_status
    assert safe_append_status["summary"]["repairs_applied"] == 1
    rendered = load_workbook(safe_append_service.artifact(safe_append_job, "excel"), data_only=False)
    assert rendered["Data"]["B2"].value == datetime(2024, 11, 3, 3, 30)
    rendered.close()
    public_manifest = json.loads(safe_append_service.artifact(safe_append_job, "manifest").read_text(encoding="utf-8"))
    append_operation = next(item for item in public_manifest["operations"] if item["type"] == "append_row")
    event_value = next(item for item in append_operation["values"] if item["cell"] == "B2")
    assert event_value["timezone"] == "America/New_York"


def test_service_renames_numeric_column_alias_as_a_string_header(tmp_path):
    command = os.environ.get("EXCEL_RENDERER_COMMAND")
    if not command:
        pytest.skip("set EXCEL_RENDERER_COMMAND to run the .NET renderer contract test")
    rules = RuleSet.model_validate({
        "schema_id": "numeric-header-alias", "schema_version": "1.0.0", "name": "Numeric header alias",
        "sheets": [{
            "id": "data", "name": "Data", "primary_key": ["id"],
            "actions": {"rename_confirmed_alias": True},
            "columns": [
                {"name": "id", "title": "ID", "required": True},
                {"name": "quantity", "title": "Quantity", "aliases": ["Qty"], "type": "integer"},
            ],
        }],
    })
    source, standard = tmp_path / "numeric-alias.xlsx", tmp_path / "numeric-alias.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID", "Qty"])
    sheet.append(["E1", 1])
    book.save(source)
    standard.write_text(json.dumps({"data": [{"id": "E1", "quantity": 1}]}), encoding="utf-8")
    service = AuditService(tmp_path / "numeric-alias-runtime", renderer=DotNetOpenXmlRenderer(Path(command)))
    job_id = service.create_job()
    service.run(job_id, source, standard, rules)

    status = service.status(job_id)
    assert status["status"] == "completed", status
    rendered = load_workbook(service.artifact(job_id, "excel"), data_only=False)
    assert rendered["Data"]["B1"].value == "Quantity"
    assert rendered["Data"]["B1"].data_type == "s"
    rendered.close()


def test_dotnet_renderer_updates_table_validation_defined_name_and_formula_ranges(tmp_path):
    command = os.environ.get("EXCEL_RENDERER_COMMAND")
    if not command:
        pytest.skip("set EXCEL_RENDERER_COMMAND to run the .NET renderer contract test")
    source, output, manifest_path = tmp_path / "structure.xlsx", tmp_path / "structure-output.xlsx", tmp_path / "manifest.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "结构"
    sheet.append(["编号", "数量", "单价", "合计", "文本引用"])
    sheet.append(["E001", 2, 5, "=B2*C2", '=IF(B2="B1",C2,0)'])
    sheet.append(["E002", 3, 4, "=B3*C3", '=IF(B3="C2",C3,0)'])
    sheet.add_table(Table(displayName="DataTable", ref="A1:C3"))
    validation = DataValidation(type="whole", operator="greaterThan", formula1="0")
    validation.add("C2:C3")
    sheet.add_data_validation(validation)
    sheet.conditional_formatting.add("B2:B3", FormulaRule(formula=["B2>0"]))
    sheet.auto_filter.ref = "A1:E3"
    sheet.auto_filter.add_filter_column(2, ["5"])
    sheet.freeze_panes = "C2"
    sheet.sheet_view.selection[0].activeCell = "C2"
    sheet.sheet_view.selection[0].sqref = "C2:D3"
    sheet["A3"]._hyperlink = Hyperlink(ref="A3", location="'结构'!B2", display="jump")
    book.defined_names.add(DefinedName("DataRange", attr_text="'结构'!$A$1:$C$3"))
    book.defined_names.add(DefinedName("LocalRange", attr_text="$A$1:$C$3", localSheetId=0))
    book.defined_names.add(DefinedName("ZZZ9999999", attr_text="'结构'!$A$2"))
    other = book.create_sheet("汇总")
    other["A1"] = "=结构!B2+'结构'!C2"
    other["A2"] = "=B1"
    other.conditional_formatting.add("A1", FormulaRule(formula=["结构!B2>0"]))
    cross_validation = DataValidation(type="custom", formula1="=结构!B2>0")
    cross_validation.add("A1")
    other.add_data_validation(cross_validation)
    sheet["G2"] = "=汇总!B1+结构!B2"
    sheet["G3"] = "=log10(b3)+ZZZ9999999"
    book.save(source)
    manifest_path.write_text(json.dumps({
        "manifest_version": "1.0", "job_id": "job_ranges", "input_sha256": hashlib.sha256(source.read_bytes()).hexdigest(),
        "operations": [{"type": "insert_column", "sheet": "结构", "before": "B", "canonical_field": "department", "header_row": 1, "header_value": "部门", "fill_color": "D9EAD3"}],
    }, ensure_ascii=False), encoding="utf-8")
    completed = subprocess.run([command, "--input", str(source), "--output", str(output), "--manifest", str(manifest_path)], capture_output=True, text=True, check=False)
    assert completed.returncode == 0, completed.stderr
    rendered = load_workbook(output)
    result_sheet = rendered["结构"]
    assert result_sheet.tables["DataTable"].ref == "A1:D3"
    assert [column.name for column in result_sheet.tables["DataTable"].tableColumns] == ["编号", "部门", "数量", "单价"]
    assert str(result_sheet.data_validations.dataValidation[0].sqref) == "D2:D3"
    target_conditional = next(iter(result_sheet.conditional_formatting))
    assert str(target_conditional.sqref) == "C2:C3"
    assert result_sheet.conditional_formatting[target_conditional][0].formula == ["C2>0"]
    assert result_sheet.auto_filter.ref == "A1:F3"
    assert result_sheet.auto_filter.filterColumn[0].colId == 3
    assert result_sheet.freeze_panes == "D2"
    assert result_sheet.sheet_view.selection[0].activeCell == "D2"
    assert str(result_sheet.sheet_view.selection[0].sqref) == "D2:E3"
    assert result_sheet["A3"].hyperlink.location == "'结构'!C2"
    assert rendered.defined_names["DataRange"].attr_text == "'结构'!$A$1:$D$3"
    assert result_sheet.defined_names["LocalRange"].attr_text == "$A$1:$D$3"
    assert result_sheet["E2"].value == "=C2*D2"
    assert result_sheet["F2"].value == '=IF(C2="B1",D2,0)'
    assert result_sheet["F3"].value == '=IF(C3="C2",D3,0)'
    assert result_sheet["H2"].value == "=汇总!B1+结构!C2"
    assert result_sheet["H3"].value == "=log10(C3)+ZZZ9999999"
    assert rendered["汇总"]["A1"].value == "=结构!C2+'结构'!D2"
    assert rendered["汇总"]["A2"].value == "=B1"
    cross_sheet = rendered["汇总"]
    cross_conditional = next(iter(cross_sheet.conditional_formatting))
    assert str(cross_conditional.sqref) == "A1"
    assert cross_sheet.conditional_formatting[cross_conditional][0].formula == ["结构!C2>0"]
    assert cross_sheet.data_validations.dataValidation[0].formula1 == "=结构!C2>0"
    rendered.close()


def test_dotnet_renderer_rejects_complex_conditional_formatting_formula(tmp_path):
    command = os.environ.get("EXCEL_RENDERER_COMMAND")
    if not command:
        pytest.skip("set EXCEL_RENDERER_COMMAND to run the .NET renderer contract test")
    source, output, manifest_path = tmp_path / "conditional.xlsx", tmp_path / "conditional-output.xlsx", tmp_path / "manifest.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID", "Amount"])
    sheet.append(["E1", 1])
    sheet.conditional_formatting.add("B2", FormulaRule(formula=["SUM(B:B)>0"]))
    book.save(source)
    manifest_path.write_text(json.dumps({
        "manifest_version": "1.0", "job_id": "job_conditional", "input_sha256": hashlib.sha256(source.read_bytes()).hexdigest(),
        "operations": [{"type": "insert_column", "sheet": "Data", "before": "B", "canonical_field": "score", "header_row": 1, "header_value": "Score", "fill_color": "D9EAD3"}],
    }), encoding="utf-8")

    completed = subprocess.run([command, "--input", str(source), "--output", str(output), "--manifest", str(manifest_path)], capture_output=True, text=True, check=False)

    assert completed.returncode != 0
    failure = json.loads(completed.stderr)
    assert failure["error_code"] == "UNSUPPORTED_FEATURE"
    assert "conditional formatting" in failure["message"]
    assert not output.exists()


def test_dotnet_renderer_supports_insert_after_contract(tmp_path):
    command = os.environ.get("EXCEL_RENDERER_COMMAND")
    if not command:
        pytest.skip("set EXCEL_RENDERER_COMMAND to run the .NET renderer contract test")
    source, output, manifest_path = tmp_path / "after.xlsx", tmp_path / "after-output.xlsx", tmp_path / "manifest.json"
    book = Workbook(); sheet = book.active; sheet.title = "Data"; sheet.append(["A", "B"]); sheet.append([1, 2]); book.save(source)
    manifest_path.write_text(json.dumps({
        "manifest_version": "1.0", "job_id": "job_after", "input_sha256": hashlib.sha256(source.read_bytes()).hexdigest(),
        "operations": [{"type": "insert_column", "sheet": "Data", "after": "B", "canonical_field": "c", "header_row": 1, "header_value": "C", "fill_color": "D9EAD3"}],
    }), encoding="utf-8")
    completed = subprocess.run([command, "--input", str(source), "--output", str(output), "--manifest", str(manifest_path)], capture_output=True, text=True, check=False)
    assert completed.returncode == 0, completed.stderr
    rendered = load_workbook(output)
    assert [rendered["Data"].cell(1, index).value for index in range(1, 4)] == ["A", "B", "C"]
    rendered.close()


def test_dotnet_renderer_rejects_unsafe_complex_formula_insertions(tmp_path):
    command = os.environ.get("EXCEL_RENDERER_COMMAND")
    if not command:
        pytest.skip("set EXCEL_RENDERER_COMMAND to run the .NET renderer contract test")
    source, output, manifest_path = tmp_path / "complex.xlsx", tmp_path / "complex-output.xlsx", tmp_path / "manifest.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID", "Amount", "Total"])
    sheet.append(["E1", 1, "=SUM(B:B)"])
    book.save(source)
    manifest_path.write_text(json.dumps({
        "manifest_version": "1.0", "job_id": "job_complex", "input_sha256": hashlib.sha256(source.read_bytes()).hexdigest(),
        "operations": [{"type": "insert_column", "sheet": "Data", "before": "B", "canonical_field": "score", "header_row": 1, "header_value": "Score", "fill_color": "D9EAD3"}],
    }), encoding="utf-8")
    completed = subprocess.run([command, "--input", str(source), "--output", str(output), "--manifest", str(manifest_path)], capture_output=True, text=True, check=False)
    assert completed.returncode != 0
    failure = json.loads(completed.stderr)
    assert failure["error_code"] == "UNSUPPORTED_FEATURE"
    assert "complex" in failure["message"]
    assert not output.exists()


def test_dotnet_renderer_rejects_formula_reference_shift_beyond_xfd(tmp_path):
    command = os.environ.get("EXCEL_RENDERER_COMMAND")
    if not command:
        pytest.skip("set EXCEL_RENDERER_COMMAND to run the .NET renderer contract test")
    source, output, manifest_path = tmp_path / "xfd-formula.xlsx", tmp_path / "xfd-formula-output.xlsx", tmp_path / "manifest.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["Formula"])
    sheet["A2"] = "=XFD1"
    book.save(source)
    manifest_path.write_text(json.dumps({
        "manifest_version": "1.0", "job_id": "job_xfd_formula", "input_sha256": hashlib.sha256(source.read_bytes()).hexdigest(),
        "operations": [{"type": "insert_column", "sheet": "Data", "before": "A", "canonical_field": "id", "header_row": 1, "header_value": "ID", "fill_color": "D9EAD3"}],
    }), encoding="utf-8")

    completed = subprocess.run([command, "--input", str(source), "--output", str(output), "--manifest", str(manifest_path)], capture_output=True, text=True, check=False)

    assert completed.returncode != 0
    failure = json.loads(completed.stderr)
    assert failure["error_code"] == "UNSUPPORTED_FEATURE"
    assert "beyond XFD" in failure["message"]
    assert not output.exists()


def test_dotnet_renderer_preserves_macro_and_signature_parts_during_real_edit(tmp_path):
    command = os.environ.get("EXCEL_RENDERER_COMMAND")
    if not command:
        pytest.skip("set EXCEL_RENDERER_COMMAND to run the .NET renderer contract test")
    xlsx, source, output, manifest_path = tmp_path / "base.xlsx", tmp_path / "macro.xlsm", tmp_path / "macro-output.xlsm", tmp_path / "manifest.json"
    book = Workbook()
    book.active.title = "Data"
    book.active.append(["ID"])
    book.active.append(["E1"])
    book.save(xlsx)
    macro_payload = b"\x00VBA-PROJECT-GOLDEN\xff\x10"
    signature_payload = b"\x00VBA-SIGNATURE-GOLDEN\xfe\x20"
    with zipfile.ZipFile(xlsx) as incoming, zipfile.ZipFile(source, "w", zipfile.ZIP_DEFLATED) as outgoing:
        for info in incoming.infolist():
            data = incoming.read(info.filename)
            if info.filename == "[Content_Types].xml":
                text = data.decode("utf-8").replace(
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
                    "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
                ).replace(
                    "</Types>",
                    '<Override PartName="/xl/vbaProject.bin" ContentType="application/vnd.ms-office.vbaProject"/>'
                    '<Override PartName="/xl/vbaProjectSignature.bin" ContentType="application/vnd.ms-office.vbaProjectSignature"/>'
                    "</Types>",
                )
                data = text.encode("utf-8")
            elif info.filename == "xl/_rels/workbook.xml.rels":
                text = data.decode("utf-8").replace("</Relationships>", '<Relationship Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" Target="vbaProject.bin" Id="rIdVba"/></Relationships>')
                data = text.encode("utf-8")
            outgoing.writestr(info, data)
        outgoing.writestr("xl/vbaProject.bin", macro_payload)
        outgoing.writestr("xl/vbaProjectSignature.bin", signature_payload)
        outgoing.writestr(
            "xl/_rels/vbaProject.bin.rels",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rIdSignature" '
            'Type="http://schemas.microsoft.com/office/2006/relationships/vbaProjectSignature" '
            'Target="vbaProjectSignature.bin"/>'
            '</Relationships>',
        )
    manifest_path.write_text(json.dumps({
        "manifest_version": "1.0", "job_id": "job_macro", "input_sha256": hashlib.sha256(source.read_bytes()).hexdigest(),
        "operations": [{
            "type": "mark_cell", "sheet": "Data", "cell": "A2", "fill_color": "FFF2CC",
            "comment": "audited", "difference_id": "diff_macro",
        }],
    }), encoding="utf-8")
    completed = subprocess.run([command, "--input", str(source), "--output", str(output), "--manifest", str(manifest_path)], capture_output=True, text=True, check=False)
    assert completed.returncode == 0, completed.stderr
    with zipfile.ZipFile(output) as archive:
        assert archive.read("xl/vbaProject.bin") == macro_payload
        assert archive.read("xl/vbaProjectSignature.bin") == signature_payload
    rendered = load_workbook(output, keep_vba=True)
    assert rendered["Data"]["A2"].comment.text == "audited"
    assert rendered["Data"]["A2"].fill.fgColor.rgb.endswith("FFF2CC")
    rendered.close()


def test_rendered_output_opens_and_resaves_in_libreoffice(tmp_path):
    command = os.environ.get("EXCEL_RENDERER_COMMAND")
    libreoffice = os.environ.get("LIBREOFFICE_COMMAND")
    if not command or not libreoffice:
        pytest.skip("set renderer and LibreOffice commands to run compatibility test")
    rules = RuleSet.model_validate({
        "schema_id": "compat", "schema_version": "1.0.0", "name": "Compatibility",
        "sheets": [{"id": "data", "name": "Data", "primary_key": ["id"], "columns": [
            {"name": "id", "title": "ID", "required": True},
            {"name": "name", "title": "Name", "required": True},
        ]}],
    })
    source, standard = tmp_path / "compat.xlsx", tmp_path / "standard.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID", "Name"])
    sheet.append(["E1", "Alice"])
    book.save(source)
    standard.write_text(json.dumps({"data": [{"id": "E1", "name": "Alice"}]}), encoding="utf-8")
    service = AuditService(tmp_path / "runtime")
    job_id = service.create_job()
    service.run(job_id, source, standard, rules)
    output = service.artifact(job_id, "excel")
    converted = tmp_path / "libreoffice"
    converted.mkdir()
    completed = subprocess.run([libreoffice, "--headless", "--convert-to", "xlsx", "--outdir", str(converted), str(output)], capture_output=True, text=True, timeout=120, check=False)
    assert completed.returncode == 0, completed.stderr
    resaved = converted / output.with_suffix(".xlsx").name
    assert resaved.is_file()
    load_workbook(resaved, read_only=True).close()
