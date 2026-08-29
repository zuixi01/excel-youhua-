import hashlib
import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from excel_auditor.engine import compare_workbook
from excel_auditor.models import RuleSet
from excel_auditor.service import AuditService
from excel_auditor.snapshots import SpilledRecords, StandardSnapshot, load_snapshot
from excel_auditor.workbook import SpilledRows, inspect_workbook


ROOT = Path("tests/golden_files")
ADVANCED = json.loads((ROOT / "advanced_scenarios.json").read_text(encoding="utf-8"))


def _input(name: str, scenario: dict) -> Path:
    path = ROOT / "workbooks" / f"{name}.xlsx"
    assert hashlib.sha256(path.read_bytes()).hexdigest() == scenario["input_sha256"]
    return path


def test_typed_compound_key_golden_is_exhaustive(tmp_path):
    scenario = ADVANCED["typed_compound"]
    rules = RuleSet.model_validate(scenario["rule"])
    result = compare_workbook(inspect_workbook(_input("typed_compound", scenario), rules), scenario["standard"], rules)
    assert dict(sorted(Counter(item.type.value for item in result.differences).items())) == scenario["expected_counts"]
    assert result.summary.matched_records == 1
    assert not {"VALUE_MISMATCH", "INVALID_VALUE", "VALIDATION_ERROR", "EXTRA_RECORD", "MISSING_RECORD"} & {item.type.value for item in result.differences}


def test_structured_multisheet_golden_preserves_key_parts(tmp_path):
    scenario = ADVANCED["structured_multisheet"]
    rules = RuleSet.model_validate(scenario["rule"])
    source = _input("structured_multisheet", scenario)
    comparison = compare_workbook(inspect_workbook(source, rules), scenario["standard"], rules)
    assert dict(sorted(Counter(item.type.value for item in comparison.differences).items())) == scenario["expected_counts"]
    standard = tmp_path / "standard.json"
    standard.write_text(json.dumps(scenario["standard"], ensure_ascii=False), encoding="utf-8")
    service = AuditService(tmp_path / "runtime")
    job_id = service.create_job()
    service.run(job_id, source, standard, rules)
    assert service.status(job_id)["status"] == "completed"
    rendered = load_workbook(service.artifact(job_id, "excel"), data_only=False)
    data = rendered["结构"]
    assert data.tables["GoldenTable"].ref == "A1:C3"
    assert str(data.data_validations.dataValidation[0].sqref) == "B2:B3"
    assert data.freeze_panes == "B2"
    assert data.row_dimensions[3].hidden is True
    assert data.column_dimensions["B"].hidden is True
    assert data["C2"].value == "=B2*2"
    assert rendered.defined_names["GoldenData"].attr_text == "'结构'!$A$1:$C$3"
    assert rendered["说明"].sheet_state == "hidden"
    assert {"核验报告", "__ExcelAuditorMetadata"} <= set(rendered.sheetnames)
    rendered.close()


def test_merged_header_golden_requires_manual_review(tmp_path):
    scenario = ADVANCED["merged_header_manual_review"]
    rules = RuleSet.model_validate(scenario["rule"])
    source = _input("merged_header_manual_review", scenario)
    standard = tmp_path / "standard.json"
    standard.write_text(json.dumps(scenario["standard"], ensure_ascii=False), encoding="utf-8")
    service = AuditService(tmp_path / "runtime")
    job_id = service.create_job()
    service.run(job_id, source, standard, rules)
    status = service.status(job_id)
    assert status["status"] == scenario["expected_status"]
    assert any("merged_cells" in warning for warning in status["warnings"])
    assert "excel" not in status["artifacts"]
    manifest = json.loads(service.artifact(job_id, "manifest").read_text(encoding="utf-8"))
    assert manifest["rendering"] == {"status": "skipped", "reason": "manual_review"}
    assert manifest["operations"] == []
    assert not (service.job_directory(job_id) / "render-manifest.private.json").exists()
    assert not (service.job_directory(job_id) / "report-render.json").exists()


def test_large_golden_uses_disk_backed_workbook_and_standard_snapshot(monkeypatch):
    scenario = ADVANCED["large_report_only"]
    rules = RuleSet.model_validate(scenario["rule"])
    workbook = inspect_workbook(_input("large_report_only", scenario), rules)
    assert workbook.large_mode is True and workbook.report_only is True
    assert isinstance(workbook.sheets["大数据"].rows, SpilledRows)
    standard_path = ROOT / scenario["standard_snapshot"]
    assert hashlib.sha256(standard_path.read_bytes()).hexdigest() == scenario["standard_sha256"]
    snapshot = StandardSnapshot(
        "std_golden_large",
        standard_path,
        scenario["standard_sha256"],
        scenario["standard_record_count"],
        datetime(2026, 8, 28, tzinfo=timezone.utc),
        {},
    )
    standard = load_snapshot(snapshot, spill_after_records=100)
    assert isinstance(standard["large"], SpilledRecords)
    monkeypatch.setenv("EXCEL_AUDITOR_POLARS_JOIN_THRESHOLD", "1")
    try:
        result = compare_workbook(workbook, standard, rules)
        assert dict(sorted(Counter(item.type.value for item in result.differences).items())) == scenario["expected_counts"]
        assert result.summary.matched_records == scenario["standard_record_count"]
        assert result.join_backends == ["polars_partitioned"]
    finally:
        standard["large"].close()
        workbook.sheets["大数据"].rows.close()
