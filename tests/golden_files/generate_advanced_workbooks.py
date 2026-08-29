"""Explicit maintainer tool for regenerating committed advanced Golden inputs."""

import hashlib
import json
import re
import zipfile
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName


ROOT = Path(__file__).resolve().parent
OUTPUT = ROOT / "workbooks"
OUTPUT.mkdir(exist_ok=True)


def save_deterministic(book: Workbook, name: str) -> None:
    book.properties.created = datetime(2026, 8, 28, tzinfo=timezone.utc)
    book.properties.modified = datetime(2026, 8, 28, tzinfo=timezone.utc)
    destination = OUTPUT / f"{name}.xlsx"
    temporary = destination.with_suffix(".tmp.xlsx")
    book.save(temporary)
    with zipfile.ZipFile(temporary) as source, zipfile.ZipFile(destination, "w", zipfile.ZIP_DEFLATED, compresslevel=9) as target:
        for source_info in source.infolist():
            info = zipfile.ZipInfo(source_info.filename, date_time=(2026, 8, 28, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = source_info.external_attr
            data = source.read(source_info.filename)
            if source_info.filename == "docProps/core.xml":
                for field in (b"created", b"modified"):
                    pattern = rb"(<dcterms:" + field + rb"[^>]*>)[^<]*(</dcterms:" + field + rb">)"
                    data = re.sub(pattern, rb"\g<1>2026-08-28T00:00:00Z\g<2>", data)
            target.writestr(info, data, compresslevel=9)
    temporary.unlink()


typed = Workbook()
typed_sheet = typed.active
typed_sheet.title = "订单"
typed_sheet.append(["员工编号", "期间", "代码", "金额", "百分比", "日期", "时间", "启用", "部门", "标签", "JSON"])
typed_sheet.append([" e001 ", "202608", "00123", "￥10,000.005", "12.5%", "2026-08-28", "2026-08-28T10:30:00+08:00", "是", "技术", "A,B,A", '{"b":2,"a":1}'])
save_deterministic(typed, "typed_compound")

structured = Workbook()
data = structured.active
data.title = "结构"
data.append(["编号", "金额", "合计"])
data.append(["E001", 5, "=B2*2"])
data.append(["E002", 6, "=B3*2"])
table = Table(displayName="GoldenTable", ref="A1:C3")
table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
data.add_table(table)
validation = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0")
validation.add("B2:B3")
data.add_data_validation(validation)
data.freeze_panes = "B2"
data.row_dimensions[3].hidden = True
data.column_dimensions["B"].hidden = True
structured.defined_names.add(DefinedName("GoldenData", attr_text="'结构'!$A$1:$C$3"))
notes = structured.create_sheet("说明")
notes.append(["编号", "备注"])
notes.append(["N001", "固定说明"])
notes.sheet_state = "hidden"
save_deterministic(structured, "structured_multisheet")

merged = Workbook()
merged_sheet = merged.active
merged_sheet.title = "合并表头"
merged_sheet["A1"] = "人员"
merged_sheet.merge_cells("A1:B1")
merged_sheet.append(["E001", "张三"])
save_deterministic(merged, "merged_header_manual_review")

large = Workbook(write_only=True)
large_sheet = large.create_sheet("大数据")
large_columns = ["ID", *[f"F{index}" for index in range(1, 20)]]
large_sheet.append(large_columns)
large_snapshot = ROOT / "large_report_only.standard.jsonl"
large_digest = hashlib.sha256()
with large_snapshot.open("wb") as standard_handle:
    for row_index in range(10_000):
        record = {"id": f"R{row_index:06d}", **{f"f{column}": f"V{row_index}-{column}" for column in range(1, 20)}}
        large_sheet.append([record["id"], *[record[f"f{column}"] for column in range(1, 20)]])
        line = (json.dumps({"sheet_id": "large", "record": record}, sort_keys=True, separators=(",", ":")) + "\n").encode("utf-8")
        standard_handle.write(line)
        large_digest.update(line)
save_deterministic(large, "large_report_only")
print(f"large standard sha256={large_digest.hexdigest()}")
