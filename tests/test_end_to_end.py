import json
import os
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment

from excel_auditor.rules import load_rules
from excel_auditor.models import RuleSet
from excel_auditor.rendering import DotNetOpenXmlRenderer, ExcelRenderer, OpenPyxlDevelopmentRenderer
from excel_auditor.service import AuditService


def _input(path: Path) -> None:
    book = Workbook()
    sheet = book.active
    sheet.title = "人员信息"
    sheet.append(["员工数据"])
    sheet.append(["工号", "姓名", "部门", "工资", "临时备注"])
    sheet.append(["E001", "张三", "技术", "10000.005", "保留"])
    sheet.append(["E003", "王五", "财务部", "8000", "保留"])
    book.save(path)


def test_large_report_only_skips_renderer_and_private_manifest(tmp_path):
    columns = [{"name": "id", "title": "ID", "required": True}] + [{"name": f"f{index}", "title": f"F{index}"} for index in range(1, 100)]
    rules = RuleSet.model_validate({
        "schema_id": "report-only", "schema_version": "1.0.0", "name": "Report only",
        "workbook": {"max_in_memory_cells": 10000, "large_file_action": "report_only"},
        "sheets": [{"id": "data", "name": "Data", "primary_key": ["id"], "data_region": {"include_hidden_rows": True}, "columns": columns}],
    })
    excel = tmp_path / "large.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append([column["title"] for column in columns])
    for index in range(100):
        sheet.append([f"E{index:03d}", *([None] * 99)])
    book.save(excel)
    standard = tmp_path / "standard.json"
    standard.write_text(json.dumps({"data": [{"id": f"E{index:03d}"} for index in range(100)]}), encoding="utf-8")

    class MustNotRender(ExcelRenderer):
        def render(self, source, destination, workbook, rules, comparison, report_payload):
            raise AssertionError("report-only jobs must not invoke the renderer")

    service = AuditService(tmp_path / "runtime", renderer=MustNotRender())
    job_id = service.create_job()
    service.run(job_id, excel, standard, rules)
    status = service.status(job_id)
    assert status["status"] == "completed" and status["mode"] == "report_only"
    manifest = json.loads(service.artifact(job_id, "manifest").read_text(encoding="utf-8"))
    assert manifest["rendering"] == {"status": "skipped", "reason": "large_file_report_only"}
    assert manifest["operations"] == []
    assert not (service.job_directory(job_id) / "render-manifest.private.json").exists()
    assert not (service.job_directory(job_id) / "report-render.json").exists()


def test_full_audit_produces_reproducible_reports_and_colored_workbook(tmp_path):
    excel = tmp_path / "input.xlsx"
    standard = tmp_path / "standard.json"
    _input(excel)
    standard.write_text(json.dumps({"employees": [
        {"employee_id": "E001", "employee_name": "张三", "department": "技术部", "salary": "10000.00", "hire_date": None},
        {"employee_id": "E002", "employee_name": "李四", "department": "人事部", "salary": "9000", "hire_date": None},
    ]}, ensure_ascii=False), encoding="utf-8")
    rules = load_rules(Path("configs/examples/employee-roster.yaml"))
    service = AuditService(tmp_path / "runtime")
    job_id = service.create_job()
    service.run(job_id, excel, standard, rules)
    status = service.status(job_id)
    assert status["status"] == "completed", status
    assert status["summary"]["matched_records"] == 1
    assert status["summary"]["extra_records"] == 1
    assert status["summary"]["missing_records"] == 1
    report = json.loads(service.artifact(job_id, "json").read_text(encoding="utf-8"))
    types = {item["type"] for item in report["differences"]}
    assert {"EXTRA_HEADER", "EXTRA_RECORD", "MISSING_RECORD", "MISSING_HEADER"} <= types
    assert "VALUE_MISMATCH" not in types  # salary is inside tolerance; enum alias is normalized
    output = load_workbook(service.artifact(job_id, "excel"))
    assert output["人员信息"]["E2"].value == "临时备注"
    assert "核验报告" in output.sheetnames
    output.close()
    second_job = service.create_job()
    service.run(second_job, excel, standard, rules)
    second_report = json.loads(service.artifact(second_job, "json").read_text(encoding="utf-8"))
    for payload in (report, second_report):
        for volatile in ("job_id", "created_at", "standard_snapshot_id", "output_sha256"):
            payload.pop(volatile, None)
        for difference in payload.get("differences", []):
            difference.pop("job_id", None)
            difference.pop("difference_id", None)
    assert report == second_report
    first_manifest = json.loads(service.artifact(job_id, "manifest").read_text(encoding="utf-8"))
    second_manifest = json.loads(service.artifact(second_job, "manifest").read_text(encoding="utf-8"))
    first_manifest.pop("job_id", None)
    second_manifest.pop("job_id", None)
    assert first_manifest == second_manifest


def test_missing_required_column_is_inserted_and_marked_green(tmp_path):
    excel = tmp_path / "missing.xlsx"
    standard = tmp_path / "standard.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "人员信息"
    sheet.append(["员工数据"])
    sheet.append(["工号", "姓名", "工资"])
    sheet.append(["E001", "张三", "100"])
    book.save(excel)
    standard.write_text(json.dumps({"employees": [{"employee_id": "E001", "employee_name": "张三", "salary": "100", "department": "技术部"}]}, ensure_ascii=False), encoding="utf-8")
    service = AuditService(tmp_path / "runtime")
    job_id = service.create_job()
    service.run(job_id, excel, standard, load_rules(Path("configs/examples/employee-roster.yaml")))
    assert service.status(job_id)["status"] == "completed"
    output = load_workbook(service.artifact(job_id, "excel"))
    headers = [output["人员信息"].cell(2, col).value for col in range(1, output["人员信息"].max_column + 1)]
    assert headers == ["工号", "姓名", "工资", "部门"]
    assert output["人员信息"]["D2"].fill.fgColor.rgb.endswith("D9EAD3")
    output.close()


def test_formula_workbook_is_routed_to_manual_review(tmp_path):
    excel = tmp_path / "formula.xlsx"
    standard = tmp_path / "standard.json"
    _input(excel)
    book = load_workbook(excel)
    book["人员信息"]["D3"] = "=5000+5000"
    book.save(excel)
    standard.write_text(json.dumps({"employees": []}, ensure_ascii=False), encoding="utf-8")
    service = AuditService(tmp_path / "runtime")
    job_id = service.create_job()
    service.run(job_id, excel, standard, load_rules(Path("configs/examples/employee-roster.yaml")))
    status = service.status(job_id)
    assert status["status"] == "manual_review"
    assert "excel" not in status["artifacts"]


def test_formula_text_mode_compares_without_executing_formula(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "formula", "schema_version": "1.0.0", "name": "Formula",
        "sheets": [{"id": "data", "name": "数据", "primary_key": ["id"], "columns": [
            {"name": "id", "title": "编号", "required": True},
            {"name": "expression", "title": "表达式", "compare": {"formula_mode": "formula"}},
        ]}],
    })
    excel, standard = tmp_path / "formula-text.xlsx", tmp_path / "formula-standard.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "数据"
    sheet.append(["编号", "表达式"])
    sheet.append(["E1", "=1+1"])
    book.save(excel)
    standard.write_text(json.dumps({"data": [{"id": "E1", "expression": "=1+1"}]}), encoding="utf-8")
    service = AuditService(tmp_path / "formula-runtime")
    job_id = service.create_job()
    service.run(job_id, excel, standard, rules)
    status = service.status(job_id)
    assert status["status"] == "completed", status
    rendered = load_workbook(service.artifact(job_id, "excel"), data_only=False)
    assert rendered["数据"]["B2"].value == "=1+1"
    rendered.close()


def test_formula_text_mode_supports_typed_fields_and_is_symmetric(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "typed-formula", "schema_version": "1.0.0", "name": "Typed formula",
        "sheets": [{
            "id": "data", "name": "Data", "primary_key": ["id"],
            "actions": {"overwrite_mismatch": True},
            "columns": [
                {"name": "id", "title": "ID", "required": True},
                {"name": "amount", "title": "Amount", "type": "decimal", "compare": {"mode": "numeric", "formula_mode": "formula"}},
            ],
        }],
    })
    excel, standard = tmp_path / "typed-formula.xlsx", tmp_path / "typed-formula.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID", "Amount"])
    sheet.append(["E1", "=1+1"])
    sheet.append(["E2", 2])
    sheet.append(["E3", "=2+2"])
    book.save(excel)
    standard.write_text(json.dumps({"data": [
        {"id": "E1", "amount": "=1+1"},
        {"id": "E2", "amount": "=1+1"},
        {"id": "E3"},
    ]}), encoding="utf-8")

    service = AuditService(tmp_path / "typed-formula-runtime")
    job_id = service.create_job()
    service.run(job_id, excel, standard, rules)
    status = service.status(job_id)
    assert status["status"] == "completed", status
    assert status["summary"]["mismatched_cells"] == 1
    assert status["summary"]["repairs_planned"] == 0
    report = json.loads(service.artifact(job_id, "json").read_text(encoding="utf-8"))
    mismatch = next(item for item in report["differences"] if item["type"] == "VALUE_MISMATCH")
    assert mismatch["business_key"] == {"id": "E2"}
    assert mismatch["rule_id"] == "amount.formula_text"
    rendered = load_workbook(service.artifact(job_id, "excel"), data_only=False)
    assert rendered["Data"]["B2"].value == "=1+1"
    assert rendered["Data"]["B3"].value == 2
    assert rendered["Data"]["B4"].value == "=2+2"
    rendered.close()


def test_formula_record_append_requires_a_matching_trusted_template(tmp_path):
    def rules_for(template=None):
        column = {
            "name": "amount", "title": "Amount", "type": "decimal",
            "compare": {"mode": "numeric", "formula_mode": "formula"},
        }
        if template is not None:
            column["formula_template"] = template
        return RuleSet.model_validate({
            "schema_id": "formula-append", "schema_version": "1.0.0", "name": "Formula append",
            "sheets": [{
                "id": "data", "name": "Data", "primary_key": ["id"],
                "actions": {"missing_record": "append_and_mark_green"},
                "columns": [{"name": "id", "title": "ID", "required": True}, column],
            }],
        })

    excel = tmp_path / "formula-append.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID", "Amount"])
    book.save(excel)
    unsafe_standard = tmp_path / "unsafe-formula.json"
    unsafe_standard.write_text(json.dumps({"data": [{"id": "E1", "amount": "=1+1"}]}), encoding="utf-8")
    unsafe_service = AuditService(tmp_path / "unsafe-formula-runtime")
    unsafe_job = unsafe_service.create_job()
    unsafe_service.run(unsafe_job, excel, unsafe_standard, rules_for())
    unsafe_status = unsafe_service.status(unsafe_job)
    assert unsafe_status["status"] == "manual_review"
    assert "excel" not in unsafe_status["artifacts"]
    unsafe_report = json.loads(unsafe_service.artifact(unsafe_job, "json").read_text(encoding="utf-8"))
    missing = next(item for item in unsafe_report["differences"] if item["type"] == "MISSING_RECORD")
    assert missing["rule_id"] == "missing_record.formula_template_required"
    assert missing["repair_status"] == "not_requested"

    trusted_standard = tmp_path / "trusted-formula.json"
    trusted_standard.write_text(json.dumps({"data": [{"id": "E1", "amount": "=ROW()+2"}]}), encoding="utf-8")
    trusted_service = AuditService(tmp_path / "trusted-formula-runtime")
    trusted_job = trusted_service.create_job()
    trusted_service.run(trusted_job, excel, trusted_standard, rules_for("=ROW()+{row}"))
    trusted_status = trusted_service.status(trusted_job)
    assert trusted_status["status"] == "completed", trusted_status
    assert trusted_status["summary"]["repairs_applied"] == 1
    rendered = load_workbook(trusted_service.artifact(trusted_job, "excel"), data_only=False)
    assert rendered["Data"]["A2"].value == "E1"
    assert rendered["Data"]["B2"].value == "=ROW()+2"
    assert rendered["Data"]["B2"].data_type == "f"
    rendered.close()


def test_missing_formula_column_uses_resolved_header_and_exact_data_rows(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "formula-column", "schema_version": "1.0.0", "name": "Formula column",
        "sheets": [{
            "id": "data", "name": "Data", "header": {"auto_detect": True},
            "data_region": {"start_row": 5}, "primary_key": ["id"],
            "actions": {"fill_empty_from_standard": True},
            "columns": [
                {"name": "id", "title": "ID", "required": True},
                {
                    "name": "calc", "title": "Calc", "type": "decimal",
                    "compare": {"mode": "numeric", "formula_mode": "formula"},
                    "missing_column_action": "insert", "formula_template": "={row}*2",
                },
            ],
        }],
    })
    excel = tmp_path / "formula-column.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet["A1"] = "Report title"
    sheet["A3"] = "ID"
    sheet["A4"] = "Preamble"
    sheet["A5"] = "E1"
    sheet["A7"] = "E2"
    book.save(excel)
    standard = tmp_path / "formula-column.json"
    standard.write_text(json.dumps({"data": [
        {"id": "E1", "calc": "=5*2"},
        {"id": "E2", "calc": "=7*2"},
    ]}), encoding="utf-8")

    service = AuditService(tmp_path / "formula-column-runtime")
    job_id = service.create_job()
    service.run(job_id, excel, standard, rules)
    status = service.status(job_id)
    assert status["status"] == "completed", status
    assert status["summary"]["repairs_applied"] == 1
    manifest = json.loads(service.artifact(job_id, "manifest").read_text(encoding="utf-8"))
    insert = next(item for item in manifest["operations"] if item["type"] == "insert_column")
    assert insert["header_row"] == 3
    assert insert["data_start_row"] == 5
    assert insert["formula_rows"] == [5, 7]
    rendered = load_workbook(service.artifact(job_id, "excel"), data_only=False)
    data = rendered["Data"]
    assert data["B3"].value == "Calc"
    assert data["B4"].value is None
    assert data["B5"].value == "=5*2"
    assert data["B6"].value is None
    assert data["B7"].value == "=7*2"
    rendered.close()

    mismatch_standard = tmp_path / "formula-column-mismatch.json"
    mismatch_standard.write_text(json.dumps({"data": [
        {"id": "E1", "calc": "=5*2"},
        {"id": "E2", "calc": "=999"},
    ]}), encoding="utf-8")
    mismatch_service = AuditService(tmp_path / "formula-column-mismatch-runtime")
    mismatch_job = mismatch_service.create_job()
    mismatch_service.run(mismatch_job, excel, mismatch_standard, rules)
    mismatch_status = mismatch_service.status(mismatch_job)
    assert mismatch_status["status"] == "manual_review"
    assert "excel" not in mismatch_status["artifacts"]
    mismatch_report = json.loads(mismatch_service.artifact(mismatch_job, "json").read_text(encoding="utf-8"))
    mismatch = next(item for item in mismatch_report["differences"] if item["rule_id"] == "calc.formula_template_mismatch")
    assert mismatch["excel_row"] == 7
    assert mismatch["standard_raw_value"] == "=999"
    assert mismatch["standard_normalized_value"] == "=7*2"


def test_missing_formula_cache_routes_to_manual_review(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "formula-cache", "schema_version": "1.0.0", "name": "Formula Cache",
        "sheets": [{"id": "data", "name": "数据", "primary_key": ["id"], "columns": [
            {"name": "id", "title": "编号", "required": True},
            {"name": "amount", "title": "金额", "type": "decimal", "compare": {"formula_mode": "cached_value", "mode": "numeric"}},
        ]}],
    })
    excel, standard = tmp_path / "formula-cache.xlsx", tmp_path / "formula-cache.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "数据"
    sheet.append(["编号", "金额"])
    sheet.append(["E1", "=1+1"])
    book.save(excel)
    standard.write_text(json.dumps({"data": [{"id": "E1", "amount": "2"}]}), encoding="utf-8")
    service = AuditService(tmp_path / "formula-cache-runtime")
    job_id = service.create_job()
    service.run(job_id, excel, standard, rules)
    status = service.status(job_id)
    assert status["status"] == "manual_review"
    assert "formula_cached_value_missing" in " ".join(status["warnings"])


def test_formula_primary_key_is_excluded_from_record_matching(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "formula-key", "schema_version": "1.0.0", "name": "Formula key",
        "sheets": [{"id": "data", "name": "Data", "primary_key": ["id"], "columns": [
            {"name": "id", "title": "ID", "required": True},
            {"name": "value", "title": "Value"},
        ]}],
    })
    excel, standard = tmp_path / "formula-key.xlsx", tmp_path / "formula-key.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID", "Value"])
    sheet.append(['="E1"', "same"])
    book.save(excel)
    standard.write_text(json.dumps({"data": [{"id": "E1", "value": "same"}]}), encoding="utf-8")

    service = AuditService(tmp_path / "formula-key-runtime")
    job_id = service.create_job()
    service.run(job_id, excel, standard, rules)

    status = service.status(job_id)
    report = json.loads(service.artifact(job_id, "json").read_text(encoding="utf-8"))
    primary_key_differences = [
        item for item in report["differences"]
        if item.get("rule_id") == "id.formula_primary_key"
    ]
    assert status["status"] == "manual_review"
    assert "formula_primary_key:id" in " ".join(status["warnings"])
    assert len(primary_key_differences) == 1
    assert primary_key_differences[0]["type"] == "UNSUPPORTED_FEATURE"
    assert report["record_summary"]["matched"] == 0


def test_csv_standard_source_maps_display_headers_and_aliases(tmp_path):
    excel = tmp_path / "input.xlsx"
    standard = tmp_path / "standard.csv"
    _input(excel)
    standard.write_text("工号,员工姓名,工资,部门\nE001,张三,10000.00,技术部\nE003,王五,8000,财务部\n", encoding="utf-8-sig")
    service = AuditService(tmp_path / "runtime")
    job_id = service.create_job()
    service.run(job_id, excel, standard, load_rules(Path("configs/examples/employee-roster.yaml")))
    status = service.status(job_id)
    assert status["status"] == "completed"
    assert status["summary"]["matched_records"] == 2
    assert status["summary"]["mismatched_cells"] == 0


def test_row_number_primary_key_survives_uploaded_standard_canonicalization(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "row-number", "schema_version": "1.0.0", "name": "Row number",
        "sheets": [{
            "id": "data", "name": "Data", "primary_key_mode": "row_number",
            "columns": [{"name": "value", "title": "Value"}],
        }],
    })
    excel, standard = tmp_path / "row-number.xlsx", tmp_path / "row-number.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["Value"])
    sheet.append(["A"])
    sheet.append(["B"])
    book.save(excel)
    standard.write_text(json.dumps({"data": [
        {"__row_number__": 2, "value": "A"},
        {"__row_number__": 3, "value": "B"},
    ]}), encoding="utf-8")
    service = AuditService(tmp_path / "row-number-runtime")
    job_id = service.create_job()
    service.run(job_id, excel, standard, rules)

    status = service.status(job_id)
    assert status["status"] == "completed", status
    assert status["summary"]["matched_records"] == 2
    assert status["summary"]["differences"] == 0


def test_fuzzy_header_suggestion_requires_manual_review(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "fuzzy-header", "schema_version": "1.0.0", "name": "Fuzzy header",
        "sheets": [{"id": "data", "name": "Data", "primary_key": ["id"], "columns": [
            {"name": "id", "title": "Identifier", "required": True},
            {"name": "employee_name", "title": "Employee Name"},
        ]}],
    })
    excel, standard = tmp_path / "fuzzy.xlsx", tmp_path / "standard.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["Identifier", "Employee Nam"])
    sheet.append(["E1", "Alice"])
    book.save(excel)
    standard.write_text(json.dumps({"data": [{"id": "E1", "employee_name": "Alice"}]}), encoding="utf-8")
    service = AuditService(tmp_path / "fuzzy-runtime")
    job_id = service.create_job()
    service.run(job_id, excel, standard, rules)
    status = service.status(job_id)
    assert status["status"] == "manual_review"
    assert "excel" not in status["artifacts"]
    assert any("ambiguous_header:Employee Nam" in warning for warning in status["warnings"])


def test_duplicate_primary_key_is_never_auto_appended(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "duplicate-append-guard", "schema_version": "1.0.0", "name": "Duplicate append guard",
        "sheets": [{
            "id": "data", "name": "Data", "primary_key": ["id"],
            "actions": {"missing_record": "append_and_mark_green"},
            "columns": [
                {"name": "id", "title": "ID", "required": True},
                {"name": "name", "title": "Name"},
            ],
        }],
    })
    excel, standard = tmp_path / "duplicate.xlsx", tmp_path / "standard.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID", "Name"])
    sheet.append(["E1", "Alice"])
    sheet.append(["E1", "Alice duplicate"])
    book.save(excel)
    standard.write_text(json.dumps({"data": [{"id": "E1", "name": "Alice"}]}), encoding="utf-8")
    service = AuditService(tmp_path / "runtime")
    job_id = service.create_job()
    service.run(job_id, excel, standard, rules)
    assert service.status(job_id)["status"] == "completed"
    report = json.loads(service.artifact(job_id, "json").read_text(encoding="utf-8"))
    missing = next(item for item in report["differences"] if item["type"] == "MISSING_RECORD")
    assert missing["repair_status"] == "not_requested"
    assert missing["render_action"] == "report_only"
    assert "禁止自动追加" in missing["message"]
    rendered = load_workbook(service.artifact(job_id, "excel"))
    assert rendered["Data"].max_row == 3
    rendered.close()


def test_multiple_missing_records_append_to_distinct_consecutive_rows(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "multi-append", "schema_version": "1.0.0", "name": "Multi append",
        "sheets": [{
            "id": "data", "name": "Data", "primary_key": ["id"],
            "actions": {"missing_record": "append_and_mark_green"},
            "columns": [
                {"name": "id", "title": "ID", "required": True},
                {"name": "name", "title": "Name"},
            ],
        }],
    })
    excel, standard = tmp_path / "multi-append.xlsx", tmp_path / "multi-append.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID", "Name"])
    sheet.append(["E0", "Existing"])
    book.save(excel)
    standard.write_text(json.dumps({"data": [
        {"id": "E0", "name": "Existing"},
        {"id": "E1", "name": "First"},
        {"id": "E2", "name": "Second"},
        {"id": "E3", "name": "Third"},
    ]}), encoding="utf-8")

    service = AuditService(tmp_path / "multi-append-runtime")
    job_id = service.create_job()
    service.run(job_id, excel, standard, rules)
    status = service.status(job_id)
    assert status["status"] == "completed", status
    assert status["summary"]["missing_records"] == 3
    assert status["summary"]["repairs_applied"] == 3
    report = json.loads(service.artifact(job_id, "json").read_text(encoding="utf-8"))
    appended_rows = sorted(
        item["excel_row"] for item in report["differences"]
        if item["type"] == "MISSING_RECORD" and item["repair_status"] == "applied"
    )
    assert appended_rows == [3, 4, 5]
    rendered = load_workbook(service.artifact(job_id, "excel"))
    assert [rendered["Data"].cell(row, 1).value for row in range(2, 6)] == ["E0", "E1", "E2", "E3"]
    assert [rendered["Data"].cell(row, 2).value for row in range(2, 6)] == ["Existing", "First", "Second", "Third"]
    rendered.close()


def test_fuzzy_value_suggestion_requires_manual_review(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "fuzzy-value", "schema_version": "1.0.0", "name": "Fuzzy value",
        "sheets": [{"id": "data", "name": "Data", "primary_key": ["id"], "columns": [
            {"name": "id", "title": "ID", "required": True},
            {"name": "address", "title": "Address", "type": "fuzzy_string"},
        ]}],
    })
    excel, standard = tmp_path / "fuzzy-value.xlsx", tmp_path / "fuzzy-value.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID", "Address"])
    sheet.append(["E1", "Shanghai Pudong"])
    book.save(excel)
    standard.write_text(json.dumps({"data": [{"id": "E1", "address": "Shanghai Pudong New Area"}]}), encoding="utf-8")
    service = AuditService(tmp_path / "fuzzy-value-runtime")
    job_id = service.create_job()
    service.run(job_id, excel, standard, rules)
    assert service.status(job_id)["status"] == "manual_review"


def test_sheet_alias_is_preserved_in_differences_and_repairs(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "sheet-alias", "schema_version": "1.0.0", "name": "Alias",
        "sheets": [{
            "id": "data", "name": "Canonical", "aliases": ["Physical"], "primary_key": ["id"],
            "actions": {"overwrite_mismatch": True},
            "columns": [
                {"name": "id", "title": "ID", "required": True},
                {"name": "value", "title": "Value"},
            ],
        }],
    })
    excel, standard = tmp_path / "alias.xlsx", tmp_path / "standard.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "Physical"
    sheet.append(["ID", "Value"])
    sheet.append(["E1", "old"])
    book.save(excel)
    standard.write_text(json.dumps({"data": [{"id": "E1", "value": "new"}]}), encoding="utf-8")
    service = AuditService(tmp_path / "alias-runtime")
    job_id = service.create_job()
    service.run(job_id, excel, standard, rules)
    assert service.status(job_id)["status"] == "completed"
    report = json.loads(service.artifact(job_id, "json").read_text(encoding="utf-8"))
    assert {item["sheet_name"] for item in report["differences"]} == {"Physical"}
    rendered = load_workbook(service.artifact(job_id, "excel"))
    assert rendered["Physical"]["B2"].value == "new"
    rendered.close()


def test_multiple_physical_sheets_matching_one_rule_require_manual_review(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "ambiguous-sheet", "schema_version": "1.0.0", "name": "Ambiguous sheet",
        "sheets": [{
            "id": "data", "name": "Canonical", "aliases": ["Physical"], "primary_key": ["id"],
            "columns": [{"name": "id", "title": "ID", "required": True}],
        }],
    })
    excel, standard = tmp_path / "ambiguous-sheet.xlsx", tmp_path / "ambiguous-sheet.json"
    book = Workbook()
    canonical = book.active
    canonical.title = "Canonical"
    canonical.append(["ID"])
    canonical.append(["E1"])
    alias = book.create_sheet("Physical")
    alias.append(["ID"])
    alias.append(["E2"])
    book.save(excel)
    standard.write_text(json.dumps({"data": [{"id": "E1"}]}), encoding="utf-8")

    service = AuditService(tmp_path / "ambiguous-sheet-runtime")
    job_id = service.create_job()
    service.run(job_id, excel, standard, rules)

    status = service.status(job_id)
    assert status["status"] == "manual_review"
    assert "excel" not in status["artifacts"]
    assert any("ambiguous_sheet:Canonical|Physical" in warning for warning in status["warnings"])
    report = json.loads(service.artifact(job_id, "json").read_text(encoding="utf-8"))
    assert [item["type"] for item in report["differences"]] == ["AMBIGUOUS_SHEET"]


def test_development_renderer_inserts_and_fills_missing_column_on_sheet_alias(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "alias-insert", "schema_version": "1.0.0", "name": "Alias insert",
        "sheets": [{
            "id": "data", "name": "Canonical", "aliases": ["Physical"], "primary_key": ["id"],
            "actions": {"fill_empty_from_standard": True},
            "columns": [
                {"name": "id", "title": "ID", "required": True},
                {"name": "value", "title": "Value", "required": True, "missing_column_action": "insert"},
            ],
        }],
    })
    excel, standard = tmp_path / "alias-insert.xlsx", tmp_path / "alias-insert.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "Physical"
    sheet.append(["ID"])
    sheet.append(["E1"])
    book.save(excel)
    standard.write_text(json.dumps({"data": [{"id": "E1", "value": "new"}]}), encoding="utf-8")

    service = AuditService(
        tmp_path / "alias-insert-runtime",
        renderer=OpenPyxlDevelopmentRenderer(),
    )
    job_id = service.create_job()
    service.run(job_id, excel, standard, rules)

    assert service.status(job_id)["status"] == "completed"
    rendered = load_workbook(service.artifact(job_id, "excel"))
    assert [rendered["Physical"].cell(1, column).value for column in (1, 2)] == ["ID", "Value"]
    assert rendered["Physical"]["B2"].value == "new"
    rendered.close()


def test_insert_order_preserves_extra_column_around_canonical_fields(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "headers", "schema_version": "1.0.0", "name": "Headers",
        "sheets": [{"id": "data", "name": "数据", "primary_key": ["name"], "columns": [
            {"name": "name", "title": "姓名", "required": True},
            {"name": "id_code", "title": "身份证号", "required": True, "missing_column_action": "insert"},
            {"name": "phone", "title": "手机号"},
        ]}],
    })
    excel, standard = tmp_path / "headers.xlsx", tmp_path / "standard.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "数据"
    sheet.append(["姓名", "部门", "手机号"])
    sheet.append(["张三", "技术部", "00123"])
    book.save(excel)
    standard.write_text(json.dumps({"data": [{"name": "张三", "id_code": "ID1", "phone": "00123"}]}, ensure_ascii=False), encoding="utf-8")
    service = AuditService(tmp_path / "runtime")
    job_id = service.create_job()
    service.run(job_id, excel, standard, rules)
    assert service.status(job_id)["status"] == "completed"
    rendered = load_workbook(service.artifact(job_id, "excel"))
    assert [rendered["数据"].cell(1, index).value for index in range(1, 5)] == ["姓名", "身份证号", "部门", "手机号"]
    rendered.close()


def test_explicit_auto_repairs_are_applied_and_public_manifest_is_redacted(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "repairs", "schema_version": "1.0.0", "name": "Repairs",
        "sheets": [{
            "id": "data", "name": "数据", "primary_key": ["id"],
            "actions": {"rename_confirmed_alias": True, "fill_empty_from_standard": True, "overwrite_mismatch": True, "missing_record": "append_and_mark_green"},
            "columns": [
                {"name": "id", "title": "编号", "aliases": ["代码"], "required": True, "validation": {"nullable": False, "unique": True}},
                {"name": "name", "title": "姓名", "required": True},
                {"name": "amount", "title": "金额", "type": "decimal"},
                {"name": "note", "title": "备注", "missing_column_action": "insert"},
            ],
        }],
    })
    excel, standard = tmp_path / "repairs.xlsx", tmp_path / "standard.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "数据"
    sheet.append(["代码", "姓名", "金额"])
    sheet.append(["E1", "Alice", "9"])
    book.save(excel)
    standard.write_text(json.dumps({"data": [
        {"id": "E1", "name": "Alice", "amount": "10", "note": "=SAFE_TEXT"},
        {"id": "E2", "name": "Bob", "amount": "20", "note": "@SAFE_TEXT"},
    ]}), encoding="utf-8")
    service = AuditService(tmp_path / "runtime")
    job_id = service.create_job()
    service.run(job_id, excel, standard, rules)
    status = service.status(job_id)
    assert status["status"] == "completed", status
    assert status["summary"]["repairs_planned"] == 5
    assert status["summary"]["repairs_applied"] == 5
    rendered = load_workbook(service.artifact(job_id, "excel"), data_only=False)
    result = rendered["数据"]
    assert [result.cell(1, index).value for index in range(1, 5)] == ["编号", "姓名", "金额", "备注"]
    assert str(result["C2"].value) == "10"
    assert result["D2"].value == "=SAFE_TEXT" and result["D2"].data_type != "f"
    assert result["A3"].value == "E2"
    assert result["D3"].value == "@SAFE_TEXT" and result["D3"].data_type != "f"
    assert all(part in result["A1"].comment.text for part in ["原值：代码", "标准值：编号", "规则：id.rename_confirmed_alias"])
    assert all(part in result["C2"].comment.text for part in ["原值：9", "标准值：10", "规则：amount.overwrite_mismatch"])
    assert all(part in result["D1"].comment.text for part in ["原值：（缺失）", "标准值：备注", "note.missing_column"])
    assert all(part in result["D2"].comment.text for part in ["原值：null", "标准值：=SAFE_TEXT", "规则：note.fill_empty_from_standard"])
    assert all(part in result["A3"].comment.text for part in ["原值：（缺失）", '"id":"E2"', '"note":"@SAFE_TEXT"', "规则：missing_record.append"])
    embedded_summary = {
        row[0]: row[1]
        for row in rendered["核验报告"].iter_rows(min_row=2, values_only=True)
        if row[0]
    }
    assert int(embedded_summary["repairs_planned"]) == 5
    assert int(embedded_summary["repairs_applied"]) == 5
    assert int(embedded_summary["repair_failures"]) == 0
    embedded_repair_statuses = sorted(
        row[12]
        for row in rendered["核验报告"].iter_rows(values_only=True)
        if len(row) > 12 and row[12] in {"not_requested", "planned", "applied", "failed"}
    )
    metadata = {
        str(row[0]): str(row[1])
        for row in rendered["__ExcelAuditorMetadata"].iter_rows(min_row=2, values_only=True)
        if row[0]
    }
    rendered.close()
    report = json.loads(service.artifact(job_id, "json").read_text(encoding="utf-8"))
    repaired = [item for item in report["differences"] if item["repair_status"] == "applied"]
    assert len(repaired) == 5
    assert embedded_repair_statuses == sorted(item["repair_status"] for item in report["differences"])
    assert metadata["schema_id"] == report["schema_id"]
    assert metadata["schema_version"] == report["schema_version"]
    assert metadata["schema_sha256"] == report["schema_sha256"]
    assert metadata["standard_snapshot_id"] == report["standard_snapshot_id"]
    assert metadata["standard_sha256"] == report["standard_sha256"]
    assert metadata["input_sha256"] == report["input_sha256"]
    assert len(metadata["result_content_sha256"]) == 64
    public_manifest = json.loads(service.artifact(job_id, "manifest").read_text(encoding="utf-8"))
    serialized = json.dumps(public_manifest, ensure_ascii=False)
    assert "SAFE_TEXT" not in serialized and '"value"' not in serialized
    redacted_repairs = [
        operation for operation in public_manifest["operations"]
        if operation["type"] in {"set_cell", "set_cell_after_insert", "append_row"}
    ]
    assert redacted_repairs and all(operation["comment_values_redacted"] is True for operation in redacted_repairs)
    assert all("原值和标准值已脱敏" in operation["comment"] for operation in redacted_repairs)
    assert any("规则：amount.overwrite_mismatch" in operation["comment"] for operation in redacted_repairs)


def test_higher_priority_invalid_cell_is_not_overwritten_by_extra_row_mark(tmp_path):
    rules = RuleSet.model_validate({
        "schema_id": "color-priority", "schema_version": "1.0.0", "name": "Color priority",
        "sheets": [{
            "id": "data", "name": "Data", "primary_key": ["id"],
            "columns": [
                {"name": "id", "title": "ID", "required": True},
                {"name": "amount", "title": "Amount", "type": "decimal", "validation": {"min": "0"}},
            ],
        }],
    })
    excel, standard = tmp_path / "color-priority.xlsx", tmp_path / "color-priority.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID", "Amount"])
    sheet.append(["EXTRA", "-1"])
    sheet["B2"].comment = Comment("用户原批注", "Alice")
    book.save(excel)
    standard.write_text(json.dumps({"data": []}), encoding="utf-8")
    matched_excel, matched_standard = tmp_path / "color-priority-matched.xlsx", tmp_path / "color-priority-matched.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID", "Amount"])
    sheet.append(["E1", "-1"])
    sheet["B2"].comment = Comment("用户原批注", "Alice")
    book.save(matched_excel)
    matched_standard.write_text(json.dumps({"data": [{"id": "E1", "amount": "5"}]}), encoding="utf-8")
    renderers = [("development", OpenPyxlDevelopmentRenderer())]
    command = os.environ.get("EXCEL_RENDERER_COMMAND")
    if command:
        renderers.append(("production", DotNetOpenXmlRenderer(Path(command))))

    for name, renderer in renderers:
        service = AuditService(tmp_path / f"color-priority-{name}", renderer=renderer)
        job_id = service.create_job()
        service.run(job_id, excel, standard, rules)
        status = service.status(job_id)
        assert status["status"] == "completed", status
        report = json.loads(service.artifact(job_id, "json").read_text(encoding="utf-8"))
        assert {item["type"] for item in report["differences"]} == {"EXTRA_RECORD", "VALIDATION_ERROR"}
        rendered = load_workbook(service.artifact(job_id, "excel"), data_only=False)
        result = rendered["Data"]
        assert result["A2"].fill.fgColor.rgb.endswith(rules.colors.extra)
        assert result["B2"].fill.fgColor.rgb.endswith(rules.colors.invalid)
        assert result["B2"].comment.author == "Alice"
        assert "用户原批注\n\n[Excel Auditor]\n" in result["B2"].comment.text
        assert "最小值" in result["B2"].comment.text
        rendered.close()
        manifest = json.loads(service.artifact(job_id, "manifest").read_text(encoding="utf-8"))
        marks = [operation for operation in manifest["operations"] if operation["type"] in {"mark_row", "mark_cell"}]
        assert [operation["type"] for operation in marks] == ["mark_row", "mark_cell"]

        matched_service = AuditService(tmp_path / f"color-priority-matched-{name}", renderer=renderer)
        matched_job = matched_service.create_job()
        matched_service.run(matched_job, matched_excel, matched_standard, rules)
        matched_status = matched_service.status(matched_job)
        assert matched_status["status"] == "completed", matched_status
        matched_rendered = load_workbook(matched_service.artifact(matched_job, "excel"), data_only=False)
        matched_cell = matched_rendered["Data"]["B2"]
        assert matched_cell.fill.fgColor.rgb.endswith(rules.colors.invalid)
        assert matched_cell.comment.author == "Alice"
        assert "用户原批注\n\n[Excel Auditor]\n" in matched_cell.comment.text
        assert "最小值" in matched_cell.comment.text
        assert "字段值与标准数据不一致" in matched_cell.comment.text
        matched_rendered.close()
        matched_manifest = json.loads(matched_service.artifact(matched_job, "manifest").read_text(encoding="utf-8"))
        matched_marks = [operation for operation in matched_manifest["operations"] if operation["type"] == "mark_cell"]
        assert len(matched_marks) == 1 and matched_marks[0]["cell"] == "B2"


def test_soft_delete_delays_then_purges_terminal_job(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_AUDITOR_DELETE_DELAY_DAYS", "0")
    rules = RuleSet.model_validate({
        "schema_id": "retention", "schema_version": "1.0.0", "name": "Retention",
        "sheets": [{"id": "data", "name": "Data", "primary_key": ["id"], "columns": [{"name": "id", "title": "ID", "required": True}]}],
    })
    excel, standard = tmp_path / "retention.xlsx", tmp_path / "retention.json"
    book = Workbook()
    sheet = book.active
    sheet.title = "Data"
    sheet.append(["ID"])
    sheet.append(["E1"])
    book.save(excel)
    standard.write_text(json.dumps({"data": [{"id": "E1"}]}), encoding="utf-8")
    service = AuditService(tmp_path / "retention-runtime")
    job_id = service.create_job()
    service.run(job_id, excel, standard, rules)
    deleted = service.soft_delete(job_id, "tester")
    assert deleted["deleted_at"] and service.job_directory(job_id).is_dir()
    assert service.purge_expired() == [job_id]
    assert not (service.jobs / job_id).exists()


def test_tenant_active_job_quota_is_atomic_across_concurrent_creates(tmp_path, monkeypatch):
    monkeypatch.setenv("EXCEL_AUDITOR_MAX_ACTIVE_JOBS_PER_TENANT", "2")
    service = AuditService(tmp_path / "quota-runtime")
    def create():
        try:
            return service.create_job(tenant_id="tenant-a")
        except ValueError as exc:
            assert "TENANT_QUOTA_EXCEEDED" in str(exc)
            return None
    with ThreadPoolExecutor(max_workers=8) as pool:
        results = list(pool.map(lambda _index: create(), range(8)))
    assert len([job_id for job_id in results if job_id]) == 2
