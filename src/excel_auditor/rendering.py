from __future__ import annotations

import hashlib
import json
import os
import posixpath
import re
import subprocess
import shutil
import tempfile
import zipfile
from xml.etree import ElementTree
from abc import ABC, abstractmethod
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from .engine import ComparisonResult
from .models import Difference, DifferenceType, RuleSet
from .workbook import WorkbookSnapshot, sha256_file


RENDER_ACTION_PRIORITY = {
    "mark_red": 1,
    "mark_row_red": 1,
    "mark_yellow": 2,
    "mark_orange": 3,
    "mark_purple": 4,
    "mark_row_purple": 4,
}


@dataclass
class RenderResult:
    output_path: Path
    sha256: str
    warnings: list[str] = field(default_factory=list)
    operation_results: list[dict[str, Any]] = field(default_factory=list)


class ExcelRenderer(ABC):
    @abstractmethod
    def render(self, source: Path, destination: Path, workbook: WorkbookSnapshot, rules: RuleSet, comparison: ComparisonResult, report_payload: dict[str, Any]) -> RenderResult:
        raise NotImplementedError


class OpenPyxlDevelopmentRenderer(ExcelRenderer):
    """Conservative development fallback; structural edits are refused on risky sheets."""

    def render(self, source: Path, destination: Path, workbook: WorkbookSnapshot, rules: RuleSet, comparison: ComparisonResult, report_payload: dict[str, Any]) -> RenderResult:
        if source.suffix.lower() == ".xlsm":
            raise RuntimeError("RENDER_FAILED: development renderer does not modify macro-enabled workbooks")
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source, destination)
        destination.chmod(0o600)
        book = load_workbook(destination, data_only=False, keep_links=False)
        warnings: list[str] = []
        operation_count = 0
        color_by_action = {
            "mark_red": rules.colors.extra,
            "mark_row_red": rules.colors.extra,
            "mark_yellow": rules.colors.mismatch,
            "mark_orange": rules.colors.invalid,
            "mark_purple": rules.colors.ambiguous,
            "mark_row_purple": rules.colors.ambiguous,
        }
        differences_by_id = {item.difference_id: item for item in comparison.differences}
        renderable_differences = [
            difference for difference in comparison.differences
            if difference.render_action in color_by_action
        ]
        for difference in sorted(
            renderable_differences,
            key=lambda item: (
                RENDER_ACTION_PRIORITY[item.render_action],
                item.sheet_name,
                item.excel_row or 0,
                item.cell or "",
                item.difference_id,
            ),
        ):
            if difference.sheet_name not in book.sheetnames:
                continue
            sheet = book[difference.sheet_name]
            color = color_by_action.get(difference.render_action)
            fill = PatternFill("solid", fgColor=color)
            if difference.cell:
                cell = sheet[difference.cell]
                cell.fill = fill
                _set_audit_comment(cell, _comment(difference))
                operation_count += 1
            elif difference.excel_row:
                for cell in sheet[difference.excel_row]:
                    cell.fill = fill
                _set_audit_comment(sheet.cell(difference.excel_row, 1), _comment(difference))
                operation_count += 1
        for repair in comparison.repairs:
            if repair.type not in {"set_cell", "set_number_format"} or repair.sheet_name not in book.sheetnames or not repair.cell:
                continue
            cell = book[repair.sheet_name][repair.cell]
            if repair.type == "set_cell":
                _set_safe_value(cell, repair.value)
                cell.fill = PatternFill("solid", fgColor=rules.colors.inserted)
            else:
                cell.number_format = str(repair.value)
            _set_audit_comment(
                cell,
                _repair_provenance_comment("自动修复", repair.rule_id, differences_by_id.get(repair.difference_id)),
            )
            operation_count += 1
        for sheet_rule in rules.sheets:
            physical_name = next(
                (name for name in [sheet_rule.name, *sheet_rule.aliases] if name in book.sheetnames),
                None,
            )
            if physical_name is None:
                continue
            sheet = book[physical_name]
            risky = workbook.sheets[physical_name].risky_features
            missing = [item for item in comparison.differences if item.sheet_id == sheet_rule.id and item.type == DifferenceType.MISSING_HEADER and item.render_action == "insert_and_mark_green"]
            if missing and risky:
                warnings.append(f"{physical_name}: development renderer refused column insertion because of {', '.join(risky)}")
                continue
            operation_count += _insert_missing_columns(sheet, sheet_rule, missing, rules.colors.inserted)
        for repair in comparison.repairs:
            if repair.sheet_name not in book.sheetnames:
                continue
            sheet = book[repair.sheet_name]
            if repair.type == "set_field" and repair.excel_row and repair.canonical_field:
                positions = _canonical_positions(sheet, next(item for item in rules.sheets if item.id == repair.sheet_id))
                if repair.canonical_field not in positions:
                    raise RuntimeError(f"RENDER_FAILED: repaired field has no final column: {repair.canonical_field}")
                cell = sheet.cell(repair.excel_row, positions[repair.canonical_field])
                _set_safe_value(cell, repair.value)
                cell.fill = PatternFill("solid", fgColor=rules.colors.inserted)
                _set_audit_comment(
                    cell,
                    _repair_provenance_comment("自动修复", repair.rule_id, differences_by_id.get(repair.difference_id)),
                )
                operation_count += 1
            elif repair.type == "append_record" and repair.excel_row and repair.values is not None:
                sheet_rule = next(item for item in rules.sheets if item.id == repair.sheet_id)
                positions = _canonical_positions(sheet, sheet_rule)
                for field, value in repair.values.items():
                    if field not in positions:
                        continue
                    cell = sheet.cell(repair.excel_row, positions[field])
                    _set_safe_value(cell, value)
                    cell.fill = PatternFill("solid", fgColor=rules.colors.inserted)
                _set_audit_comment(
                    sheet.cell(repair.excel_row, 1),
                    _repair_provenance_comment("自动追加标准记录", repair.rule_id, differences_by_id.get(repair.difference_id)),
                )
                operation_count += 1
        if "核验报告" in book.sheetnames:
            del book["核验报告"]
        report_sheet = book.create_sheet("核验报告")
        report_sheet.append(["项目", "值"])
        summary = report_payload.get("summary", {})
        for key, value in summary.items():
            report_sheet.append([key, value])
        report_sheet.append([])
        report_sheet.append(["类型", "工作表", "单元格", "业务主键", "说明"])
        for item in report_payload.get("differences", []):
            report_sheet.append([item["type"], item["sheet_name"], item.get("cell"), json.dumps(item.get("business_key"), ensure_ascii=False), item["message"]])
        report_sheet.sheet_state = "visible"
        operation_count += 1
        if "__ExcelAuditorMetadata" in book.sheetnames:
            del book["__ExcelAuditorMetadata"]
        metadata_sheet = book.create_sheet("__ExcelAuditorMetadata")
        metadata_sheet.append(["key", "value"])
        metadata = [
            ("job_id", report_payload.get("job_id")),
            ("schema_id", report_payload.get("schema_id")),
            ("schema_version", report_payload.get("schema_version")),
            ("schema_sha256", report_payload.get("schema_sha256")),
            ("standard_snapshot_id", report_payload.get("standard_snapshot_id")),
            ("standard_sha256", report_payload.get("standard_sha256")),
            ("input_sha256", report_payload.get("input_sha256")),
            ("result_content_sha256", ""),
            ("operation_count", operation_count),
        ]
        for key, value in metadata:
            metadata_sheet.append([key, "" if value is None else str(value)])
        metadata_sheet.sheet_state = "veryHidden"
        book.save(destination)
        book.close()
        metadata_entry = _worksheet_package_entry(destination, "__ExcelAuditorMetadata")
        _set_result_content_hash(destination, metadata_entry)
        load_workbook(destination, read_only=True, data_only=False).close()
        applied = [{"difference_id": item.difference_id, "status": "applied"} for item in comparison.differences if item.repair_status == "planned"]
        return RenderResult(destination, sha256_file(destination), warnings, applied)


class DotNetOpenXmlRenderer(ExcelRenderer):
    def __init__(self, command: Path) -> None:
        if not command.is_file():
            raise FileNotFoundError(f"renderer command not found: {command}")
        self.command = command

    def self_check(self) -> str:
        completed = subprocess.run(
            [str(self.command), "--version"],
            capture_output=True,
            text=True,
            timeout=10,
            check=False,
        )
        if completed.returncode != 0:
            raise RuntimeError("renderer self-check failed")
        try:
            payload = json.loads(completed.stdout)
        except json.JSONDecodeError as exc:
            raise RuntimeError("renderer self-check returned invalid JSON") from exc
        version = payload.get("version")
        if payload.get("name") != "ExcelRenderer" or not isinstance(version, str) or not re.fullmatch(r"\d+\.\d+\.\d+", version):
            raise RuntimeError("renderer self-check returned an invalid identity or version")
        return version

    def render(self, source: Path, destination: Path, workbook: WorkbookSnapshot, rules: RuleSet, comparison: ComparisonResult, report_payload: dict[str, Any]) -> RenderResult:
        manifest = destination.parent / "render-manifest.private.json"
        completed = subprocess.run(
            [str(self.command), "--input", str(source), "--output", str(destination), "--manifest", str(manifest)],
            capture_output=True,
            text=True,
            timeout=rules.workbook.processing_timeout_seconds,
            check=False,
        )
        if completed.returncode != 0:
            try:
                failure = json.loads(completed.stderr)
            except json.JSONDecodeError:
                raise RuntimeError("RENDER_FAILED: renderer process failed without structured JSON")
            error_code = failure.get("error_code")
            known_codes = {
                "ARGUMENT_INVALID",
                "MANIFEST_OR_STRUCTURE_INVALID",
                "OUTPUT_VERIFICATION_FAILED",
                "RENDER_FAILED",
                "UNSUPPORTED_FEATURE",
            }
            if error_code not in known_codes:
                raise RuntimeError("RENDER_FAILED: renderer returned an unknown error code")
            raise RuntimeError(f"{error_code}: renderer process failed")
        try:
            payload = json.loads(completed.stdout)
        except json.JSONDecodeError as exc:
            raise RuntimeError("RENDER_FAILED: renderer returned invalid JSON") from exc
        digest = sha256_file(destination)
        if not payload.get("success") or payload.get("output_sha256") != digest:
            raise RuntimeError("OUTPUT_VERIFICATION_FAILED: renderer hash mismatch")
        operation_results = payload.get("operation_results", [])
        if not isinstance(operation_results, list):
            raise RuntimeError("RENDER_FAILED: renderer returned invalid operation results")
        return RenderResult(destination, digest, operation_results=operation_results)


def _insert_missing_columns(sheet: Any, sheet_rule: Any, missing: list[Difference], color: str) -> int:
    missing_by_name = {item.canonical_field: item for item in missing}
    missing_names = set(missing_by_name)
    if not missing_names:
        return 0
    header_row = sheet_rule.header.row
    known_positions: dict[str, int] = {}
    normalized_headers = {str(sheet.cell(header_row, col).value).strip(): col for col in range(1, sheet.max_column + 1)}
    for rule in sheet_rule.columns:
        for candidate in [rule.title, rule.name, *rule.aliases]:
            if candidate in normalized_headers:
                known_positions[rule.name] = normalized_headers[candidate]
                break
    plans: list[tuple[int, int, Any]] = []
    for index, rule in enumerate(sheet_rule.columns):
        if rule.name not in missing_names:
            continue
        previous = [known_positions[previous_rule.name] for previous_rule in sheet_rule.columns[:index] if previous_rule.name in known_positions]
        before = previous[-1] + 1 if previous else 1
        plans.append((before, index, rule))
    for before, _index, rule in sorted(plans, key=lambda item: (item[0], item[1]), reverse=True):
        sheet.insert_cols(before)
        cell = sheet.cell(header_row, before, rule.title)
        if before > 1:
            source = sheet.cell(header_row, before - 1)
            cell.font, cell.border, cell.alignment, cell.number_format, cell.protection = copy(source.font), copy(source.border), copy(source.alignment), source.number_format, copy(source.protection)
        cell.fill = PatternFill("solid", fgColor=color)
        difference = missing_by_name[rule.name]
        _set_audit_comment(
            cell,
            _repair_provenance_comment("自动插入缺失表头", difference.rule_id or f"{rule.name}.missing_column", difference),
        )
    return len(plans)


def _workbook_content_hash(path: Path, excluded_entry: str) -> str:
    """Hash all uncompressed package parts except the self-referential metadata sheet."""
    digest = hashlib.sha256()
    with zipfile.ZipFile(path) as archive:
        for name in sorted(archive.namelist()):
            if name == excluded_entry:
                continue
            digest.update(name.encode("utf-8"))
            digest.update(b"\0")
            digest.update(archive.read(name))
            digest.update(b"\0")
    return digest.hexdigest()


def _worksheet_package_entry(path: Path, sheet_name: str) -> str:
    spreadsheet_namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    office_relationship_namespace = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    package_relationship_namespace = "http://schemas.openxmlformats.org/package/2006/relationships"
    with zipfile.ZipFile(path) as archive:
        workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        sheet = workbook.find(f".//{{{spreadsheet_namespace}}}sheet[@name='{sheet_name}']")
        if sheet is None:
            raise RuntimeError(f"OUTPUT_VERIFICATION_FAILED: worksheet is missing: {sheet_name}")
        relationship_id = sheet.get(f"{{{office_relationship_namespace}}}id")
        relationships = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        relationship = relationships.find(f".//{{{package_relationship_namespace}}}Relationship[@Id='{relationship_id}']")
        if relationship is None or not relationship.get("Target"):
            raise RuntimeError(f"OUTPUT_VERIFICATION_FAILED: worksheet relationship is missing: {sheet_name}")
        target = str(relationship.get("Target"))
        return target.lstrip("/") if target.startswith("/") else posixpath.normpath(posixpath.join("xl", target))


def _set_result_content_hash(path: Path, metadata_entry: str) -> None:
    content_hash = _workbook_content_hash(path, metadata_entry)
    spreadsheet_namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ElementTree.register_namespace("", spreadsheet_namespace)
    temporary_name: str | None = None
    try:
        with tempfile.NamedTemporaryFile(dir=path.parent, prefix=f".{path.name}.", suffix=".tmp", delete=False) as temporary:
            temporary_name = temporary.name
        with zipfile.ZipFile(path) as source, zipfile.ZipFile(temporary_name, "w") as destination:
            destination.comment = source.comment
            replaced = False
            for info in source.infolist():
                payload = source.read(info.filename)
                if info.filename == metadata_entry:
                    replaced = True
                    root = ElementTree.fromstring(payload)
                    cell = root.find(f".//{{{spreadsheet_namespace}}}c[@r='B9']")
                    if cell is None:
                        raise RuntimeError("OUTPUT_VERIFICATION_FAILED: metadata result hash cell is missing")
                    cell.clear()
                    cell.set("r", "B9")
                    cell.set("t", "inlineStr")
                    inline = ElementTree.SubElement(cell, f"{{{spreadsheet_namespace}}}is")
                    text = ElementTree.SubElement(inline, f"{{{spreadsheet_namespace}}}t")
                    text.text = content_hash
                    payload = ElementTree.tostring(root, encoding="utf-8", xml_declaration=True)
                destination.writestr(info, payload)
            if not replaced:
                raise RuntimeError("OUTPUT_VERIFICATION_FAILED: metadata worksheet package part is missing")
        os.replace(temporary_name, path)
        temporary_name = None
        path.chmod(0o600)
        if _workbook_content_hash(path, metadata_entry) != content_hash:
            raise RuntimeError("OUTPUT_VERIFICATION_FAILED: development renderer content hash mismatch")
    finally:
        if temporary_name is not None:
            Path(temporary_name).unlink(missing_ok=True)


def _canonical_positions(sheet: Any, sheet_rule: Any) -> dict[str, int]:
    positions: dict[str, int] = {}
    normalized_headers = {str(sheet.cell(sheet_rule.header.row, col).value).strip(): col for col in range(1, sheet.max_column + 1)}
    for rule in sheet_rule.columns:
        for candidate in [rule.title, rule.name, *rule.aliases]:
            if candidate in normalized_headers:
                positions[rule.name] = normalized_headers[candidate]
                break
    return positions


def _set_safe_value(cell: Any, value: Any) -> None:
    cell.value = value
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        cell.data_type = "s"


def _comment(item: Difference) -> str:
    parts = [item.message]
    if item.excel_raw_value is not None:
        parts.append(f"Excel值：{item.excel_raw_value}")
    if item.standard_raw_value is not None:
        parts.append(f"标准值：{item.standard_raw_value}")
    if item.rule_id:
        parts.append(f"规则：{item.rule_id}")
    return "；".join(parts)[:32000]


def _set_audit_comment(cell: Any, text: str) -> None:
    audit_marker = "\n\n[Excel Auditor]\n"
    author = "Excel Auditor"
    if cell.comment is not None:
        existing_text = cell.comment.text or ""
        author = cell.comment.author or author
        if cell.comment.author == "Excel Auditor":
            audit_text = _merge_audit_comment_text(existing_text, text)
            output = audit_text
        else:
            original, marker, previous_audit = existing_text.partition(audit_marker)
            audit_text = _merge_audit_comment_text(previous_audit if marker else "", text)
            output = original + audit_marker + audit_text
    else:
        output = text
    if len(output.encode("utf-16-le")) // 2 > 32767:
        raise RuntimeError("UNSUPPORTED_FEATURE: existing comment leaves insufficient room for an audit comment without data loss")
    cell.comment = Comment(output, author)


def _merge_audit_comment_text(existing: str, next_text: str) -> str:
    if not existing:
        return next_text
    if existing == next_text or next_text in existing.split(" | "):
        return existing
    return f"{existing} | {next_text}"


def _repair_provenance_comment(prefix: str, rule_id: str, difference: Difference | None) -> str:
    """Build a deterministic, privacy-safe repair comment.

    Difference values have already passed the field masking policy. Never fall
    back to the actual repair payload here because it may contain sensitive
    business data that is intentionally absent from reports and comments.
    """
    if difference is None:
        original = standard = "（报告中不可用）"
    else:
        original = "（缺失）" if difference.type in {DifferenceType.MISSING_HEADER, DifferenceType.MISSING_RECORD} else _comment_value(difference.excel_raw_value)
        standard_source = difference.standard_raw_value
        if standard_source is None and difference.standard_normalized_value is not None:
            standard_source = difference.standard_normalized_value
        standard = _comment_value(standard_source)
    return f"{prefix}；原值：{original}；标准值：{standard}；规则：{rule_id}"


def _comment_value(value: Any) -> str:
    if isinstance(value, str):
        text = value
    else:
        text = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)
    # Two values of at most 7,000 Unicode code points remain below Excel's
    # 32,767 UTF-16 code-unit comment limit even for surrogate-pair characters.
    return text if len(text) <= 7000 else f"{text[:6999]}…"
