from __future__ import annotations

import json
import shutil
import time
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from ..ids import new_ulid
from ..models import RuleSet
from ..persistence import DatabaseRepository
from ..rendering import DotNetOpenXmlRenderer, _set_result_content_hash, _worksheet_package_entry
from ..service import AuditService, JobCancelled
from ..workbook import inspect_workbook, sha256_file
from .catalog import CatalogAdapter, FrozenCatalogAdapter
from .models import ProductNormalizationResult, ProductReviewDecision
from .normalizer import normalize_product_workbook


class ProductWorkflowService:
    """Execute and persist a versioned dynamic-product normalization analysis."""

    def __init__(self, audit_service: AuditService, database: DatabaseRepository | None = None) -> None:
        self.audit_service = audit_service
        self.database = database

    def run(
        self,
        job_id: str,
        excel_path: Path,
        rules: RuleSet,
        catalog: CatalogAdapter,
        *,
        tenant_id: str = "local",
        actor_id: str = "local",
        confirmed_aliases: dict[str, str] | None = None,
        confirmed_aliases_by_category: dict[str, dict[str, str]] | None = None,
        category_overrides: dict[int, str] | None = None,
        forced_extra_columns: set[int] | None = None,
        parent_revision_id: str | None = None,
    ) -> None:
        directory = self.audit_service.job_directory(job_id)
        current = self.audit_service.status(job_id)
        self.audit_service._write_status(job_id, {
            **current,
            "status": "validating",
            "progress": 15,
            "completed_at": None,
            "workflow": "product_normalization",
            "schema_id": rules.schema_id,
            "schema_version": rules.schema_version,
        })
        workbook = None
        result = None
        started = time.monotonic()

        def checkpoint() -> None:
            self.audit_service._check_cancelled(job_id)
            if time.monotonic() - started > rules.workbook.processing_timeout_seconds:
                raise TimeoutError("product normalization exceeded its processing timeout")

        try:
            checkpoint()
            workbook = inspect_workbook(excel_path, rules)
            checkpoint()
            self.audit_service._write_status(job_id, {
                **self.audit_service.status(job_id),
                "status": "catalog_resolving",
                "progress": 40,
                "input_sha256": workbook.sha256,
            })
            result = normalize_product_workbook(
                workbook,
                rules,
                catalog,
                confirmed_aliases=confirmed_aliases,
                confirmed_aliases_by_category=confirmed_aliases_by_category,
                category_overrides=category_overrides,
                forced_extra_columns=forced_extra_columns,
                checkpoint=checkpoint,
            )
            checkpoint()
            payload = result.model_dump(mode="json")
            result_name = "product-result.json"
            _write_json_atomic(directory / result_name, payload)
            issues_name = "product-issues.jsonl"
            _write_jsonl_atomic(
                directory / issues_name,
                [issue.model_dump(mode="json") for issue in result.issues],
            )
            if self.database:
                self.database.save_product_category_snapshot(result.category_catalog_snapshot, tenant_id)
            for snapshot in result.catalog_snapshots:
                if self.database:
                    self.database.save_product_catalog_snapshot(snapshot, tenant_id)
            if self.database:
                revision = self.database.create_product_revision(
                    job_id,
                    result,
                    actor_id=actor_id,
                    parent_revision_id=parent_revision_id,
                    tenant_id=tenant_id,
                )
                reviews = self.database.list_product_reviews(
                    job_id,
                    revision_id=revision["revision_id"],
                    tenant_id=tenant_id,
                )
            else:
                previous_state = self._state(job_id) if (directory / "product-workflow.json").is_file() else {}
                previous_revision = previous_state.get("current_revision", {})
                next_number = int(previous_revision.get("revision_number", 0)) + 1
                revision = {
                    "revision_id": new_ulid("prev_"),
                    "revision_number": next_number,
                    "parent_revision_id": parent_revision_id or previous_revision.get("revision_id"),
                    "status": "manual_review" if result.requires_manual_review else "ready",
                }
                reviews = [
                    {
                        "review_id": new_ulid("review_"),
                        "job_id": job_id,
                        "revision_id": revision["revision_id"],
                        "review_key": item.key,
                        "review_type": item.review_type,
                        "status": "pending",
                        "payload": item.model_dump(mode="json"),
                        "decision": None,
                        "created_at": current["created_at"],
                        "decided_at": None,
                        "decided_by": None,
                    }
                    for item in result.review_items
                ]
            previous_state = self._state(job_id) if (directory / "product-workflow.json").is_file() else {}
            history = list(previous_state.get("revision_history", []))
            history.append(revision)
            state = {
                "job_id": job_id,
                "current_revision": revision,
                "revision_history": history,
                "reviews": reviews,
            }
            _write_json_atomic(directory / "product-workflow.json", state)
            _write_json_atomic(directory / f"product-result-r{revision['revision_number']}.json", payload)
            artifacts = {"product_result": result_name, "product_issues": issues_name}
            artifact_names = [excel_path.name, result_name, issues_name, f"product-result-r{revision['revision_number']}.json", "product-workflow.json"]
            output_sha256 = None
            status = "manual_review" if result.requires_manual_review else "completed"
            if not workbook.report_only:
                checkpoint()
                status = "rendering"
                self.audit_service._write_status(job_id, {
                    **self.audit_service.status(job_id),
                    "status": status,
                    "progress": 75,
                })
                destination = directory / f"product-result{excel_path.suffix.lower()}"
                manifest_name = "product-render-manifest.json"
                manifest_path = directory / manifest_name
                manifest = {
                    "manifest_version": "1.0",
                    "job_id": job_id,
                    "input_sha256": workbook.sha256,
                    "metadata": {
                        "schema_id": rules.schema_id,
                        "schema_version": rules.schema_version,
                        "schema_sha256": rules.content_sha256,
                    },
                    "operations": [{
                        "type": "add_or_replace_product_sheets",
                        "source_json": result_name,
                    }],
                }
                _write_json_atomic(manifest_path, manifest)
                if isinstance(self.audit_service.renderer, DotNetOpenXmlRenderer):
                    render = self.audit_service.renderer.render_manifest(
                        excel_path,
                        destination,
                        manifest_path,
                        rules.workbook.processing_timeout_seconds,
                    )
                    output_sha256 = render.sha256
                else:
                    output_sha256 = _render_product_development(
                        excel_path,
                        destination,
                        result,
                        job_id=job_id,
                        rules=rules,
                    )
                checkpoint()
                artifacts.update({"product_excel": destination.name, "product_manifest": manifest_name})
                artifact_names.extend([destination.name, manifest_name])
                status = "manual_review" if result.requires_manual_review else "completed"
            object_keys = self.audit_service._upload_artifacts(job_id, directory, artifact_names)
            checkpoint()
            self.audit_service._write_status(job_id, {
                **self.audit_service.status(job_id),
                "status": status,
                "progress": 100,
                "completed_at": datetime.now(UTC).isoformat(),
                "category_count": len(result.category_sheets),
                "unresolved_row_count": len(result.unresolved_rows),
                "review_count": len(reviews),
                "issue_count": len(result.issues),
                "revision_id": revision["revision_id"],
                "revision_number": revision["revision_number"],
                "output_sha256": output_sha256,
                "report_only": workbook.report_only,
                "artifacts": artifacts,
                "object_keys": object_keys,
            })
        except JobCancelled:
            previous = self.audit_service.status(job_id)
            self.audit_service._write_status(job_id, {
                **previous,
                "status": "cancelled",
                "completed_at": datetime.now(UTC).isoformat(),
            })
        except Exception as exc:
            _write_json_atomic(directory / "product-diagnostic.json", {
                "exception_type": type(exc).__name__,
                "stage": self.audit_service.status(job_id).get("status"),
            })
            self.audit_service._write_status(job_id, {
                **self.audit_service.status(job_id),
                "status": "failed",
                "completed_at": datetime.now(UTC).isoformat(),
                "error_code": (
                    "PRODUCT_PROCESSING_TIMEOUT"
                    if isinstance(exc, TimeoutError)
                    else "PRODUCT_NORMALIZATION_FAILED"
                ),
                "error_message_safe": "商品表格规范化失败，请检查输入、规则和平台目录配置。",
            })
        finally:
            if result is not None:
                result.close()
            if workbook is not None:
                workbook.close()

    def list_reviews(
        self,
        job_id: str,
        *,
        tenant_id: str = "local",
        status: str | None = None,
    ) -> list[dict[str, Any]]:
        if self.database:
            return self.database.list_product_reviews(job_id, status=status, tenant_id=tenant_id)
        state = self._state(job_id)
        reviews = state.get("reviews", [])
        return [item for item in reviews if status is None or item.get("status") == status]

    def decide_review(
        self,
        job_id: str,
        review_id: str,
        decision: ProductReviewDecision,
        *,
        tenant_id: str = "local",
        actor_id: str = "local",
    ) -> dict[str, Any]:
        state = self._state(job_id)
        review = next((item for item in state.get("reviews", []) if item.get("review_id") == review_id), None)
        if review is None:
            raise FileNotFoundError("product review item not found")
        _validate_review_decision(review, decision)
        if self.database:
            updated = self.database.decide_product_review(
                job_id,
                review_id,
                decision.model_dump(mode="json", exclude_none=True),
                actor_id=actor_id,
                tenant_id=tenant_id,
            )
        else:
            if review.get("status") != "pending":
                raise ValueError("product review item has already been decided")
            review["status"] = "resolved"
            review["decision"] = decision.model_dump(mode="json", exclude_none=True)
            review["decided_by"] = actor_id
            review["decided_at"] = datetime.now(UTC).isoformat()
            updated = review
        for index, item in enumerate(state.get("reviews", [])):
            if item.get("review_id") == review_id:
                state["reviews"][index] = updated
                break
        _write_json_atomic(self.audit_service.job_directory(job_id) / "product-workflow.json", state)
        if all(item.get("status") == "resolved" for item in state.get("reviews", [])):
            self.audit_service._write_status(job_id, {
                **self.audit_service.status(job_id),
                "status": "review_resolved",
                "pending_review_count": 0,
            })
        return updated

    def rerun_after_reviews(
        self,
        job_id: str,
        rules: RuleSet,
        *,
        tenant_id: str = "local",
        actor_id: str = "local",
    ) -> None:
        state = self._state(job_id)
        reviews = state.get("reviews", [])
        if any(review.get("status") != "resolved" for review in reviews):
            raise ValueError("all review items must be resolved before creating a revision")
        decisions = [review.get("decision") or {} for review in reviews]
        if any(decision.get("action") == "reject" for decision in decisions):
            raise ValueError("rejected review items must be corrected before creating a revision")
        category_overrides: dict[int, str] = {}
        confirmed_aliases_by_category: dict[str, dict[str, str]] = {}
        forced_extra_columns: set[int] = set()
        for review, decision in zip(reviews, decisions, strict=True):
            action = decision.get("action")
            payload = review.get("payload", {})
            if action == "confirm_category":
                category_overrides[int(payload["excel_row"])] = str(decision["category_id"])
            elif action == "confirm_mapping":
                category_id = str(payload["category_id"])
                confirmed_aliases_by_category.setdefault(category_id, {})[
                    str(decision["raw_header"])
                ] = str(decision["field_id"])
            elif action == "keep_extra" and payload.get("physical_column") is not None:
                forced_extra_columns.add(int(payload["physical_column"]))
        directory = self.audit_service.job_directory(job_id)
        inputs = [path for path in directory.glob("product-input.*") if path.is_file()]
        if len(inputs) != 1:
            raise FileNotFoundError("original product workbook is unavailable")
        current_revision = state.get("current_revision", {})
        revision_number = int(current_revision.get("revision_number", 0))
        result_path = directory / f"product-result-r{revision_number}.json"
        if revision_number < 1 or not result_path.is_file():
            raise FileNotFoundError("current product result snapshot is unavailable")
        frozen_result = ProductNormalizationResult.model_validate(
            json.loads(result_path.read_text(encoding="utf-8"))
        )
        frozen_catalog = FrozenCatalogAdapter(
            frozen_result.category_catalog_snapshot,
            frozen_result.catalog_snapshots,
        )
        parent = current_revision.get("revision_id")
        revision_lock = directory / ".revision.lock"
        try:
            revision_lock.mkdir()
        except FileExistsError as exc:
            try:
                stale_after = rules.workbook.processing_timeout_seconds + 60
                if time.time() - revision_lock.stat().st_mtime > stale_after:
                    revision_lock.rmdir()
                    revision_lock.mkdir()
                else:
                    raise ValueError("a product revision is already running") from exc
            except FileNotFoundError:
                revision_lock.mkdir()
        try:
            self.run(
                job_id,
                inputs[0],
                rules,
                frozen_catalog,
                tenant_id=tenant_id,
                actor_id=actor_id,
                confirmed_aliases_by_category=confirmed_aliases_by_category,
                category_overrides=category_overrides,
                forced_extra_columns=forced_extra_columns,
                parent_revision_id=parent,
            )
        finally:
            frozen_result.close()
            try:
                revision_lock.rmdir()
            except FileNotFoundError:
                pass

    def rerun_after_reviews_safe(
        self,
        job_id: str,
        rules: RuleSet,
        *,
        tenant_id: str = "local",
        actor_id: str = "local",
    ) -> None:
        try:
            self.rerun_after_reviews(
                job_id,
                rules,
                tenant_id=tenant_id,
                actor_id=actor_id,
            )
        except Exception as exc:
            directory = self.audit_service.job_directory(job_id)
            _write_json_atomic(directory / "product-diagnostic.json", {
                "exception_type": type(exc).__name__,
                "stage": "revision_queued",
            })
            self.audit_service._write_status(job_id, {
                **self.audit_service.status(job_id),
                "status": "failed",
                "completed_at": datetime.now(UTC).isoformat(),
                "error_code": "PRODUCT_REVISION_FAILED",
                "error_message_safe": "商品表格修订失败，请重新检查审核决定和原始文件。",
            })

    def _state(self, job_id: str) -> dict[str, Any]:
        path = self.audit_service.job_directory(job_id) / "product-workflow.json"
        if not path.is_file():
            raise FileNotFoundError("product workflow state not found")
        return json.loads(path.read_text(encoding="utf-8"))


def _validate_review_decision(review: dict[str, Any], decision: ProductReviewDecision) -> None:
    review_type = review.get("review_type")
    if review_type == "category" and decision.action not in {"confirm_category", "reject"}:
        raise ValueError("category review requires confirm_category or reject")
    if review_type == "field_mapping" and decision.action not in {"confirm_mapping", "keep_extra", "reject"}:
        raise ValueError("field mapping review requires confirm_mapping, keep_extra, or reject")
    if review_type == "duplicate_header" and decision.action not in {"keep_extra", "reject"}:
        raise ValueError("duplicate header review cannot be auto-mapped")
    if decision.action in {"confirm_category", "confirm_mapping"}:
        target = decision.category_id if decision.action == "confirm_category" else decision.field_id
        candidates = {
            candidate.get("field_id")
            for candidate in review.get("payload", {}).get("candidates", [])
        }
        if not candidates:
            raise ValueError("review decision has no frozen candidate target")
        if target not in candidates:
            raise ValueError("review decision target is not one of the recorded candidates")
    if decision.action == "confirm_mapping" and decision.raw_header != review.get("payload", {}).get("raw_header"):
        raise ValueError("confirmed raw_header does not match the reviewed physical header")


def _write_json_atomic(path: Path, payload: dict[str, Any]) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    temporary.replace(path)


def _write_jsonl_atomic(path: Path, rows: list[dict[str, Any]]) -> None:
    temporary = path.with_suffix(path.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8", newline="\n") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False, sort_keys=True, default=str))
            handle.write("\n")
    temporary.replace(path)


def _render_product_development(
    source: Path,
    destination: Path,
    result: Any,
    *,
    job_id: str,
    rules: RuleSet,
) -> str:
    """Local fallback mirroring the strict product-sheet operation used by .NET."""
    shutil.copy2(source, destination)
    book = load_workbook(destination, data_only=False, keep_links=False)
    issue_colors = {
        (issue.category_id, issue.excel_row, issue.field_id): issue.color
        for issue in result.issues
        if issue.category_id and issue.field_id
    }
    used_names = {name.casefold() for name in book.sheetnames}
    validation_cache: dict[tuple[str, ...], str] = {}
    validation_sheet_name = _unique_local_sheet_name("__ExcelAuditorLists", "lists", used_names)
    for category in result.category_sheets:
        name = category.worksheet_name
        if name in book.sheetnames:
            del book[name]
            used_names.discard(name.casefold())
        sheet = book.create_sheet(name)
        used_names.add(name.casefold())
        fields = [item.field for item in category.plan.fields]
        _write_product_sheet(
            book, sheet, fields, category.rows, category.source_excel_rows, category.category_id,
            issue_colors, result.merchant_extra_header_color, validation_cache, validation_sheet_name,
        )
        if category.sku_rows:
            sku_name = _unique_local_sheet_name(f"{name}-SKU", category.category_id, used_names)
            sku_sheet = book.create_sheet(sku_name)
            sku_fields = [
                field for field in fields
                if field.source.value in {"fixed", "platform_specification"}
            ]
            _write_product_sheet(
                book,
                sku_sheet,
                sku_fields,
                category.sku_rows,
                category.sku_source_excel_rows,
                category.category_id,
                issue_colors,
                result.merchant_extra_header_color,
                validation_cache,
                validation_sheet_name,
            )
            used_names.add(sku_name.casefold())
    if result.unresolved_rows:
        review_name = _unique_local_sheet_name("待审核商品", "review", used_names)
        review_sheet = book.create_sheet(review_name)
        review_headers = ["源行", "状态", "匹配类型", "原类目ID", "原类目", "候选类目", *result.source_headers]
        review_rows = []
        for item in result.unresolved_rows:
            resolution = item.category_resolution
            candidates = " | ".join(f"{candidate.field_id}:{candidate.title}" for candidate in resolution.candidates)
            review_rows.append([
                item.excel_row, resolution.status, resolution.match_type, resolution.raw_category_id,
                resolution.raw_category, candidates, *item.values,
            ])
        _write_product_review_sheet(review_sheet, review_headers, review_rows, "D9D2E9")
        used_names.add(review_name.casefold())
    if result.issues:
        issue_name = _unique_local_sheet_name("问题清单", "issues", used_names)
        issue_sheet = book.create_sheet(issue_name)
        _write_product_review_sheet(
            issue_sheet,
            ["源行", "类目ID", "字段ID", "问题类型", "原值", "说明"],
            [[
                item.excel_row, item.category_id, item.field_id, item.issue_type,
                item.raw_value, item.message,
            ] for item in result.issues],
            "F9CB9C",
        )
        used_names.add(issue_name.casefold())
    if "__ExcelAuditorMetadata" in book.sheetnames:
        del book["__ExcelAuditorMetadata"]
    metadata = book.create_sheet("__ExcelAuditorMetadata")
    metadata.append(["key", "value"])
    for key, value in [
        ("job_id", job_id),
        ("schema_id", rules.schema_id),
        ("schema_version", rules.schema_version),
        ("schema_sha256", rules.content_sha256),
        ("standard_snapshot_id", ""),
        ("standard_sha256", ""),
        ("input_sha256", sha256_file(source)),
        ("result_content_sha256", ""),
        ("operation_count", len(result.category_sheets)),
    ]:
        metadata.append([key, value])
    metadata.sheet_state = "veryHidden"
    book.save(destination)
    book.close()
    metadata_entry = _worksheet_package_entry(destination, "__ExcelAuditorMetadata")
    _set_result_content_hash(destination, metadata_entry)
    load_workbook(destination, read_only=True, data_only=False).close()
    return sha256_file(destination)


def _write_product_sheet(
    book: Any,
    sheet: Any,
    fields: list[Any],
    rows: list[dict[str, Any]],
    source_rows: list[int],
    category_id: str,
    issue_colors: dict[tuple[str, int, str], str],
    merchant_extra_header_color: str,
    validation_cache: dict[tuple[str, ...], str],
    validation_sheet_name: str,
) -> None:
    header_colors = {
        "fixed": "DDEBF7",
        "platform_attribute": "E2F0D9",
        "platform_specification": "FFF2CC",
        "merchant_extra": merchant_extra_header_color,
    }
    for column, field in enumerate(fields, start=1):
        cell = sheet.cell(1, column, field.title)
        cell.fill = PatternFill("solid", fgColor=header_colors[field.source.value])
    for output_row, (source_row, values) in enumerate(zip(source_rows, rows, strict=True), start=2):
        for column, field in enumerate(fields, start=1):
            value = values.get(field.field_id)
            if isinstance(value, datetime):
                target = ZoneInfo(field.timezone) if field.timezone else UTC
                value = value.astimezone(target).replace(tzinfo=None) if value.tzinfo else value
            cell = sheet.cell(output_row, column, value)
            if isinstance(value, str):
                cell.data_type = "s"
            if field.number_format:
                cell.number_format = field.number_format
            color = issue_colors.get((category_id, source_row, field.field_id))
            if color:
                cell.fill = PatternFill("solid", fgColor=color)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(fields)).coordinate[:-1]}{max(1, len(rows) + 1)}"
    for column, field in enumerate(fields, start=1):
        validation = None
        allow_blank = not field.required and field.validation.nullable
        if field.enum_values:
            range_name = _ensure_openpyxl_validation_list(
                book,
                tuple(field.enum_values),
                validation_cache,
                validation_sheet_name,
            )
            validation = DataValidation(
                type="list",
                formula1=range_name,
                allow_blank=allow_blank,
                errorTitle="Invalid value",
                error="Select a value from the platform catalog list.",
                showErrorMessage=True,
            )
        elif field.field_type.value in {"integer", "decimal"} and (
            field.validation.min is not None or field.validation.max is not None
        ):
            minimum, maximum = field.validation.min, field.validation.max
            operator = "between" if minimum is not None and maximum is not None else (
                "greaterThanOrEqual" if minimum is not None else "lessThanOrEqual"
            )
            validation = DataValidation(
                type="whole" if field.field_type.value == "integer" else "decimal",
                operator=operator,
                formula1=str(minimum if minimum is not None else maximum),
                formula2=str(maximum) if minimum is not None and maximum is not None else None,
                allow_blank=allow_blank,
                errorTitle="Invalid number",
                error="Enter a value within the platform catalog bounds.",
                showErrorMessage=True,
            )
        if validation is not None:
            sheet.add_data_validation(validation)
            column_name = get_column_letter(column)
            validation.add(f"{column_name}2:{column_name}1048576")


def _write_product_review_sheet(sheet: Any, headers: list[str], rows: list[list[Any]], color: str) -> None:
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(1, column, header)
        cell.fill = PatternFill("solid", fgColor=color)
    for row_index, values in enumerate(rows, start=2):
        for column, value in enumerate(values, start=1):
            if isinstance(value, (dict, list, tuple, set)):
                value = json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
            cell = sheet.cell(row_index, column, "" if value is None else value)
            if isinstance(cell.value, str):
                cell.data_type = "s"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(rows) + 1)}"


def _ensure_openpyxl_validation_list(
    book: Any,
    values: tuple[str, ...],
    cache: dict[tuple[str, ...], str],
    sheet_name: str,
) -> str:
    if values in cache:
        return cache[values]
    if sheet_name in book.sheetnames:
        lookup = book[sheet_name]
    else:
        lookup = book.create_sheet(sheet_name)
        lookup.sheet_state = "veryHidden"
    column = len(cache) + 1
    for row, value in enumerate(values, start=1):
        lookup.cell(row, column, value)
    ordinal = 1
    existing = set(book.defined_names)
    while f"_ExcelAuditorList{ordinal}" in existing:
        ordinal += 1
    name = f"_ExcelAuditorList{ordinal}"
    column_name = get_column_letter(column)
    quoted_sheet = sheet_name.replace("'", "''")
    book.defined_names.add(DefinedName(
        name,
        attr_text=f"'{quoted_sheet}'!${column_name}$1:${column_name}${len(values)}",
    ))
    cache[values] = name
    return name


def _unique_local_sheet_name(base: str, category_id: str, used: set[str]) -> str:
    cleaned = "".join("-" if character in "[]:*?/\\" else character for character in base)[:31]
    if cleaned.casefold() not in used:
        return cleaned
    suffix = f"-{category_id}"[:12]
    candidate = f"{cleaned[:31-len(suffix)]}{suffix}"
    counter = 2
    while candidate.casefold() in used:
        marker = f"-{counter}"
        candidate = f"{cleaned[:31-len(suffix)-len(marker)]}{suffix}{marker}"
        counter += 1
    return candidate
