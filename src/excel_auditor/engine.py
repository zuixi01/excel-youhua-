from __future__ import annotations

import json
import math
import re
import os
from collections import Counter, defaultdict
from dataclasses import dataclass
from decimal import Decimal
from typing import Any, Sequence

from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz, process

from .models import (
    Difference,
    DifferenceType,
    HeaderMapping,
    ReportSummary,
    RuleSet,
    SheetRule,
    normalize_header,
)
from .normalization import ParsedValue, is_formula_text, parse_excel_value, parse_row_number, parse_value, values_equal
from .record_store import DiskBackedRecordMap
from .snapshots import SpilledRecords
from .spill import SpillableSequence
from .workbook import SheetSnapshot, WorkbookSnapshot, locate_header_row
from .validators import run_validator
from .ids import new_ulid


_EXCEL_MIN_POSITIVE_NUMBER = Decimal("2.2251E-308")
_EXCEL_MAX_ABSOLUTE_NUMBER = Decimal("9.99999999999999E307")


@dataclass
class ComparisonResult:
    mappings: list[HeaderMapping]
    differences: Sequence[Difference]
    summary: ReportSummary
    repairs: Sequence["RepairOperation"]
    manual_review_reasons: list[str] | None = None
    join_backends: list[str] | None = None
    report_only: bool = False
    storage_backends: list[str] | None = None
    matched_records_by_sheet: dict[str, int] | None = None
    validated_records_by_sheet: dict[str, int] | None = None
    resolved_header_rows_by_sheet: dict[str, int] | None = None
    data_start_rows_by_sheet: dict[str, int] | None = None
    formula_rows_by_sheet: dict[str, list[int]] | None = None

    def close(self) -> None:
        for sequence in (self.differences, self.repairs):
            close = getattr(sequence, "close", None)
            if close is not None:
                close()


@dataclass(frozen=True)
class RepairOperation:
    type: str
    sheet_id: str
    sheet_name: str
    rule_id: str
    difference_id: str
    cell: str | None = None
    excel_row: int | None = None
    canonical_field: str | None = None
    value: Any = None
    values: dict[str, Any] | None = None


def _difference(kind: DifferenceType, sheet: SheetRule, message: str, **kwargs: Any) -> Difference:
    sheet_name = kwargs.pop("sheet_name", sheet.name)
    return Difference(difference_id=new_ulid("diff_"), type=kind, sheet_id=sheet.id, sheet_name=sheet_name, message=message, **kwargs)


def map_headers(sheet: SheetRule, snapshot: SheetSnapshot) -> tuple[list[HeaderMapping], dict[str, int]]:
    header_values = snapshot.rows[sheet.header.row - 1][1] if len(snapshot.rows) >= sheet.header.row else []
    normalized_values = [normalize_header(value) for value in header_values]
    duplicate_values = {value for value, count in Counter(value for value in normalized_values if value).items() if count > 1}
    exact: dict[str, tuple[str, str]] = {}
    for column in sheet.columns:
        exact.setdefault(normalize_header(column.name), (column.name, "canonical"))
        exact.setdefault(normalize_header(column.title), (column.name, "exact_title"))
        for alias in column.aliases:
            exact.setdefault(normalize_header(alias), (column.name, "confirmed_alias"))
    choices = {column.name: normalize_header(column.title) for column in sheet.columns}
    mappings: list[HeaderMapping] = []
    for index, (raw, normalized) in enumerate(zip(header_values, normalized_values), start=1):
        raw_text = "" if raw is None else str(raw)
        if normalized in duplicate_values:
            mapping = HeaderMapping(sheet_id=sheet.id, physical_column=index, raw_header=raw_text, normalized_header=normalized, match_type="ambiguous", status="duplicate")
        elif normalized in exact:
            canonical, match_type = exact[normalized]
            mapping = HeaderMapping(sheet_id=sheet.id, physical_column=index, raw_header=raw_text, normalized_header=normalized, canonical_field=canonical, match_type=match_type, confidence=100, status="matched")
        elif normalized:
            match = process.extractOne(normalized, choices, scorer=fuzz.ratio, score_cutoff=sheet.header.fuzzy_suggestion_threshold)
            if match:
                _title, score, canonical = match
                mapping = HeaderMapping(sheet_id=sheet.id, physical_column=index, raw_header=raw_text, normalized_header=normalized, canonical_field=canonical, match_type="fuzzy_suggestion", confidence=score, status="ambiguous")
            else:
                mapping = HeaderMapping(sheet_id=sheet.id, physical_column=index, raw_header=raw_text, normalized_header=normalized, match_type="unmatched", status="extra")
        else:
            mapping = HeaderMapping(sheet_id=sheet.id, physical_column=index, raw_header=raw_text, normalized_header=normalized, match_type="unmatched", status="extra")
        mappings.append(mapping)
    matched_by_canonical: dict[str, list[int]] = defaultdict(list)
    for mapping_index, mapping in enumerate(mappings):
        if mapping.status == "matched" and mapping.canonical_field:
            matched_by_canonical[mapping.canonical_field].append(mapping_index)
    semantic_duplicates = {
        canonical: indexes
        for canonical, indexes in matched_by_canonical.items()
        if len(indexes) > 1
    }
    for canonical, indexes in semantic_duplicates.items():
        for mapping_index in indexes:
            mappings[mapping_index] = mappings[mapping_index].model_copy(
                update={"status": "duplicate", "canonical_field": canonical}
            )
    canonical_columns = {
        mapping.canonical_field: mapping.physical_column
        for mapping in mappings
        if mapping.status == "matched" and mapping.canonical_field
    }
    return mappings, canonical_columns


def compare_workbook(
    workbook: WorkbookSnapshot,
    standard: dict[str, Sequence[dict[str, Any]]],
    rules: RuleSet,
    *,
    job_id: str | None = None,
    difference_spill_threshold: int | None = None,
    spill_to_disk: bool | None = None,
) -> ComparisonResult:
    threshold = difference_spill_threshold if difference_spill_threshold is not None else int(os.environ.get("EXCEL_AUDITOR_DIFFERENCE_SPILL_THRESHOLD", "50000"))
    if threshold < 1:
        raise ValueError("EXCEL_AUDITOR_DIFFERENCE_SPILL_THRESHOLD must be positive")

    def attach_job_id(item: Difference) -> None:
        if job_id is not None:
            item.job_id = job_id

    spill_enabled = spill_to_disk if spill_to_disk is not None else (
        workbook.report_only or any(isinstance(rows, SpilledRecords) for rows in standard.values())
    )
    differences: list[Difference] | SpillableSequence[Difference]
    repairs: list[RepairOperation] | SpillableSequence[RepairOperation]
    differences = SpillableSequence[Difference](threshold, on_append=attach_job_id) if spill_enabled else []
    repairs = SpillableSequence[RepairOperation](threshold) if spill_enabled else []
    all_mappings: list[HeaderMapping] = []
    summary = ReportSummary()
    manual_review_reasons: list[str] = []
    join_backends: list[str] = []
    storage_backends: list[str] = []
    matched_records_by_sheet: dict[str, int] = {}
    validated_records_by_sheet: dict[str, int] = {}
    resolved_header_rows_by_sheet: dict[str, int] = {}
    data_start_rows_by_sheet: dict[str, int] = {}
    formula_rows_by_sheet: dict[str, list[int]] = {}
    consumed_sheets: set[str] = set()
    for sheet_rule in rules.sheets:
        matching_names = [name for name in [sheet_rule.name, *sheet_rule.aliases] if name in workbook.sheets]
        if len(matching_names) > 1:
            consumed_sheets.update(matching_names)
            differences.append(_difference(
                DifferenceType.AMBIGUOUS_SHEET,
                sheet_rule,
                f"工作表规则同时匹配到多个物理工作表：{matching_names}",
                severity="error",
                render_action="report_only",
            ))
            manual_review_reasons.append(
                f"{sheet_rule.id}: ambiguous_sheet:{'|'.join(matching_names)}"
            )
            continue
        actual_name = matching_names[0] if matching_names else None
        if actual_name is None:
            if sheet_rule.required:
                differences.append(_difference(DifferenceType.MISSING_SHEET, sheet_rule, f"缺少必需工作表：{sheet_rule.name}"))
            continue
        consumed_sheets.add(actual_name)
        snapshot = workbook.sheets[actual_name]
        header_row, header_problem = locate_header_row(sheet_rule, snapshot)
        if header_problem:
            differences.append(_difference(DifferenceType.HEADER_NOT_FOUND, sheet_rule, header_problem, sheet_name=actual_name, severity="error"))
            manual_review_reasons.append(f"{actual_name}: header_not_found_or_ambiguous")
            continue
        if sheet_rule.data_region.start_row is not None and sheet_rule.data_region.start_row <= header_row:
            differences.append(_difference(
                DifferenceType.HEADER_NOT_FOUND,
                sheet_rule,
                f"数据起始行 {sheet_rule.data_region.start_row} 必须晚于定位表头行 {header_row}",
                sheet_name=actual_name,
                severity="error",
            ))
            manual_review_reasons.append(f"{actual_name}: data_region_overlaps_header")
            continue
        # Keep the stable rule id, but carry the physical worksheet name through
        # every difference and render operation when a configured alias matched.
        actual_sheet_rule = sheet_rule.model_copy(update={"name": actual_name, "header": sheet_rule.header.model_copy(update={"row": header_row})})
        resolved_header_rows_by_sheet[sheet_rule.id] = header_row
        data_start_rows_by_sheet[sheet_rule.id] = sheet_rule.data_region.start_row or header_row + 1
        mappings, canonical_columns = map_headers(actual_sheet_rule, snapshot)
        all_mappings.extend(mappings)
        differences.extend(_header_differences(actual_sheet_rule, mappings, canonical_columns, repairs))
        if any(mapping.status == "duplicate" for mapping in mappings) or (
            sheet_rule.primary_key_mode == "fields" and any(key not in canonical_columns for key in sheet_rule.primary_key)
        ):
            continue
        sheet_standard = standard.get(sheet_rule.id, standard.get(sheet_rule.name, []))
        join_backend, record_storage, matched_count, validated_count, formula_rows = _compare_records(
            actual_sheet_rule,
            snapshot,
            canonical_columns,
            sheet_standard,
            differences,
            repairs,
            summary,
            workbook.excel_epoch,
        )
        join_backends.append(join_backend)
        storage_backends.append(record_storage)
        matched_records_by_sheet[sheet_rule.id] = matched_count
        validated_records_by_sheet[sheet_rule.id] = validated_count
        formula_rows_by_sheet[sheet_rule.id] = formula_rows
    for extra_name in set(workbook.sheets) - consumed_sheets:
        pseudo = rules.sheets[0]
        differences.append(_difference(DifferenceType.EXTRA_SHEET, pseudo, f"存在规则未声明的工作表：{extra_name}", sheet_name=extra_name, severity="warning"))
    summary.differences = len(differences)
    summary.mismatched_cells = sum(item.type == DifferenceType.VALUE_MISMATCH for item in differences)
    summary.validation_errors = sum(item.type in {DifferenceType.INVALID_VALUE, DifferenceType.VALIDATION_ERROR} for item in differences)
    summary.repairs_planned = sum(item.repair_status == "planned" for item in differences)
    manual_review_reasons.extend(sorted({
        f"{item.sheet_name}: fuzzy_value_suggestion:{item.canonical_field}"
        for item in differences
        if item.rule_id and item.rule_id.endswith(".fuzzy_suggestion")
    }))
    manual_review_reasons.extend(sorted({
        f"{item.sheet_name}: formula_primary_key:{item.canonical_field}"
        for item in differences
        if item.rule_id and item.rule_id.endswith(".formula_primary_key")
    }))
    manual_review_reasons.extend(sorted({
        f"{item.sheet_name}: formula_append_requires_trusted_template"
        for item in differences
        if item.rule_id == "missing_record.formula_template_required"
    }))
    manual_review_reasons.extend(sorted({
        f"{item.sheet_name}: formula_template_mismatch:{item.canonical_field}"
        for item in differences
        if item.rule_id and item.rule_id.endswith(".formula_template_mismatch")
    }))
    manual_review_reasons.extend(sorted({
        f"{item.sheet_name}: excel_numeric_write_precision:{item.canonical_field}"
        for item in differences
        if item.rule_id and item.rule_id.endswith(".excel_write_precision")
    }))
    spilled = (
        isinstance(differences, SpillableSequence) and differences.spilled
    ) or (
        isinstance(repairs, SpillableSequence) and repairs.spilled
    )
    storage_backends.append("disk_differences" if spilled else "memory_differences")
    return ComparisonResult(
        all_mappings,
        differences.finish() if isinstance(differences, SpillableSequence) else differences,
        summary,
        repairs.finish() if isinstance(repairs, SpillableSequence) else repairs,
        manual_review_reasons,
        join_backends,
        report_only=spilled,
        storage_backends=sorted(set(storage_backends)),
        matched_records_by_sheet=matched_records_by_sheet,
        validated_records_by_sheet=validated_records_by_sheet,
        resolved_header_rows_by_sheet=resolved_header_rows_by_sheet,
        data_start_rows_by_sheet=data_start_rows_by_sheet,
        formula_rows_by_sheet=formula_rows_by_sheet,
    )


def _header_differences(sheet: SheetRule, mappings: list[HeaderMapping], canonical_columns: dict[str, int], repairs: list[RepairOperation]) -> list[Difference]:
    result: list[Difference] = []
    by_name = {column.name: column for column in sheet.columns}
    ordered_positions = [canonical_columns[column.name] for column in sheet.columns if column.name in canonical_columns]
    if ordered_positions != sorted(ordered_positions):
        result.append(_difference(DifferenceType.HEADER_ORDER_MISMATCH, sheet, "标准字段的物理顺序与规则不一致", severity="warning", render_action="report_only"))
    for mapping in mappings:
        cell = f"{get_column_letter(mapping.physical_column)}{sheet.header.row}"
        if mapping.status == "duplicate":
            result.append(_difference(DifferenceType.DUPLICATE_HEADER, sheet, f"重复表头：{mapping.raw_header}", cell=cell, canonical_field=mapping.canonical_field, excel_raw_value=mapping.raw_header, render_action=sheet.actions.duplicate_key))
        elif mapping.status == "ambiguous":
            result.append(_difference(DifferenceType.AMBIGUOUS_HEADER, sheet, f"表头仅产生候选，未自动映射：{mapping.raw_header}", cell=cell, canonical_field=mapping.canonical_field, excel_raw_value=mapping.raw_header, severity="warning", render_action="mark_purple"))
        elif mapping.status == "extra":
            result.append(_difference(DifferenceType.EXTRA_HEADER, sheet, f"多余表头：{mapping.raw_header}", cell=cell, excel_raw_value=mapping.raw_header, render_action=sheet.actions.extra_header))
        elif mapping.match_type == "confirmed_alias" and sheet.actions.rename_confirmed_alias and mapping.canonical_field:
            column = by_name[mapping.canonical_field]
            difference = _difference(
                DifferenceType.NORMALIZED_MATCH,
                sheet,
                f"已授权将确认别名改为标准表头：{column.title}",
                cell=cell,
                canonical_field=column.name,
                excel_raw_value=mapping.raw_header,
                standard_raw_value=column.title,
                rule_id=f"{column.name}.rename_confirmed_alias",
                severity="info",
                render_action="set_cell_green",
                repair_status="planned",
            )
            result.append(difference)
            repairs.append(RepairOperation("set_cell", sheet.id, sheet.name, difference.rule_id or "", difference.difference_id, cell=cell, canonical_field=column.name, value=column.title))
    for name, column in by_name.items():
        if name not in canonical_columns and not any(mapping.canonical_field == name and mapping.status == "ambiguous" for mapping in mappings):
            action = "insert_and_mark_green" if _should_insert_missing(sheet, column) else "report_only"
            result.append(_difference(
                DifferenceType.MISSING_HEADER,
                sheet,
                f"缺少表头：{column.title}",
                canonical_field=name,
                standard_raw_value=column.title,
                rule_id=f"{name}.missing_column",
                render_action=action,
                repair_status="planned" if action == "insert_and_mark_green" else "not_requested",
            ))
    return result


def _compare_records(sheet: SheetRule, snapshot: SheetSnapshot, columns: dict[str, int], standard_rows: Sequence[dict[str, Any]], differences: list[Difference], repairs: list[RepairOperation], summary: ReportSummary, excel_epoch: Any) -> tuple[str, str, int, int, list[int]]:
    rules_by_name = {column.name: column for column in sheet.columns}
    start_row = sheet.data_region.start_row or sheet.header.row + 1
    excel_records: dict[tuple[Any, ...], tuple[int, dict[str, Any]]] = {}
    excel_duplicates: dict[tuple[Any, ...], list[int]] = defaultdict(list)
    formula_target_rows: list[int] = []
    for row_number, values in snapshot.rows[start_row - 1 :]:
        if row_number in snapshot.hidden_rows and not sheet.data_region.include_hidden_rows:
            continue
        record = {name: values[index - 1] for name, index in columns.items() if index <= len(values) and values[index - 1] not in {None, ""}}
        if all(value is None or value == "" for value in record.values()):
            continue
        formula_target_rows.append(row_number)
        formula_key_fields = [
            name
            for name in sheet.primary_key
            if isinstance(record.get(name), str) and record[name].startswith("=")
        ]
        if formula_key_fields:
            for name in formula_key_fields:
                col = columns.get(name)
                differences.append(_difference(
                    DifferenceType.UNSUPPORTED_FEATURE,
                    sheet,
                    "Primary-key formulas are not evaluated and cannot participate in automatic record matching",
                    cell=f"{get_column_letter(col)}{row_number}" if col else None,
                    excel_row=row_number,
                    canonical_field=name,
                    excel_raw_value=_safe_value(record.get(name), rules_by_name[name]),
                    rule_id=f"{name}.formula_primary_key",
                    render_action="mark_purple",
                ))
            continue
        key, key_valid = _key(record, sheet, rules_by_name, row_number=row_number, excel_epoch=excel_epoch)
        if not key_valid:
            if sheet.empty_primary_key_action == "skip_row":
                continue
            if sheet.empty_primary_key_action == "use_row_number":
                key, key_valid = (("row_number_fallback", row_number - start_row + 1),), True
        if not key_valid:
            key_action = "mark_purple" if sheet.actions.duplicate_key == "mark_purple" else "report_only"
            key_fields = sheet.primary_key or [sheet.row_number_field]
            for name in key_fields:
                col = columns.get(name)
                differences.append(_difference(DifferenceType.EMPTY_PRIMARY_KEY, sheet, "主键为空或无法解析", cell=f"{get_column_letter(col)}{row_number}" if col else None, excel_row=row_number, canonical_field=name, excel_raw_value=_safe_value(record.get(name), rules_by_name[name]), render_action=key_action))
            continue
        if key in excel_records:
            excel_duplicates[key].extend([excel_records[key][0], row_number])
        else:
            excel_records[key] = (row_number, record)
    for key, rows in excel_duplicates.items():
        excel_records.pop(key, None)
        for row_number in sorted(set(rows)):
            duplicate_action = "mark_row_purple" if sheet.actions.duplicate_key == "mark_purple" else "report_only"
            differences.append(_difference(DifferenceType.DUPLICATE_PRIMARY_KEY, sheet, "Excel 中主键重复，不参与自动匹配", excel_row=row_number, business_key=_business_key(key, sheet), render_action=duplicate_action))
    _validate_excel_records(sheet, snapshot, columns, excel_records, rules_by_name, differences, excel_epoch)
    join_threshold = int(os.environ.get("EXCEL_AUDITOR_POLARS_JOIN_THRESHOLD", "50000"))
    if join_threshold < 1:
        raise ValueError("EXCEL_AUDITOR_POLARS_JOIN_THRESHOLD must be positive")
    # Do not duplicate an already-resident caller list into SQLite: that only
    # adds I/O without releasing the source payload. Service large-data paths
    # arrive as SpilledRecords, where the disk map replaces the Python payload
    # dictionary and materially bounds memory.
    disk_record_threshold = int(os.environ.get("EXCEL_AUDITOR_DISK_RECORD_THRESHOLD", "250000"))
    if disk_record_threshold < 1:
        raise ValueError("EXCEL_AUDITOR_DISK_RECORD_THRESHOLD must be positive")
    use_disk_records = isinstance(standard_rows, SpilledRecords) or len(standard_rows) >= disk_record_threshold
    standard_records: dict[tuple[Any, ...], dict[str, Any]] | DiskBackedRecordMap
    standard_records = DiskBackedRecordMap() if use_disk_records else {}
    standard_duplicates: set[tuple[Any, ...]] = set()
    for standard_ordinal, record in enumerate(standard_rows, start=1):
        standard_row_number = record.get(sheet.row_number_field) if sheet.primary_key_mode == "row_number" else standard_ordinal
        key, valid = _key(record, sheet, rules_by_name, row_number=standard_row_number)
        if not valid and sheet.empty_primary_key_action == "skip_row":
            continue
        if not valid and sheet.empty_primary_key_action == "use_row_number":
            key, valid = (("row_number_fallback", standard_ordinal),), True
        if not valid or key in standard_records:
            if key:
                standard_duplicates.add(key)
            continue
        standard_records[key] = record
    for key in standard_duplicates:
        standard_records.pop(key, None)
        differences.append(_difference(DifferenceType.DUPLICATE_PRIMARY_KEY, sheet, "标准数据中主键重复，不参与自动匹配", business_key=_business_key(key, sheet), render_action="report_only"))
    if len(excel_records) + len(standard_records) >= join_threshold:
        from .partitioned_join import PolarsPartitionedKeyConnector

        excel_key_order = list(excel_records)
        standard_key_order = standard_records.iter_join_keys() if isinstance(standard_records, DiskBackedRecordMap) else list(standard_records)
        joined = PolarsPartitionedKeyConnector().classify(excel_key_order, standard_key_order)
        excel_only = [excel_key_order[index] for index in joined.excel_only]
        standard_only = joined.standard_only if isinstance(standard_records, DiskBackedRecordMap) else [standard_key_order[index] for index in joined.standard_only]
        matched_keys = [excel_key_order[excel_index] for excel_index, _standard_index in joined.matched]
        join_backend = "polars_partitioned"
    else:
        excel_keys, standard_keys = set(excel_records), set(standard_records)
        excel_only = sorted(excel_keys - standard_keys, key=str)
        standard_only = sorted(standard_keys - excel_keys, key=str)
        matched_keys = sorted(excel_keys & standard_keys, key=str)
        join_backend = "python_in_memory"
    if os.environ.get("EXCEL_AUDITOR_VERIFY_DATACOMPY") == "1":
        from .dataset_adapter import DataComPyAdapter

        excel_keys, standard_keys = set(excel_records), set(standard_records)
        observed = DataComPyAdapter().all_rows_overlap(excel_keys, standard_keys)
        if observed is not (excel_keys == standard_keys):
            raise RuntimeError("DataComPy join cross-check disagrees with the deterministic key connector")
    for key in excel_only:
        row_number, record = excel_records[key]
        summary.extra_records += 1
        extra_action = "mark_row_red" if sheet.actions.extra_record == "mark_red" else "report_only"
        differences.append(_difference(DifferenceType.EXTRA_RECORD, sheet, "仅 Excel 存在的记录", excel_row=row_number, business_key=_business_key(key, sheet), render_action=extra_action))
        for name, col_index in columns.items():
            rule = rules_by_name[name]
            parsed = parse_excel_value(record.get(name), rule, excel_epoch)
            if parsed.valid and parsed.normalized is None and rule.fill_static_default:
                cell = f"{get_column_letter(col_index)}{row_number}"
                default = parse_value(rule.static_default, rule)
                if not _excel_numeric_write_safe(default, rule):
                    _append_excel_write_precision_difference(differences, sheet, rule, default, row_number=row_number, cell=cell, key=key)
                    continue
                difference = _difference(
                    DifferenceType.NORMALIZED_MATCH,
                    sheet,
                    "已授权使用静态默认值填充空单元格",
                    cell=cell,
                    excel_row=row_number,
                    canonical_field=name,
                    business_key=_business_key(key, sheet),
                    excel_raw_value=_safe_value(parsed.raw, rule),
                    standard_raw_value=_safe_value(rule.static_default, rule),
                    rule_id=f"{name}.fill_static_default",
                    severity="info",
                    render_action="set_cell_green",
                    repair_status="planned",
                )
                differences.append(difference)
                default_value = _excel_write_value(default, rule)
                repairs.append(RepairOperation("set_cell", sheet.id, sheet.name, difference.rule_id or "", difference.difference_id, cell=cell, canonical_field=name, value=default_value))
    append_row = snapshot.max_row + 1
    for standard_reference in standard_only:
        summary.missing_records += 1
        if isinstance(standard_records, DiskBackedRecordMap):
            if join_backend == "polars_partitioned":
                key, record = standard_records.item_at_join_index(int(standard_reference))
            else:
                key = standard_reference
                record = standard_records[key]
        else:
            key = standard_reference
            record = standard_records[key]
        duplicate_in_excel = key in excel_duplicates
        requested_row = key[0][1] if sheet.primary_key_mode == "row_number" else append_row
        unsafe_formula_fields = [
            name
            for name, rule in rules_by_name.items()
            if rule.compare.formula_mode == "formula"
            and is_formula_text(record.get(name))
            and (
                rule.formula_template is None
                or rule.formula_template.replace("{row}", str(append_row)) != record.get(name)
            )
        ]
        formula_append_blocked = bool(unsafe_formula_fields) and sheet.actions.missing_record == "append_and_mark_green"
        parsed_record = {name: parse_value(record.get(name), rule) for name, rule in rules_by_name.items()}
        unsafe_numeric_fields = [
            name for name, rule in rules_by_name.items()
            if not _excel_numeric_write_safe(parsed_record[name], rule)
        ]
        numeric_append_blocked = bool(unsafe_numeric_fields) and sheet.actions.missing_record == "append_and_mark_green"
        if numeric_append_blocked:
            for name in unsafe_numeric_fields:
                _append_excel_write_precision_difference(differences, sheet, rules_by_name[name], parsed_record[name], key=key)
        append = (
            sheet.actions.missing_record == "append_and_mark_green"
            and not duplicate_in_excel
            and requested_row == append_row
            and not formula_append_blocked
            and not numeric_append_blocked
        )
        row_number_mismatch = sheet.primary_key_mode == "row_number" and requested_row != append_row
        if row_number_mismatch:
            missing_message = "标准数据行号与可追加物理行不一致，禁止自动追加"
        elif duplicate_in_excel:
            missing_message = "标准数据存在但 Excel 主键重复，禁止自动追加"
        elif formula_append_blocked:
            missing_message = f"标准记录包含无法由受信任模板重建的公式字段 {unsafe_formula_fields}，禁止自动追加"
        elif numeric_append_blocked:
            missing_message = f"标准记录包含超出 Excel 安全数值写入精度的字段 {unsafe_numeric_fields}，禁止自动追加"
        else:
            missing_message = "标准数据存在但 Excel 缺失的记录"
        if append:
            missing_rule_id = "missing_record.append"
        elif formula_append_blocked:
            missing_rule_id = "missing_record.formula_template_required"
        elif numeric_append_blocked:
            missing_rule_id = "missing_record.numeric_write_blocked"
        else:
            missing_rule_id = None
        difference = _difference(
            DifferenceType.MISSING_RECORD,
            sheet,
            missing_message,
            excel_row=append_row if append else None,
            business_key=_business_key(key, sheet),
            standard_raw_value=_safe_record(record, rules_by_name),
            rule_id=missing_rule_id,
            render_action="append_row_green" if append else "report_only",
            repair_status="planned" if append else "not_requested",
        )
        differences.append(difference)
        if append:
            write_record = {
                name: _excel_write_value(parsed_record[name], rule)
                for name, rule in rules_by_name.items()
            }
            repairs.append(RepairOperation("append_record", sheet.id, sheet.name, difference.rule_id or "", difference.difference_id, excel_row=append_row, values=write_record))
            append_row += 1
    for key in matched_keys:
        summary.matched_records += 1
        row_number, excel_record = excel_records[key]
        standard_record = standard_records[key]
        for name, rule in rules_by_name.items():
            col_index = columns.get(name)
            if col_index is None:
                if not _should_insert_missing(sheet, rule):
                    continue
                if rule.formula_template is not None:
                    if name in standard_record:
                        expected_formula = rule.formula_template.replace("{row}", str(row_number))
                        if standard_record.get(name) != expected_formula:
                            differences.append(_difference(
                                DifferenceType.VALUE_MISMATCH,
                                sheet,
                                "标准公式与受信任公式模板在目标行展开后的文本不一致",
                                excel_row=row_number,
                                canonical_field=name,
                                business_key=_business_key(key, sheet),
                                standard_raw_value=_safe_value(standard_record.get(name), rule),
                                standard_normalized_value=_safe_value(expected_formula, rule),
                                rule_id=f"{name}.formula_template_mismatch",
                                render_action="report_only",
                            ))
                    continue
                right = parse_value(standard_record.get(name), rule)
                repair_value = None
                write_value = None
                repair_rule = None
                if sheet.actions.fill_empty_from_standard and right.valid and right.normalized is not None:
                    if not _excel_numeric_write_safe(right, rule):
                        _append_excel_write_precision_difference(differences, sheet, rule, right, row_number=row_number, key=key)
                    else:
                        repair_value, write_value, repair_rule = right.raw, _excel_write_value(right, rule), f"{name}.fill_empty_from_standard"
                elif rule.fill_static_default:
                    default = parse_value(rule.static_default, rule)
                    if not _excel_numeric_write_safe(default, rule):
                        _append_excel_write_precision_difference(differences, sheet, rule, default, row_number=row_number, key=key)
                    else:
                        repair_value, write_value, repair_rule = rule.static_default, _excel_write_value(default, rule), f"{name}.fill_static_default"
                if repair_rule is not None:
                    difference = _difference(
                        DifferenceType.NORMALIZED_MATCH,
                        sheet,
                        "已授权在插入列中填充值",
                        excel_row=row_number,
                        canonical_field=name,
                        business_key=_business_key(key, sheet),
                        standard_raw_value=_safe_value(repair_value, rule),
                        rule_id=repair_rule,
                        severity="info",
                        render_action="set_field_green",
                        repair_status="planned",
                    )
                    differences.append(difference)
                    repairs.append(RepairOperation("set_field", sheet.id, sheet.name, repair_rule, difference.difference_id, excel_row=row_number, canonical_field=name, value=write_value))
                continue
            rule = rules_by_name[name]
            left_raw = excel_record.get(name)
            right_raw = standard_record.get(name)
            cell = f"{get_column_letter(col_index)}{row_number}"
            # An omitted optional standard field means the source supplied no
            # authoritative value. Preserve the distinction from an explicit
            # null/empty value so overwrite_mismatch cannot silently clear data.
            if name not in standard_record and not (
                (left_raw is None or left_raw == "") and rule.fill_static_default
            ):
                continue
            left_is_formula = is_formula_text(left_raw)
            right_is_formula = is_formula_text(right_raw)
            if rule.compare.formula_mode == "formula" and (left_is_formula or right_is_formula):
                if not (left_is_formula and right_is_formula and left_raw == right_raw):
                    differences.append(_difference(
                        DifferenceType.VALUE_MISMATCH,
                        sheet,
                        "公式存在性或公式文本与标准不一致；系统未执行公式",
                        cell=cell,
                        excel_row=row_number,
                        canonical_field=name,
                        business_key=_business_key(key, sheet),
                        excel_raw_value=_safe_value(left_raw, rule),
                        standard_raw_value=_safe_value(right_raw, rule),
                        rule_id=f"{name}.formula_text",
                        render_action=sheet.actions.mismatched_value,
                    ))
                continue
            if left_is_formula:
                if rule.compare.formula_mode == "reject":
                    continue
                left_raw = snapshot.cached_values.get(cell)
            if (left_raw is None or left_raw == "") and (right_raw is None or right_raw == "") and not rule.fill_static_default:
                continue
            left = parse_excel_value(left_raw, rule, excel_epoch)
            right = parse_value(right_raw, rule)
            if not left.valid:
                continue
            if not right.valid:
                differences.append(_difference(DifferenceType.INVALID_VALUE, sheet, f"标准值无法按 {rule.type.value} 解析：{right.error}", cell=cell, excel_row=row_number, canonical_field=name, business_key=_business_key(key, sheet), standard_raw_value=_safe_value(right.raw, rule), rule_id=f"{name}.standard_parse", render_action="report_only"))
            elif not values_equal(left, right, rule):
                if rule.type.value == "fuzzy_string":
                    score = fuzz.ratio(str(left.normalized), str(right.normalized))
                    differences.append(_difference(
                        DifferenceType.VALUE_MISMATCH,
                        sheet,
                        f"模糊字段仅生成候选，未自动认定相同（相似度 {score:.1f}）",
                        cell=cell,
                        excel_row=row_number,
                        canonical_field=name,
                        business_key=_business_key(key, sheet),
                        excel_raw_value=_safe_value(left.raw, rule),
                        excel_normalized_value=_safe_value(left.normalized, rule),
                        standard_raw_value=_safe_value(right.raw, rule),
                        standard_normalized_value=_safe_value(right.normalized, rule),
                        rule_id=f"{name}.fuzzy_suggestion",
                        severity="warning",
                        render_action="mark_purple",
                    ))
                    continue
                can_fill = left.normalized is None and right.normalized is not None and sheet.actions.fill_empty_from_standard
                can_overwrite = sheet.actions.overwrite_mismatch
                repair = can_fill or can_overwrite
                if repair and not _excel_numeric_write_safe(right, rule):
                    _append_excel_write_precision_difference(differences, sheet, rule, right, row_number=row_number, cell=cell, key=key)
                    repair = False
                repair_rule = f"{name}.fill_empty_from_standard" if can_fill else f"{name}.overwrite_mismatch"
                difference = _difference(
                    DifferenceType.VALUE_MISMATCH,
                    sheet,
                    "字段值与标准数据不一致",
                    cell=cell,
                    excel_row=row_number,
                    canonical_field=name,
                    business_key=_business_key(key, sheet),
                    excel_raw_value=_safe_value(left.raw, rule),
                    excel_normalized_value=_safe_value(left.normalized, rule),
                    standard_raw_value=_safe_value(right.raw, rule),
                    standard_normalized_value=_safe_value(right.normalized, rule),
                    rule_id=f"{name}.{rule.compare.mode}",
                    render_action="set_cell_green" if repair else sheet.actions.mismatched_value,
                    repair_status="planned" if repair else "not_requested",
                )
                differences.append(difference)
                if repair:
                    repairs.append(RepairOperation("set_cell", sheet.id, sheet.name, repair_rule, difference.difference_id, cell=cell, canonical_field=name, value=_excel_write_value(right, rule)))
            elif left.normalized is None and rule.fill_static_default:
                default = parse_value(rule.static_default, rule)
                if not _excel_numeric_write_safe(default, rule):
                    _append_excel_write_precision_difference(differences, sheet, rule, default, row_number=row_number, cell=cell, key=key)
                    continue
                difference = _difference(
                    DifferenceType.NORMALIZED_MATCH,
                    sheet,
                    "已授权使用静态默认值填充空单元格",
                    cell=cell,
                    excel_row=row_number,
                    canonical_field=name,
                    business_key=_business_key(key, sheet),
                    standard_raw_value=_safe_value(rule.static_default, rule),
                    rule_id=f"{name}.fill_static_default",
                    severity="info",
                    render_action="set_cell_green",
                    repair_status="planned",
                )
                differences.append(difference)
                default_value = _excel_write_value(default, rule)
                repairs.append(RepairOperation("set_cell", sheet.id, sheet.name, difference.rule_id or "", difference.difference_id, cell=cell, canonical_field=name, value=default_value))
    if isinstance(standard_records, DiskBackedRecordMap):
        standard_records.close()
    return (
        join_backend,
        "disk_standard_records" if use_disk_records else "memory_standard_records",
        len(matched_keys),
        len(excel_records),
        formula_target_rows,
    )


def _validate_excel_records(sheet: SheetRule, snapshot: SheetSnapshot, columns: dict[str, int], records: dict[tuple[Any, ...], tuple[int, dict[str, Any]]], rules_by_name: dict[str, Any], differences: list[Difference], excel_epoch: Any) -> None:
    unique_values: dict[str, dict[Any, list[tuple[int, tuple[Any, ...]]]]] = defaultdict(lambda: defaultdict(list))
    cross_fields = {
        value
        for cross_rule in sheet.cross_field_rules
        for key, value in cross_rule.params.items()
        if key.endswith("_field") and isinstance(value, str)
    }
    for key, (row_number, record) in records.items():
        parsed_record: dict[str, ParsedValue] = {}
        for name, col_index in columns.items():
            rule = rules_by_name[name]
            raw = record.get(name)
            cell = f"{get_column_letter(col_index)}{row_number}"
            if isinstance(raw, str) and raw.startswith("="):
                if rule.compare.formula_mode == "formula":
                    parsed_record[name] = ParsedValue(raw, raw, True)
                    continue
                if rule.compare.formula_mode == "cached_value":
                    raw = snapshot.cached_values.get(cell)
                else:
                    differences.append(_difference(DifferenceType.UNSUPPORTED_FEATURE, sheet, "字段规则禁止公式，系统未执行公式", cell=cell, excel_row=row_number, canonical_field=name, business_key=_business_key(key, sheet), rule_id=f"{name}.formula_reject", render_action="mark_purple"))
                    continue
            if (raw is None or raw == "") and not rule.required and rule.validation.nullable and not rule.validation.unique:
                if name in cross_fields:
                    parsed_record[name] = ParsedValue(raw, None, True)
                continue
            parsed = parse_excel_value(raw, rule, excel_epoch)
            parsed_record[name] = parsed
            if not parsed.valid:
                differences.append(_difference(DifferenceType.INVALID_VALUE, sheet, f"Excel 值无法按 {rule.type.value} 解析：{parsed.error}", cell=cell, excel_row=row_number, canonical_field=name, business_key=_business_key(key, sheet), excel_raw_value=_safe_value(parsed.raw, rule), rule_id=f"{name}.parse", render_action=sheet.actions.invalid_value))
                continue
            validation_message = _validate(parsed, rule)
            if validation_message:
                differences.append(_difference(DifferenceType.VALIDATION_ERROR, sheet, validation_message, cell=cell, excel_row=row_number, canonical_field=name, business_key=_business_key(key, sheet), excel_raw_value=_safe_value(parsed.raw, rule), excel_normalized_value=_safe_value(parsed.normalized, rule), rule_id=f"{name}.validation", render_action=sheet.actions.invalid_value))
            if rule.validation.unique and parsed.normalized is not None:
                unique_values[name][parsed.normalized].append((row_number, key))
        _validate_cross_fields(sheet, columns, row_number, key, parsed_record, rules_by_name, differences)
    for name, values in unique_values.items():
        for _value, occurrences in values.items():
            if len(occurrences) <= 1:
                continue
            for row_number, key in occurrences:
                differences.append(_difference(DifferenceType.VALIDATION_ERROR, sheet, "字段值不唯一", cell=f"{get_column_letter(columns[name])}{row_number}", excel_row=row_number, canonical_field=name, business_key=_business_key(key, sheet), rule_id=f"{name}.unique", render_action=sheet.actions.invalid_value))


def _validate_cross_fields(sheet: SheetRule, columns: dict[str, int], row_number: int, key: tuple[Any, ...], parsed: dict[str, ParsedValue], rules_by_name: dict[str, Any], differences: list[Difference]) -> None:
    for rule in sheet.cross_field_rules:
        params = dict(rule.params)
        when = params.get("when_field")
        if rule.validator == "conditional_required" and when in rules_by_name:
            params["equals"] = parse_value(params.get("equals"), rules_by_name[when]).normalized
        outcome = run_validator(rule.validator, parsed, params)
        target, message = outcome if outcome is not None else (None, None)
        if target and message and target in columns:
            differences.append(_difference(DifferenceType.VALIDATION_ERROR, sheet, message, cell=f"{get_column_letter(columns[target])}{row_number}", excel_row=row_number, canonical_field=target, business_key=_business_key(key, sheet), rule_id=rule.rule_id, severity=rule.severity, render_action=sheet.actions.invalid_value))


def _key(record: dict[str, Any], sheet: SheetRule, rules_by_name: dict[str, Any], row_number: Any = None, excel_epoch: Any = None) -> tuple[tuple[Any, ...], bool]:
    if sheet.primary_key_mode == "row_number":
        try:
            parsed_row = parse_row_number(row_number)
        except (TypeError, ValueError):
            return tuple(), False
        return (("row_number", parsed_row),), True
    values: list[Any] = []
    for name in sheet.primary_key:
        parsed = (
            parse_excel_value(record.get(name), rules_by_name[name], excel_epoch)
            if excel_epoch is not None
            else parse_value(record.get(name), rules_by_name[name])
        )
        if not parsed.valid or parsed.normalized is None or parsed.normalized == "":
            return tuple(values), False
        values.append((rules_by_name[name].type.value, parsed.normalized))
    return tuple(values), True


def _business_key(key: tuple[Any, ...], sheet: SheetRule) -> dict[str, Any]:
    if sheet.primary_key_mode == "row_number":
        return {sheet.row_number_field: key[0][1]}
    rules_by_name = {column.name: column for column in sheet.columns}
    return {name: _safe_value(value[1], rules_by_name[name]) for name, value in zip(sheet.primary_key, key)}


def _safe_value(value: Any, rule: Any) -> Any:
    if not rule.sensitive or value is None:
        return value
    text = str(value)
    if len(text) <= 4:
        return "****"
    return f"{text[:2]}***{text[-2:]}"


def _safe_record(record: dict[str, Any], rules_by_name: dict[str, Any]) -> dict[str, Any]:
    return {name: _safe_value(value, rules_by_name[name]) if name in rules_by_name else value for name, value in record.items()}


def _excel_numeric_write_safe(parsed: ParsedValue, rule: Any) -> bool:
    """Return whether Excel can round-trip the normalized numeric value exactly.

    Excel stores numeric cells as IEEE-754 doubles and documents 15 significant
    decimal digits. Comparison can remain arbitrary-precision Decimal, but an
    authorized repair must not silently round or change a numeric value to text.
    """
    if rule.type.value not in {"integer", "decimal"} or not parsed.valid or parsed.normalized is None:
        return True
    # Formula-mode values are never written as numeric cells. Their safety is
    # governed by the separate trusted formula-template checks.
    if rule.compare.formula_mode == "formula" and is_formula_text(parsed.normalized):
        return True
    value = Decimal(str(parsed.normalized))
    if not value.is_finite():
        return False
    absolute = abs(value)
    if absolute != 0 and (
        absolute < _EXCEL_MIN_POSITIVE_NUMBER
        or absolute > _EXCEL_MAX_ABSOLUTE_NUMBER
    ):
        return False
    digits = list(value.as_tuple().digits)
    while len(digits) > 1 and digits[-1] == 0:
        digits.pop()
    if len(digits) > 15:
        return False
    try:
        as_float = float(value)
    except (OverflowError, ValueError):
        return False
    if not math.isfinite(as_float) or (value != 0 and as_float == 0):
        return False
    return Decimal(str(as_float)) == value


def _append_excel_write_precision_difference(
    differences: list[Difference] | SpillableSequence[Difference],
    sheet: SheetRule,
    rule: Any,
    parsed: ParsedValue,
    *,
    row_number: int | None = None,
    cell: str | None = None,
    key: tuple[Any, ...] | None = None,
) -> None:
    differences.append(_difference(
        DifferenceType.UNSUPPORTED_FEATURE,
        sheet,
        "标准数值超出 Excel 数值单元格可安全往返的 15 位有效数字或指数范围，禁止自动写回",
        cell=cell,
        excel_row=row_number,
        canonical_field=rule.name,
        business_key=_business_key(key, sheet) if key else None,
        standard_raw_value=_safe_value(parsed.raw, rule),
        standard_normalized_value=_safe_value(parsed.normalized, rule),
        rule_id=f"{rule.name}.excel_write_precision",
        render_action="report_only",
    ))


def _excel_write_value(parsed: ParsedValue, rule: Any) -> Any:
    """Return a deterministic, type-preserving value for the renderer.

    Reports retain ``parsed.raw`` separately. Repairs must use the same typed
    normalization that comparison used, otherwise aliases, percentages,
    booleans and timezone-aware datetimes can be written back as different
    semantics or as plain text.
    """
    if not parsed.valid or parsed.normalized is None:
        return None
    if rule.type.value == "decimal":
        return str(parsed.normalized)
    if rule.type.value in {"date", "datetime"}:
        return parsed.normalized.isoformat()
    if rule.type.value == "set":
        return rule.separator.join(str(value) for value in parsed.normalized)
    if rule.type.value == "json":
        return str(parsed.normalized)
    if rule.type.value in {"integer", "boolean", "enum"}:
        return parsed.normalized
    return parsed.raw


def _should_insert_missing(sheet: SheetRule, column: Any) -> bool:
    if column.missing_column_action is not None:
        return column.missing_column_action == "insert"
    configured = sheet.actions.missing_required_header if column.required else sheet.actions.missing_optional_header
    return configured == "insert_and_mark_green"


def _validate(value: ParsedValue, rule: Any) -> str | None:
    config = rule.validation
    if value.normalized is None:
        if rule.required or not config.nullable:
            return "必填字段为空"
        return None
    text = str(value.normalized)
    if config.min_length is not None and len(text) < config.min_length:
        return f"长度小于 {config.min_length}"
    if config.max_length is not None and len(text) > config.max_length:
        return f"长度大于 {config.max_length}"
    if config.regex and re.fullmatch(config.regex, text) is None:
        return "值不符合正则规则"
    if config.min is not None and Decimal(str(value.normalized)) < config.min:
        return f"值小于最小值 {config.min}"
    if config.max is not None and Decimal(str(value.normalized)) > config.max:
        return f"值大于最大值 {config.max}"
    return None
