from __future__ import annotations

import hashlib
import pickle
import re
import tempfile
import zipfile
from array import array
from datetime import datetime
from xml.parsers import expat
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable, Sequence, overload

from openpyxl import load_workbook
from openpyxl.utils.datetime import WINDOWS_EPOCH

from .models import RuleSet, SheetRule, normalize_header


UNSUPPORTED_PACKAGE_FEATURES = frozenset(
    {
        "external_links",
        "external_connections",
        "pivot_tables",
        "drawings",
        "embedded_objects",
        "activex_controls",
        "legacy_controls",
        "charts",
    }
)

UNSUPPORTED_SHEET_FEATURES = frozenset(
    {
        "protected_sheet",
        "merged_header",
        "complex_formula_references",
        "shared_formulas",
        "array_formulas",
        "formulas",
        "large_mode_hidden_row_filter_unavailable",
    }
)

AUTO_HEADER_SCAN_ROWS = 50


class WorkbookSafetyError(ValueError):
    pass


class SpilledRows(Sequence[tuple[int, list[Any]]]):
    """Disk-backed random-access row sequence used by large-file mode."""

    def __init__(self) -> None:
        self._file = tempfile.TemporaryFile(prefix="excel-auditor-rows-", suffix=".bin")
        self._offsets = array("Q")

    def append(self, item: tuple[int, list[Any]]) -> None:
        self._offsets.append(self._file.tell())
        pickle.dump(item, self._file, protocol=5)

    def __len__(self) -> int:
        return len(self._offsets)

    @overload
    def __getitem__(self, index: int) -> tuple[int, list[Any]]: ...
    @overload
    def __getitem__(self, index: slice) -> Iterable[tuple[int, list[Any]]]: ...

    def __getitem__(self, index: int | slice) -> tuple[int, list[Any]] | Iterable[tuple[int, list[Any]]]:
        if isinstance(index, slice):
            return (self[position] for position in range(*index.indices(len(self))))
        if index < 0:
            index += len(self)
        if not 0 <= index < len(self):
            raise IndexError(index)
        self._file.seek(self._offsets[index])
        return pickle.load(self._file)

    def close(self) -> None:
        self._file.close()


@dataclass
class SheetSnapshot:
    name: str
    max_row: int
    max_column: int
    rows: Sequence[tuple[int, list[Any]]]
    hidden_rows: set[int] = field(default_factory=set)
    risky_features: list[str] = field(default_factory=list)
    cached_values: dict[str, Any] = field(default_factory=dict)


@dataclass
class WorkbookSnapshot:
    path: Path
    sha256: str
    sheets: dict[str, SheetSnapshot]
    warnings: list[str] = field(default_factory=list)
    manual_review_reasons: list[str] = field(default_factory=list)
    large_mode: bool = False
    report_only: bool = False
    excel_epoch: datetime = WINDOWS_EPOCH

    def close(self) -> None:
        for sheet in self.sheets.values():
            close = getattr(sheet.rows, "close", None)
            if close is not None:
                close()


def locate_header_row(sheet: SheetRule, snapshot: SheetSnapshot) -> tuple[int, str | None]:
    """Resolve one header row for both safety inspection and comparison."""
    if not sheet.header.auto_detect:
        if sheet.header.row > len(snapshot.rows):
            return sheet.header.row, f"指定表头行 {sheet.header.row} 超出工作表范围"
        return sheet.header.row, None
    exact: dict[str, str] = {}
    for column in sheet.columns:
        for candidate in [column.name, column.title, *column.aliases]:
            exact[normalize_header(candidate)] = column.name
    candidates: list[tuple[int, int]] = []
    for row_number, values in snapshot.rows[:AUTO_HEADER_SCAN_ROWS]:
        matched = {exact[value] for value in map(normalize_header, values) if value in exact}
        if (sheet.primary_key_mode == "fields" and set(sheet.primary_key) <= matched) or (
            sheet.primary_key_mode == "row_number" and matched
        ):
            candidates.append((len(matched), row_number))
    if not candidates:
        return sheet.header.row, "未找到包含完整主键的候选表头行"
    best_score = max(score for score, _row in candidates)
    best_rows = [row for score, row in candidates if score == best_score]
    if len(best_rows) != 1:
        return best_rows[0], f"表头自动定位存在并列候选行：{best_rows}"
    return best_rows[0], None


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def inspect_workbook(path: Path, rules: RuleSet, max_size: int | None = None, max_in_memory_cells: int | None = None) -> WorkbookSnapshot:
    max_size = max_size if max_size is not None else rules.workbook.max_upload_mib * 1024 * 1024
    max_in_memory_cells = max_in_memory_cells if max_in_memory_cells is not None else rules.workbook.max_in_memory_cells
    if path.suffix.lower().lstrip(".") not in rules.workbook.allowed_extensions:
        raise WorkbookSafetyError("FILE_UNSUPPORTED_FORMAT")
    if path.stat().st_size > max_size:
        raise WorkbookSafetyError("FILE_LIMIT_EXCEEDED")
    if not zipfile.is_zipfile(path):
        raise WorkbookSafetyError("FILE_CORRUPTED")
    package_warnings: list[str] = []
    package_sheet_features: set[str] = set()
    with zipfile.ZipFile(path) as archive:
        infos = archive.infolist()
        if len(infos) > 10_000:
            raise WorkbookSafetyError("FILE_LIMIT_EXCEEDED: too many ZIP entries")
        if any(info.file_size > 256 * 1024 * 1024 for info in infos):
            raise WorkbookSafetyError("FILE_LIMIT_EXCEEDED: oversized ZIP entry")
        normalized_names = [info.filename.replace("\\", "/").casefold() for info in infos]
        if len(normalized_names) != len(set(normalized_names)):
            raise WorkbookSafetyError("FILE_CORRUPTED: duplicate ZIP entry")
        compressed = sum(max(info.compress_size, 1) for info in infos)
        uncompressed = sum(info.file_size for info in infos)
        if uncompressed > 512 * 1024 * 1024 or uncompressed / compressed > 100:
            raise WorkbookSafetyError("FILE_LIMIT_EXCEEDED: suspicious compression ratio")
        if any(".." in Path(info.filename).parts or info.filename.startswith(("/", "\\")) for info in infos):
            raise WorkbookSafetyError("FILE_CORRUPTED: unsafe ZIP entry")
        names = {info.filename.lower() for info in infos}
        has_vba = "xl/vbaproject.bin" in names
        if path.suffix.lower() == ".xlsm" and (not rules.workbook.preserve_macros or not has_vba):
            raise WorkbookSafetyError("FILE_UNSUPPORTED_FORMAT: macro preservation is not enabled or VBA project is missing")
        if has_vba:
            package_warnings.append("workbook: vba_project")
        package_features = {
            "external_links": any(name.startswith("xl/externallinks/") for name in names),
            "external_connections": "xl/connections.xml" in names,
            "pivot_tables": any(name.startswith("xl/pivottables/") for name in names),
            # Legacy VML Note drawings are the storage mechanism for ordinary
            # cell comments and are assessed separately by legacy_controls.
            # Only DrawingML worksheet drawing parts are unsafe here.
            "drawings": any(name.startswith("xl/drawings/") and name.endswith(".xml") for name in names),
            "embedded_objects": any(name.startswith("xl/embeddings/") for name in names),
            "activex_controls": any(name.startswith("xl/activex/") for name in names),
            "legacy_controls": any(name.endswith(".vml") and _vml_has_controls(archive, info) for name, info in ((item.filename.lower(), item) for item in infos)),
            "excel_tables": any(name.startswith("xl/tables/") for name in names),
            "charts": any(name.startswith("xl/charts/") for name in names),
            "calculation_chain": "xl/calcchain.xml" in names,
            "defined_names": any(
                info.filename.lower() == "xl/workbook.xml" and re.search(rb"<definedname(?:\s|>)", archive.read(info).lower()) is not None
                for info in infos
            ),
        }
        package_warnings.extend(f"workbook: {feature}" for feature, present in package_features.items() if present)
        for info in infos:
            if info.filename.lower().endswith((".xml", ".rels")):
                _validate_xml_part(archive, info)
            if info.filename.lower().startswith("xl/worksheets/") and info.filename.lower().endswith(".xml"):
                package_sheet_features.update(_detect_sheet_xml_features(archive, info))
    try:
        probe = load_workbook(path, read_only=True, data_only=False, keep_links=False, keep_vba=path.suffix.lower() == ".xlsm")
        large_mode = any(
            sheet.max_row is None
            or sheet.max_column is None
            or sheet.max_row * sheet.max_column > max_in_memory_cells
            for sheet in probe.worksheets
        )
        if large_mode and rules.workbook.large_file_action == "reject":
            probe.close()
            raise WorkbookSafetyError("FILE_LIMIT_EXCEEDED: workbook requires large-file report mode")
        if large_mode:
            book = probe
        else:
            probe.close()
            book = load_workbook(path, read_only=False, data_only=False, keep_links=False, keep_vba=path.suffix.lower() == ".xlsm")
    except Exception as exc:
        raise WorkbookSafetyError("FILE_CORRUPTED") from exc
    snapshots: dict[str, SheetSnapshot] = {}
    warnings: list[str] = list(package_warnings)
    manual_review_reasons: list[str] = [
        f"workbook: {feature}"
        for feature, present in package_features.items()
        if present and feature in UNSUPPORTED_PACKAGE_FEATURES
    ]
    cached_value_requests: dict[str, set[str]] = {}
    if len(book.worksheets) > rules.workbook.max_worksheets:
        book.close()
        raise WorkbookSafetyError("FILE_LIMIT_EXCEEDED: workbook has more than 50 worksheets")
    if rules.workbook.reject_protected and book.security.lockStructure:
        raise WorkbookSafetyError("WORKBOOK_PROTECTED")
    for sheet in book.worksheets:
        matching_rule = sheet_rule_for_name(rules, sheet.title)
        configured_data_start = matching_rule.data_region.start_row
        if configured_data_start is not None:
            preliminary_data_start = configured_data_start
        elif matching_rule.header.auto_detect:
            # Header discovery is bounded to the first 50 rows. Permit that
            # bounded prefix while reading, then enforce the exact limit once
            # the physical header row has been resolved below.
            preliminary_data_start = AUTO_HEADER_SCAN_ROWS + 1
        else:
            preliminary_data_start = matching_rule.header.row + 1
        row_limit = rules.workbook.max_rows_per_sheet + preliminary_data_start - 1
        max_row, max_column = sheet.max_row, sheet.max_column
        if (max_row is not None and max_row > row_limit) or (max_column is not None and max_column > rules.workbook.max_columns_per_sheet):
            raise WorkbookSafetyError("FILE_LIMIT_EXCEEDED")
        risky: list[str] = []
        if sheet.sheet_state != "visible":
            risky.append(f"sheet_state_{sheet.sheet_state}")
        if getattr(sheet, "freeze_panes", None):
            risky.append("freeze_panes")
        if not large_mode and any(dimension.hidden for dimension in sheet.column_dimensions.values()):
            risky.append("hidden_columns")
        if getattr(getattr(sheet, "protection", None), "sheet", False):
            if rules.workbook.reject_protected:
                raise WorkbookSafetyError("WORKBOOK_PROTECTED")
            risky.append("protected_sheet")
        if not large_mode and sheet.merged_cells.ranges:
            risky.append("merged_cells")
        if not large_mode and sheet.tables:
            risky.append("excel_tables")
        if not large_mode and sheet.data_validations.count:
            risky.append("data_validations")
        if not large_mode and sheet.auto_filter.ref:
            risky.append("auto_filter")
        if large_mode:
            risky.extend(sorted(package_sheet_features))
        has_comments = False
        has_formulas = False
        formula_cells: list[tuple[int, int]] = []
        has_external_hyperlinks = False
        rows: list[tuple[int, list[Any]]] | SpilledRows = SpilledRows() if large_mode else []
        row_iterator = sheet.iter_rows() if max_row is None or max_column is None else sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_column)
        observed_max_column = 0
        for row_index, row in enumerate(row_iterator, start=1):
            if row_index > row_limit:
                raise WorkbookSafetyError("FILE_LIMIT_EXCEEDED")
            values: list[Any] = []
            for cell in row:
                has_comments = has_comments or getattr(cell, "comment", None) is not None
                has_formulas = has_formulas or cell.data_type == "f"
                if cell.data_type == "f":
                    formula_cells.append((row_index, len(values) + 1))
                    if _complex_formula_reference(str(cell.value)) and "complex_formula_references" not in risky:
                        risky.append("complex_formula_references")
                hyperlink = getattr(cell, "hyperlink", None)
                has_external_hyperlinks = has_external_hyperlinks or bool(hyperlink and getattr(hyperlink, "target", None))
                values.append(cell.value)
            observed_max_column = max(observed_max_column, len(values))
            if observed_max_column > rules.workbook.max_columns_per_sheet:
                raise WorkbookSafetyError("FILE_LIMIT_EXCEEDED")
            rows.append((row_index, values))
        max_row = len(rows)
        max_column = observed_max_column if rows else 0
        inspection_snapshot = SheetSnapshot(sheet.title, max_row, max_column, rows, set(), risky)
        resolved_header_row, _header_problem = locate_header_row(matching_rule, inspection_snapshot)
        actual_data_start = configured_data_start or resolved_header_row + 1
        actual_row_limit = rules.workbook.max_rows_per_sheet + actual_data_start - 1
        if max_row > actual_row_limit:
            close_rows = getattr(rows, "close", None)
            if close_rows is not None:
                close_rows()
            for existing_snapshot in snapshots.values():
                close_existing_rows = getattr(existing_snapshot.rows, "close", None)
                if close_existing_rows is not None:
                    close_existing_rows()
            book.close()
            raise WorkbookSafetyError("FILE_LIMIT_EXCEEDED")
        if not large_mode and sheet.merged_cells.ranges and any(
            merged.min_row <= resolved_header_row <= merged.max_row
            for merged in sheet.merged_cells.ranges
        ):
            risky.append("merged_header")
        if has_comments:
            risky.append("existing_comments")
        if has_formulas:
            warnings.append(f"{sheet.title}: formulas_present_not_calculated")
            header_values = rows[resolved_header_row - 1][1] if len(rows) >= resolved_header_row else []
            rule_by_header: dict[str, Any] = {}

            for column_rule in matching_rule.columns:
                for candidate in [column_rule.name, column_rule.title, *column_rule.aliases]:
                    rule_by_header[normalize_header(candidate)] = column_rule
            for row_index, column_index in formula_cells:
                raw_header = header_values[column_index - 1] if column_index <= len(header_values) else None
                column_rule = rule_by_header.get(normalize_header(raw_header))
                if column_rule is None or column_rule.compare.formula_mode == "reject":
                    if "formulas" not in risky:
                        risky.append("formulas")
                elif column_rule.compare.formula_mode == "cached_value":
                    cached_value_requests.setdefault(sheet.title, set()).add(f"{_column_letter(column_index)}{row_index}")
        if has_external_hyperlinks:
            risky.append("external_hyperlinks")
        hidden_rows = set() if large_mode else {index for index, dimension in sheet.row_dimensions.items() if dimension.hidden}
        if large_mode and not matching_rule.data_region.include_hidden_rows:
            risky.append("large_mode_hidden_row_filter_unavailable")
        snapshots[sheet.title] = SheetSnapshot(sheet.title, max_row, max_column, rows, hidden_rows, risky)
        warnings.extend(f"{sheet.title}: {feature}" for feature in risky)
        manual_review_reasons.extend(
            f"{sheet.title}: {feature}"
            for feature in risky
            if feature in UNSUPPORTED_SHEET_FEATURES
        )
    excel_epoch = book.epoch
    book.close()
    if cached_value_requests:
        cached_book = load_workbook(path, read_only=True, data_only=True, keep_links=False, keep_vba=path.suffix.lower() == ".xlsm")
        for sheet_name, references in cached_value_requests.items():
            pending = set(references)
            cached_sheet = cached_book[sheet_name]
            for row in cached_sheet.iter_rows():
                for cell in row:
                    if cell.coordinate in pending:
                        snapshots[sheet_name].cached_values[cell.coordinate] = cell.value
                        pending.remove(cell.coordinate)
                if not pending:
                    break
            missing_cache = {reference for reference in references if snapshots[sheet_name].cached_values.get(reference) is None}
            if missing_cache:
                reason = f"{sheet_name}: formula_cached_value_missing"
                warnings.append(reason)
                manual_review_reasons.append(reason)
        cached_book.close()
    if large_mode:
        warnings.append("workbook: large_file_report_only")
    return WorkbookSnapshot(
        path,
        sha256_file(path),
        snapshots,
        warnings,
        manual_review_reasons,
        large_mode,
        large_mode,
        excel_epoch,
    )


def sheet_rule_for_name(rules: RuleSet, name: str) -> Any:
    return next((sheet for sheet in rules.sheets if name in {sheet.name, *sheet.aliases}), rules.sheets[0])


def _detect_sheet_xml_features(archive: zipfile.ZipFile, info: zipfile.ZipInfo) -> set[str]:
    markers = {
        b"<mergecells": "merged_cells",
        b"<datavalidations": "data_validations",
        b"<autofilter": "auto_filter",
        b"<sheetprotection": "protected_sheet",
        b"<conditionalformatting": "conditional_formatting",
        b"<f t=\"shared\"": "shared_formulas",
        b"<f t=\"array\"": "array_formulas",
        b"<hyperlink": "internal_or_external_hyperlinks",
        b"<sortstate": "sort_state",
        b"<filtercolumn": "filter_columns",
    }
    found: set[str] = set()
    tail = b""
    with archive.open(info) as handle:
        while chunk := handle.read(256 * 1024):
            data = (tail + chunk).lower()
            for marker, feature in markers.items():
                if marker in data:
                    found.add(feature)
            tail = data[-64:]
            if len(found) == len(markers):
                break
    return found


def _vml_has_controls(archive: zipfile.ZipFile, info: zipfile.ZipInfo) -> bool:
    shape_count = 0
    note_count = 0
    unsafe = False

    def start_element(name: str, attributes: dict[str, str]) -> None:
        nonlocal shape_count, note_count, unsafe
        local_name = name.rsplit(":", 1)[-1].casefold()
        if local_name == "shape":
            shape_count += 1
        elif local_name == "macro":
            unsafe = True
        elif local_name == "clientdata":
            object_type = next(
                (value for key, value in attributes.items() if key.rsplit(":", 1)[-1].casefold() == "objecttype"),
                None,
            )
            if object_type is not None and object_type.casefold() == "note":
                note_count += 1
            else:
                unsafe = True

    def reject_declaration(*_arguments: Any) -> None:
        raise ValueError("VML declarations are not safe to inspect")

    try:
        parser = expat.ParserCreate()
        parser.StartElementHandler = start_element
        parser.StartDoctypeDeclHandler = reject_declaration
        parser.EntityDeclHandler = reject_declaration
        parser.ExternalEntityRefHandler = lambda *_arguments: 0
        with archive.open(info) as handle:
            while chunk := handle.read(256 * 1024):
                parser.Parse(chunk, False)
            parser.Parse(b"", True)
        # Comment VML has one Note ClientData element per shape. Any other
        # shape would be destroyed by rebuilding the comment drawing.
        return unsafe or shape_count != note_count
    except (OSError, ValueError, expat.ExpatError):
        return True


def _complex_formula_reference(formula: str) -> bool:
    return bool(
        "[" in formula
        or "#" in formula
        or "{" in formula
        or re.search(r"(?:^|[^A-Z0-9_])\$?[A-Z]{1,3}:\$?[A-Z]{1,3}(?:$|[^A-Z0-9_])", formula, re.IGNORECASE)
        or re.search(r"(?:^|[^0-9])\$?\d+:\$?\d+(?:$|[^0-9])", formula)
        or re.search(r"[^!]+:[^!]+!", formula)
        or re.search(r"\b[A-Z_][A-Z0-9_.]*\[[^\]]+\]", formula, re.IGNORECASE)
    )


def _column_letter(index: int) -> str:
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _validate_xml_part(archive: zipfile.ZipFile, info: zipfile.ZipInfo, max_depth: int = 100) -> None:
    with archive.open(info) as handle:
        prefix = handle.read(min(info.file_size, 64 * 1024)).upper()
    if b"<!DOCTYPE" in prefix or b"<!ENTITY" in prefix:
        raise WorkbookSafetyError("FILE_CORRUPTED: DTD and entity declarations are forbidden")
    if info.file_size > 2 * 1024 * 1024:
        _scan_large_xml_depth(archive, info, max_depth)
        return
    depth = 0
    parser = expat.ParserCreate()
    def start_element(_name: str, _attributes: dict[str, str]) -> None:
        nonlocal depth
        depth += 1
        if depth > max_depth:
            raise WorkbookSafetyError("FILE_LIMIT_EXCEEDED: XML nesting depth exceeded")
    def end_element(_name: str) -> None:
        nonlocal depth
        depth -= 1
    parser.StartElementHandler = start_element
    parser.EndElementHandler = end_element
    parser.ExternalEntityRefHandler = lambda *_args: 0
    try:
        with archive.open(info) as handle:
            parser.ParseFile(handle)
    except WorkbookSafetyError:
        raise
    except expat.ExpatError as exc:
        raise WorkbookSafetyError(f"FILE_CORRUPTED: invalid XML part {info.filename}") from exc


def _scan_large_xml_depth(archive: zipfile.ZipFile, info: zipfile.ZipInfo, max_depth: int) -> None:
    depth = 0
    remainder = b""
    with archive.open(info) as handle:
        while chunk := handle.read(256 * 1024):
            data = remainder + chunk
            last_close = data.rfind(b">")
            if last_close < 0:
                if len(data) > 1024 * 1024:
                    raise WorkbookSafetyError("FILE_LIMIT_EXCEEDED: oversized XML token")
                remainder = data
                continue
            complete, remainder = data[: last_close + 1], data[last_close + 1 :]
            upper = complete.upper()
            if b"<!DOCTYPE" in upper or b"<!ENTITY" in upper:
                raise WorkbookSafetyError("FILE_CORRUPTED: DTD and entity declarations are forbidden")
            for match in re.finditer(rb"<\s*(/)?\s*([A-Za-z_][A-Za-z0-9_.:-]*)(?:\s[^<>]*?)?\s*(/)?\s*>", complete):
                closing, self_closing = match.group(1), match.group(3)
                if closing:
                    depth -= 1
                    if depth < 0:
                        raise WorkbookSafetyError(f"FILE_CORRUPTED: invalid XML part {info.filename}")
                elif not self_closing:
                    depth += 1
                    if depth > max_depth:
                        raise WorkbookSafetyError("FILE_LIMIT_EXCEEDED: XML nesting depth exceeded")
    if remainder.strip() or depth != 0:
        raise WorkbookSafetyError(f"FILE_CORRUPTED: invalid XML part {info.filename}")
